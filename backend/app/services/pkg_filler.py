"""PKG 自动填入服务

根据解析结果，从费率数据库查询匹配费率，自动填入 Excel。
"""
import shutil
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

from app.services.pkg_parser import (
    PkgParseResult,
    LaneInfo,
    _extract_dest_code,
)
from app.services.rate_db import AirRate, query_rate


@dataclass
class FillResult:
    """单条航线的填入结果"""
    row: int
    origin: str
    destination: str
    cost_type: str
    status: str           # "filled" / "no_rate" / "already_filled" / "skipped"
    confidence: float
    # 填入的值
    unit_price: float | None = None
    lead_time: str | None = None
    carrier_route: str | None = None
    remarks: str | None = None
    # 原始值（用于对比）
    original_price: float | None = None


@dataclass
class PkgFillSummary:
    """填入结果汇总"""
    input_file: str
    output_file: str
    total_lanes: int
    filled_count: int
    no_rate_count: int
    already_filled_count: int
    skipped_count: int
    results: list[FillResult] = field(default_factory=list)


def fill_pkg(
    parse_result: PkgParseResult,
    input_path: str | Path,
    output_path: str | Path,
    overwrite_existing: bool = False,
) -> PkgFillSummary:
    """自动填入 PKG Excel

    Args:
        parse_result: 解析结果
        input_path: 输入 Excel 路径
        output_path: 输出 Excel 路径
        overwrite_existing: 是否覆盖已有值

    Returns:
        PkgFillSummary 填入结果汇总
    """
    input_path = Path(input_path)
    output_path = Path(output_path)

    # 复制原始文件
    shutil.copy2(input_path, output_path)

    # 打开副本进行编辑
    wb = openpyxl.load_workbook(str(output_path))
    ws = wb.active

    results: list[FillResult] = []
    filled_count = 0
    no_rate_count = 0
    already_filled_count = 0
    skipped_count = 0

    for section in parse_result.sections:
        for lane in section.lanes:
            result = _fill_lane(ws, section, lane, overwrite_existing)
            results.append(result)

            if result.status == 'filled':
                filled_count += 1
            elif result.status == 'no_rate':
                no_rate_count += 1
            elif result.status == 'already_filled':
                already_filled_count += 1
            else:
                skipped_count += 1

    wb.save(str(output_path))
    wb.close()

    return PkgFillSummary(
        input_file=input_path.name,
        output_file=output_path.name,
        total_lanes=parse_result.total_lanes,
        filled_count=filled_count,
        no_rate_count=no_rate_count,
        already_filled_count=already_filled_count,
        skipped_count=skipped_count,
        results=results,
    )


def _fill_lane(
    ws,
    section,
    lane: LaneInfo,
    overwrite_existing: bool,
) -> FillResult:
    """填入单条航线"""
    dest_code = _extract_dest_code(lane.destination_clean)

    # 检查是否已有有效数据
    current_price = lane.unit_price
    has_data = current_price is not None and current_price != 0

    if has_data and not overwrite_existing:
        return FillResult(
            row=lane.row,
            origin=lane.origin,
            destination=lane.destination_clean,
            cost_type=lane.cost_type,
            status='already_filled',
            confidence=1.0,
            unit_price=current_price,
            lead_time=lane.lead_time,
            carrier_route=lane.carrier_route,
            remarks=lane.remarks,
            original_price=current_price,
        )

    # LOCAL_DELIVERY 行中 F/G 列是 "－"，需要跳过填入
    is_local_delivery = lane.cost_type == 'LOCAL_DELIVERY'

    # 查询费率
    rate = query_rate(
        origin_code=section.origin_code,
        destination_code=dest_code,
        currency=section.currency,
        cost_type=lane.cost_type,
    )

    if rate is None:
        return FillResult(
            row=lane.row,
            origin=lane.origin,
            destination=lane.destination_clean,
            cost_type=lane.cost_type,
            status='no_rate',
            confidence=0.0,
            original_price=current_price,
        )

    # 写入 Excel 单元格
    row = lane.row

    # E列: 単価
    if rate.unit_price is not None:
        ws.cell(row=row, column=5).value = rate.unit_price

    # F列: Lead Time（LOCAL_DELIVERY 保持 "－"）
    if is_local_delivery:
        ws.cell(row=row, column=6).value = '－'
    elif rate.lead_time:
        ws.cell(row=row, column=6).value = rate.lead_time

    # G列: キャリアとルート（LOCAL_DELIVERY 保持 "－"）
    if is_local_delivery:
        ws.cell(row=row, column=7).value = '－'
    elif rate.carrier_route:
        ws.cell(row=row, column=7).value = rate.carrier_route

    # H列: 備考
    if rate.remarks:
        ws.cell(row=row, column=8).value = rate.remarks

    return FillResult(
        row=lane.row,
        origin=lane.origin,
        destination=lane.destination_clean,
        cost_type=lane.cost_type,
        status='filled',
        confidence=rate.confidence,
        unit_price=rate.unit_price,
        lead_time=rate.lead_time if not is_local_delivery else '－',
        carrier_route=rate.carrier_route if not is_local_delivery else '－',
        remarks=rate.remarks,
        original_price=current_price,
    )


def fill_summary_to_dict(summary: PkgFillSummary) -> dict:
    """将填入结果转换为可 JSON 序列化的字典"""
    return {
        'input_file': summary.input_file,
        'output_file': summary.output_file,
        'total_lanes': summary.total_lanes,
        'filled_count': summary.filled_count,
        'no_rate_count': summary.no_rate_count,
        'already_filled_count': summary.already_filled_count,
        'skipped_count': summary.skipped_count,
        'results': [
            {
                'row': r.row,
                'origin': r.origin,
                'destination': r.destination,
                'cost_type': r.cost_type,
                'status': r.status,
                'confidence': r.confidence,
                'unit_price': r.unit_price,
                'lead_time': r.lead_time,
                'carrier_route': r.carrier_route,
                'remarks': r.remarks,
                'original_price': r.original_price,
            }
            for r in summary.results
        ],
    }
