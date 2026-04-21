"""PKG（入札包）Excel 解析服务

解析客户入札模板 Excel，识别：
- 发地段（Section）：按发地分组（日本/中国/台湾/オランダ/韓国）
- 航线（Lane）：每行一条航线（発地→着地）
- 需填字段：単価(E列)、Lead Time(F列)、主要キャリアとルート(G列)、備考(H列)
"""
import re
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl


@dataclass
class LaneInfo:
    """一条航线信息"""
    row: int                    # Excel 行号
    origin: str                 # 発地（如 "中国  (上海)"）
    destination: str            # 着地（如 "アメリカ  (アトランタ)\nAIR FREIGHT COST"）
    destination_clean: str      # 着地简称（如 "アメリカ (アトランタ)"）
    cost_type: str              # 费用类型: "AIR_FREIGHT" / "LOCAL_DELIVERY" / "OTHER"
    volume_desc: str            # 想定物量描述
    currency: str               # 币种（円/CNY/ユーロ/USD）
    # 当前值（可能是空/0）
    unit_price: float | int | None = None
    lead_time: str | None = None
    carrier_route: str | None = None
    remarks: str | None = None


@dataclass
class SectionInfo:
    """一个发地段"""
    header_row: int             # 表头行号
    origin: str                 # 発地名
    origin_code: str            # 発地代码（如 PVG, NRT）
    currency: str               # 币种
    currency_unit: str          # 币种列标题（如 "単価 (CNY/kg)"）
    lanes: list[LaneInfo] = field(default_factory=list)


@dataclass
class PkgParseResult:
    """PKG 解析结果"""
    filename: str
    sheet_name: str
    period: str                 # 入札期间（如 "1月"）
    total_sections: int
    total_lanes: int
    sections: list[SectionInfo] = field(default_factory=list)


# 币种提取正则
CURRENCY_PATTERN = re.compile(r'単価\s*[\(（]\s*(.+?)\s*[\)）]')

# 发地代码映射
ORIGIN_CODE_MAP = {
    '成田': 'NRT',
    '日本': 'NRT',
    '上海': 'PVG',
    '中国': 'PVG',
    'アムステルダム': 'AMS',
    'オランダ': 'AMS',
    '台北': 'TPE',
    '台湾': 'TPE',
    'インチョン': 'ICN',
    '韓国': 'ICN',
}

# 着地代码映射
DEST_CODE_MAP = {
    'アトランタ': 'ATL',
    'マイアミ': 'MIA',
    'アムステルダム': 'AMS',
    'サンパウロ': 'GRU',
    'シドニー': 'SYD',
    '台北': 'TPE',
    '上海': 'PVG',
    'アメリカ': 'US',
    'オランダ': 'NL',
    'ブラジル': 'BR',
    'オーストラリア': 'AU',
}


def _extract_origin_code(origin_text: str) -> str:
    """从発地文本提取代码"""
    for keyword, code in ORIGIN_CODE_MAP.items():
        if keyword in origin_text:
            return code
    return 'UNKNOWN'


def _extract_dest_code(dest_text: str) -> str:
    """从着地文本提取代码"""
    for keyword, code in DEST_CODE_MAP.items():
        if keyword in dest_text:
            return code
    return 'UNKNOWN'


def _clean_destination(dest_text: str) -> str:
    """清理着地文本，去掉 AIR FREIGHT COST / LOCAL DELIVERY COST 等"""
    if not dest_text:
        return ''
    # 去掉费用类型标识
    cleaned = re.sub(r'\n?AIR FREIGHT COST', '', dest_text)
    cleaned = re.sub(r'\n?LOCAL DELIVERY COST', '', cleaned)
    # 清理多余空白
    cleaned = re.sub(r'\s+', ' ', cleaned).strip()
    return cleaned


def _detect_cost_type(dest_text: str) -> str:
    """检测费用类型"""
    if not dest_text:
        return 'OTHER'
    if 'LOCAL DELIVERY COST' in dest_text:
        return 'LOCAL_DELIVERY'
    if 'AIR FREIGHT COST' in dest_text:
        return 'AIR_FREIGHT'
    return 'AIR_FREIGHT'  # 默认为运费


def _extract_currency(header_text: str) -> str:
    """从表头文本提取币种"""
    m = CURRENCY_PATTERN.search(header_text)
    if m:
        raw = m.group(1).strip()
        # 标准化币种
        if '円' in raw or 'JPY' in raw:
            return 'JPY'
        if 'CNY' in raw:
            return 'CNY'
        if 'ユーロ' in raw or 'EUR' in raw:
            return 'EUR'
        if 'USD' in raw:
            return 'USD'
        return raw
    return 'UNKNOWN'


def _cell_str(cell_value) -> str:
    """安全获取单元格字符串值"""
    if cell_value is None:
        return ''
    return str(cell_value).strip()


def _cell_num(cell_value) -> float | int | None:
    """安全获取单元格数值"""
    if cell_value is None:
        return None
    if isinstance(cell_value, (int, float)):
        return cell_value
    try:
        return float(str(cell_value).strip())
    except (ValueError, TypeError):
        return None


def parse_pkg(filepath: str | Path) -> PkgParseResult:
    """解析入札包 Excel 文件

    Args:
        filepath: Excel 文件路径

    Returns:
        PkgParseResult 包含解析出的所有段和航线
    """
    filepath = Path(filepath)
    wb = openpyxl.load_workbook(str(filepath), data_only=True)
    ws = wb.active
    sheet_name = ws.title

    # 获取入札期间（B1 单元格）
    period = _cell_str(ws['B1'].value) or '不明'

    sections: list[SectionInfo] = []
    current_section: SectionInfo | None = None

    for row_idx in range(1, ws.max_row + 1):
        b_val = _cell_str(ws.cell(row=row_idx, column=2).value)  # B列
        c_val = _cell_str(ws.cell(row=row_idx, column=3).value)  # C列
        e_val = _cell_str(ws.cell(row=row_idx, column=5).value)  # E列

        # 检测表头行：B列=発地 且 C列=着地
        if '発地' in b_val and '着地' in c_val:
            # 提取币种
            currency_unit = e_val
            currency = _extract_currency(currency_unit)

            current_section = SectionInfo(
                header_row=row_idx,
                origin='',
                origin_code='',
                currency=currency,
                currency_unit=currency_unit,
            )
            sections.append(current_section)
            continue

        # 检测数据行：当前在某个 section 内
        if current_section is not None:
            # B列有值 = 新的发地
            if b_val and '発地' not in b_val and '※' not in b_val:
                current_section.origin = b_val
                current_section.origin_code = _extract_origin_code(b_val)

            # C列有值 = 一条航线
            if c_val:
                dest_clean = _clean_destination(c_val)
                cost_type = _detect_cost_type(c_val)

                d_val = _cell_str(ws.cell(row=row_idx, column=4).value)  # D列 想定物量
                e_num = _cell_num(ws.cell(row=row_idx, column=5).value)  # E列 単価
                f_val = _cell_str(ws.cell(row=row_idx, column=6).value)  # F列 Lead Time
                g_val = _cell_str(ws.cell(row=row_idx, column=7).value)  # G列 キャリアとルート
                h_val = _cell_str(ws.cell(row=row_idx, column=8).value)  # H列 備考

                # 跳过记入例行
                if '記入例' in g_val or '記入例' in h_val:
                    # 仍然记录这行，但标记为示例
                    pass

                lane = LaneInfo(
                    row=row_idx,
                    origin=current_section.origin,
                    destination=c_val,
                    destination_clean=dest_clean,
                    cost_type=cost_type,
                    volume_desc=d_val,
                    currency=current_section.currency,
                    unit_price=e_num,
                    lead_time=f_val if f_val and f_val != '－' else None,
                    carrier_route=g_val if g_val and g_val != '－' else None,
                    remarks=h_val if h_val else None,
                )
                current_section.lanes.append(lane)

    wb.close()

    total_lanes = sum(len(s.lanes) for s in sections)
    return PkgParseResult(
        filename=filepath.name,
        sheet_name=sheet_name,
        period=period,
        total_sections=len(sections),
        total_lanes=total_lanes,
        sections=sections,
    )


def parse_result_to_dict(result: PkgParseResult) -> dict:
    """将解析结果转换为可 JSON 序列化的字典"""
    return {
        'filename': result.filename,
        'sheet_name': result.sheet_name,
        'period': result.period,
        'total_sections': result.total_sections,
        'total_lanes': result.total_lanes,
        'sections': [
            {
                'header_row': s.header_row,
                'origin': s.origin,
                'origin_code': s.origin_code,
                'currency': s.currency,
                'currency_unit': s.currency_unit,
                'lanes': [
                    {
                        'row': l.row,
                        'origin': l.origin,
                        'destination': l.destination_clean,
                        'cost_type': l.cost_type,
                        'volume_desc': l.volume_desc,
                        'currency': l.currency,
                        'unit_price': l.unit_price,
                        'lead_time': l.lead_time,
                        'carrier_route': l.carrier_route,
                        'remarks': l.remarks,
                    }
                    for l in s.lanes
                ],
            }
            for s in result.sections
        ],
    }
