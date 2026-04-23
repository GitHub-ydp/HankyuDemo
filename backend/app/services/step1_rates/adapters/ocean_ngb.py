from __future__ import annotations

from collections.abc import Iterable
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
import math
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.services.step1_rates.entities import ParsedRateBatch, ParsedRateRecord, Step1FileType


class OceanNgbAdapter:
    """Step1 Ocean-NGB adapter（HHE/NGB 宁波法人 FCL+LCL Lv.1/2/3 三档运价）。"""

    key = "ocean_ngb"
    file_type = Step1FileType.ocean_ngb
    priority = 30

    _RATE_SHEET = "Rate"
    _SKIP_SHEETS = frozenset({"sample", "Shipping line name"})
    _DEFAULT_ORIGIN_PORT = "NINGBO"
    _RATE_LEVEL_VALUES = frozenset({"Lv.1", "Lv.2", "Lv.3"})
    _FCL_MODE = "FCL"
    _LCL_MODE = "LCL"
    _DEFAULT_CURRENCY = "USD"

    # 1-based 列号常量（与业务需求 §3 表对齐）
    _COL_AGENT = 1
    _COL_LEVEL = 2
    _COL_VALID_FROM = 3
    _COL_VALID_TO = 4
    _COL_MODE = 5
    _COL_SHIPPING_LINE = 6
    _COL_ORIGIN_AREA = 7
    _COL_ORIGIN_COUNTRY = 8
    _COL_ORIGIN_PORT = 9
    _COL_POD_CODE = 10
    _COL_DEST_AREA = 11
    _COL_DEST_COUNTRY = 12
    _COL_PLACE_OF_DELIVERY_CODE = 13
    _COL_PLACE_OF_DELIVERY_FULL = 14
    _COL_SERVICE_CODE = 15
    _COL_CONTAINER_TYPE = 16
    _COL_FCL_CURRENCY = 17
    _COL_20GP = 18
    _COL_40GP = 19
    _COL_40HC = 20
    _COL_FAF_CCY = 21
    _COL_FAF_VALUE = 22
    _COL_YAS_CCY = 23
    _COL_YAS_VALUE = 24
    _COL_THC_CCY = 25
    _COL_THC_20 = 26
    _COL_THC_40 = 27
    _COL_DOC_CCY = 28
    _COL_DOC = 29
    _COL_SEAL_CCY = 30
    _COL_SEAL = 31
    _COL_ENS = 32
    _COL_LSF = 33
    _COL_ERS = 34
    _COL_ISPS = 35
    _COL_HWS = 36
    _COL_OTHER1_FCL = 37
    _COL_OTHER2_FCL = 38
    _COL_OTHER3_FCL = 39
    _COL_LCL_CURRENCY = 40
    _COL_LCL_PER_CBM = 41
    _COL_LCL_PER_TON = 42
    _COL_LCL_BAF_CCY = 43
    _COL_LCL_BAF = 44
    _COL_LCL_CAF_CCY = 45
    _COL_LCL_CAF = 46
    _COL_LCL_THC_CCY = 47
    _COL_LCL_THC = 48
    _COL_CFS_CCY = 49
    _COL_CFS = 50
    _COL_BL_CCY = 51
    _COL_BL = 52
    _COL_LCL_OTHER1 = 53
    _COL_LCL_OTHER2 = 54
    _COL_REMARKS = 55
    _MAX_COL = 55

    _FCL_RATE_COLS = (_COL_20GP, _COL_40GP, _COL_40HC)
    _LCL_RATE_COLS = (_COL_LCL_PER_CBM, _COL_LCL_PER_TON)
    _FCL_RATE_COL_NAMES = {_COL_20GP: "20GP", _COL_40GP: "40GP", _COL_40HC: "40HC"}
    _LCL_RATE_COL_NAMES = {_COL_LCL_PER_CBM: "LCL_PER_CBM", _COL_LCL_PER_TON: "LCL_PER_TON"}

    _WHITESPACE_RE = re.compile(r"\s+")
    _MULTI_NEWLINE_RE = re.compile(r"\r\n?|\n")
    _INLINE_WS_RE = re.compile(r"[ \t]+")

    # 表头校验关键字（小写包含）
    _HEADER_EXPECT = {
        _COL_AGENT: "agent",
        _COL_LEVEL: "p/s rate",
        _COL_MODE: "fcl/lcl",
        _COL_SHIPPING_LINE: "shipping line",
        _COL_ORIGIN_PORT: "origin port",
        _COL_PLACE_OF_DELIVERY_CODE: "place of",
        _COL_PLACE_OF_DELIVERY_FULL: "place of delivery",
    }

    def detect(self, path: Path, *, file_type_hint: Step1FileType | None = None) -> bool:
        if file_type_hint == self.file_type:
            return True
        normalized_name = path.name.lower()
        return "ocean-ngb" in normalized_name or "ocean ngb" in normalized_name or "ngb" in normalized_name

    def parse(self, path: Path, db: Session | None = None) -> ParsedRateBatch:
        workbook = load_workbook(path, data_only=True)
        warnings: list[str] = []
        records: list[ParsedRateRecord] = []
        sheet_summary: dict[str, Any] = {
            "sheet_name": self._RATE_SHEET,
            "total_rows": 0,
            "effective_from": None,
            "effective_to": None,
        }
        formula_fallback_count = 0

        if self._RATE_SHEET not in workbook.sheetnames:
            warnings.append("Rate sheet missing in NGB workbook; nothing to parse")
            return ParsedRateBatch(
                file_type=self.file_type,
                source_file=path.name,
                effective_from=None,
                effective_to=None,
                records=[],
                warnings=warnings,
                adapter_key=self.key,
                metadata={
                    "file_name": path.name,
                    "source_type": "excel",
                    "parser_version": "ocean_ngb_v1",
                    "ngb_origin_assumption": "default origin = NINGBO when I column is empty",
                    "formula_fallback_count": 0,
                    "sheets": [sheet_summary],
                    "record_kind_distribution": {"ocean_ngb_fcl": 0, "ocean_ngb_lcl": 0},
                },
            )

        ws = workbook[self._RATE_SHEET]
        sheet_records, sheet_warnings, sheet_fallback_count, date_pairs = self._parse_rate_sheet(
            ws, source_file=path.name
        )
        records.extend(sheet_records)
        warnings.extend(sheet_warnings)
        formula_fallback_count += sheet_fallback_count

        # 批次有效期：从 Lv.1 行 (C, D) 集合取
        effective_from, effective_to, range_warning = self._resolve_batch_range(date_pairs)
        if range_warning:
            warnings.append(range_warning)

        sheet_summary["total_rows"] = len(sheet_records)
        sheet_summary["effective_from"] = effective_from
        sheet_summary["effective_to"] = effective_to

        if not records:
            warnings.append("NGB workbook produced 0 records; check Rate sheet structure")

        kind_dist = {"ocean_ngb_fcl": 0, "ocean_ngb_lcl": 0}
        for r in records:
            if r.record_kind in kind_dist:
                kind_dist[r.record_kind] += 1

        return ParsedRateBatch(
            file_type=self.file_type,
            source_file=path.name,
            effective_from=effective_from,
            effective_to=effective_to,
            records=records,
            warnings=self._dedupe_warnings(warnings),
            adapter_key=self.key,
            metadata={
                "file_name": path.name,
                "source_type": "excel",
                "parser_version": "ocean_ngb_v1",
                "ngb_origin_assumption": "default origin = NINGBO when I column is empty",
                "formula_fallback_count": formula_fallback_count,
                "sheets": [sheet_summary],
                "record_kind_distribution": kind_dist,
            },
        )

    def _parse_rate_sheet(
        self,
        ws,
        *,
        source_file: str,
    ) -> tuple[list[ParsedRateRecord], list[str], int, list[tuple[date | None, date | None]]]:
        warnings: list[str] = []
        records: list[ParsedRateRecord] = []
        formula_fallback_count = 0
        date_pairs: list[tuple[date | None, date | None]] = []

        # 表头校验（不 fail-fast）
        for col, expected_kw in self._HEADER_EXPECT.items():
            raw = ws.cell(1, col).value
            text = self._normalize_text(raw) or ""
            if expected_kw not in text.lower():
                warnings.append(
                    f"NGB Rate header mismatch at column {col}: expected '{expected_kw}', got '{raw}'"
                )

        last_lv1_rates: dict[int, Decimal | None] | None = None
        consecutive_empty = 0
        max_row = ws.max_row or 0

        for row_index in range(2, max_row + 1):
            row = self._read_row_cells(ws, row_index)
            if self._is_empty_row(row):
                consecutive_empty += 1
                if consecutive_empty >= 3:
                    break
                continue
            consecutive_empty = 0

            record, row_warnings, fallback_hits, lv1_rates_seen = self._build_ngb_record(
                row,
                row_index,
                source_file=source_file,
                last_lv1_rates=last_lv1_rates,
            )
            warnings.extend(row_warnings)
            formula_fallback_count += fallback_hits
            if lv1_rates_seen is not None:
                last_lv1_rates = lv1_rates_seen
            if record is None:
                continue
            records.append(record)
            if record.rate_level == "Lv.1":
                date_pairs.append((record.valid_from, record.valid_to))

        return records, warnings, formula_fallback_count, date_pairs

    def _read_row_cells(self, ws, row_index: int) -> list[Any]:
        return [ws.cell(row=row_index, column=c).value for c in range(1, self._MAX_COL + 1)]

    def _is_empty_row(self, row: Iterable[Any]) -> bool:
        return all(value is None or (isinstance(value, str) and value.strip() == "") for value in row)

    def _build_ngb_record(
        self,
        row: list[Any],
        row_index: int,
        *,
        source_file: str,
        last_lv1_rates: dict[int, Decimal | None] | None,
    ) -> tuple[ParsedRateRecord | None, list[str], int, dict[int, Decimal | None] | None]:
        warnings: list[str] = []
        fallback_hits = 0

        agent = self._normalize_text(row[self._COL_AGENT - 1])
        level = self._normalize_text(row[self._COL_LEVEL - 1])
        if level not in self._RATE_LEVEL_VALUES:
            warnings.append(f"NGB row {row_index}: unrecognized rate_level '{level}', skipped")
            return None, warnings, 0, None

        mode = self._normalize_text(row[self._COL_MODE - 1])
        if mode not in {self._FCL_MODE, self._LCL_MODE}:
            warnings.append(f"NGB row {row_index}: unrecognized mode '{mode}', skipped")
            return None, warnings, 0, None

        valid_from = self._to_date(row[self._COL_VALID_FROM - 1])
        valid_to = self._to_date(row[self._COL_VALID_TO - 1])
        if valid_from is None or valid_to is None:
            warnings.append(
                f"NGB row {row_index}: cannot parse valid_from/valid_to from C/D, "
                f"raw=({row[self._COL_VALID_FROM - 1]!r}, {row[self._COL_VALID_TO - 1]!r})"
            )

        shipping_line = self._normalize_text(row[self._COL_SHIPPING_LINE - 1])

        origin_raw = self._normalize_text(row[self._COL_ORIGIN_PORT - 1])
        if origin_raw:
            origin_port_name = origin_raw
            origin_source = "raw"
        else:
            origin_port_name = self._DEFAULT_ORIGIN_PORT
            origin_source = "default_NINGBO"

        destination_port_code = self._normalize_text(row[self._COL_PLACE_OF_DELIVERY_CODE - 1])
        destination_port_name = self._normalize_text(row[self._COL_PLACE_OF_DELIVERY_FULL - 1])
        pod_code = self._normalize_text(row[self._COL_POD_CODE - 1])

        # 主运费分流
        container_20gp = container_40gp = container_40hq = None
        freight_per_cbm = freight_per_ton = None
        currency: str | None
        formula_fallback_columns: list[int] = []
        new_lv1_rates: dict[int, Decimal | None] | None = None

        if mode == self._FCL_MODE:
            currency = self._normalize_text(row[self._COL_FCL_CURRENCY - 1]) or self._DEFAULT_CURRENCY
            kind = "ocean_ngb_fcl"
            (
                container_20gp,
                container_40gp,
                container_40hq,
                rate_warnings,
                rate_fallbacks,
                lv1_seen,
            ) = self._extract_main_rates(
                row=row,
                row_index=row_index,
                level=level,
                cols=self._FCL_RATE_COLS,
                col_names=self._FCL_RATE_COL_NAMES,
                last_lv1_rates=last_lv1_rates,
            )
            warnings.extend(rate_warnings)
            fallback_hits += len(rate_fallbacks)
            formula_fallback_columns = rate_fallbacks
            new_lv1_rates = lv1_seen
        else:
            currency = self._normalize_text(row[self._COL_LCL_CURRENCY - 1]) or self._DEFAULT_CURRENCY
            kind = "ocean_ngb_lcl"
            (
                freight_per_cbm,
                freight_per_ton,
                _unused,
                rate_warnings,
                rate_fallbacks,
                lv1_seen,
            ) = self._extract_main_rates(
                row=row,
                row_index=row_index,
                level=level,
                cols=self._LCL_RATE_COLS + (None,),  # 占位让方法保持 3 槽
                col_names=self._LCL_RATE_COL_NAMES,
                last_lv1_rates=last_lv1_rates,
            )
            warnings.extend(rate_warnings)
            fallback_hits += len(rate_fallbacks)
            formula_fallback_columns = rate_fallbacks
            new_lv1_rates = lv1_seen

        remarks = self._extract_text_keep_newline(row[self._COL_REMARKS - 1])

        # extras（通用）
        extras: dict[str, Any] = {
            "sheet_name": self._RATE_SHEET,
            "row_index": row_index,
            "agent": agent,
            "pod_code": pod_code,
            "place_of_delivery_code": destination_port_code,
            "origin_area": row[self._COL_ORIGIN_AREA - 1],
            "origin_country": self._normalize_text(row[self._COL_ORIGIN_COUNTRY - 1]),
            "dest_area": row[self._COL_DEST_AREA - 1],
            "dest_country": self._normalize_text(row[self._COL_DEST_COUNTRY - 1]),
            "container_type": self._normalize_text(row[self._COL_CONTAINER_TYPE - 1]),
            "mode": mode,
            "origin_source": origin_source,
        }

        # 附加费 extras（按 mode 分流）
        if mode == self._FCL_MODE:
            extras.update(self._collect_fcl_surcharges(row))
        else:
            extras.update(self._collect_lcl_surcharges(row))

        if formula_fallback_columns:
            extras["formula_fallback_columns"] = formula_fallback_columns
            extras["formula_fallback_note"] = (
                "main rate computed by ROUNDUP(lv1 * factor, -1) due to missing data_only cache"
            )

        # column_index_map（业务需求 §11.1）
        column_index_map = self._build_column_index_map(
            row=row, row_index=row_index, level=level, mode=mode
        )
        extras["column_index_map"] = column_index_map

        record = ParsedRateRecord(
            record_kind=kind,
            carrier_name=shipping_line,
            origin_port_name=origin_port_name,
            destination_port_name=destination_port_name,
            rate_level=level,
            container_20gp=container_20gp,
            container_40gp=container_40gp,
            container_40hq=container_40hq,
            freight_per_cbm=freight_per_cbm,
            freight_per_ton=freight_per_ton,
            currency=currency or self._DEFAULT_CURRENCY,
            valid_from=valid_from,
            valid_to=valid_to,
            remarks=remarks,
            source_type="excel",
            source_file=source_file,
            extras=extras,
        )
        return record, warnings, fallback_hits, new_lv1_rates

    def _extract_main_rates(
        self,
        *,
        row: list[Any],
        row_index: int,
        level: str,
        cols: tuple[int | None, ...],
        col_names: dict[int, str],
        last_lv1_rates: dict[int, Decimal | None] | None,
    ) -> tuple[
        Decimal | None,
        Decimal | None,
        Decimal | None,
        list[str],
        list[int],
        dict[int, Decimal | None] | None,
    ]:
        """读取主运费三槽（FCL: 20/40/40HC; LCL: per_cbm/per_ton/None）。

        Lv.1：直接取 data_only 数值，并把读到的值返回供后续 Lv.2/Lv.3 fallback 使用。
        Lv.2/Lv.3：data_only 优先；为 None 则按 ROUNDUP(lv1 × 1.1 / 1.2, -1) 兜底，并打 W-N08。
        """
        warnings: list[str] = []
        results: list[Decimal | None] = [None, None, None]
        fallback_cols: list[int] = []
        lv1_rates_seen: dict[int, Decimal | None] | None = None

        if level == "Lv.1":
            lv1_rates_seen = {}
            for slot, col in enumerate(cols):
                if col is None:
                    continue
                value = self._to_decimal_or_none(row[col - 1])
                results[slot] = value
                lv1_rates_seen[col] = value
                if value is None:
                    warnings.append(
                        f"NGB row {row_index}: Lv.1 main rate is empty for column {col_names.get(col, col)}, "
                        "cannot seed Lv.2/Lv.3 fallback"
                    )
        else:
            for slot, col in enumerate(cols):
                if col is None:
                    continue
                value = self._to_decimal_or_none(row[col - 1])
                if value is None:
                    lv1_value = (last_lv1_rates or {}).get(col)
                    fallback = self._compute_lv_fallback(lv1_value, level)
                    if fallback is not None:
                        warnings.append(
                            f"NGB row {row_index}: {level} {col_names.get(col, col)} "
                            f"formula not cached, fallback computed = {fallback}"
                        )
                        results[slot] = fallback
                        fallback_cols.append(col)
                    else:
                        results[slot] = None
                else:
                    results[slot] = value

        return results[0], results[1], results[2], warnings, fallback_cols, lv1_rates_seen

    def _compute_lv_fallback(self, lv1_value: Decimal | None, level: str) -> Decimal | None:
        if lv1_value is None:
            return None
        factor = Decimal("1.1") if level == "Lv.2" else Decimal("1.2") if level == "Lv.3" else None
        if factor is None:
            return None
        # ROUNDUP(value, -1) → 向上取整到 10 的倍数
        scaled = lv1_value * factor
        rounded_units = math.ceil(scaled / Decimal("10"))
        return Decimal(rounded_units) * Decimal("10")

    def _collect_fcl_surcharges(self, row: list[Any]) -> dict[str, Any]:
        return {
            "faf_currency": self._normalize_text(row[self._COL_FAF_CCY - 1]),
            "faf_value_raw": self._normalize_text(row[self._COL_FAF_VALUE - 1]),
            "yas_caf_currency": self._normalize_text(row[self._COL_YAS_CCY - 1]),
            "yas_caf_value_raw": self._normalize_text(row[self._COL_YAS_VALUE - 1]),
            "thc_currency": self._normalize_text(row[self._COL_THC_CCY - 1]),
            "thc_20": row[self._COL_THC_20 - 1],
            "thc_40": row[self._COL_THC_40 - 1],
            "doc_currency": self._normalize_text(row[self._COL_DOC_CCY - 1]),
            "doc_value": row[self._COL_DOC - 1],
            "seal_currency": self._normalize_text(row[self._COL_SEAL_CCY - 1]),
            "seal_value": row[self._COL_SEAL - 1],
            "ens_raw": self._normalize_text(row[self._COL_ENS - 1]),
            "lsf_raw": self._normalize_text(row[self._COL_LSF - 1]),
            "ers_raw": self._normalize_text(row[self._COL_ERS - 1]),
            "isps_raw": self._extract_text_keep_newline(row[self._COL_ISPS - 1]),
            "hws_raw": self._extract_text_keep_newline(row[self._COL_HWS - 1]),
            "other1_raw": self._extract_text_keep_newline(row[self._COL_OTHER1_FCL - 1]),
            "other2_raw": self._extract_text_keep_newline(row[self._COL_OTHER2_FCL - 1]),
            "other3_raw": self._extract_text_keep_newline(row[self._COL_OTHER3_FCL - 1]),
        }

    def _collect_lcl_surcharges(self, row: list[Any]) -> dict[str, Any]:
        return {
            "lcl_baf_currency": self._normalize_text(row[self._COL_LCL_BAF_CCY - 1]),
            "lcl_baf_raw": self._normalize_text(row[self._COL_LCL_BAF - 1]),
            "lcl_caf_currency": self._normalize_text(row[self._COL_LCL_CAF_CCY - 1]),
            "lcl_caf_value": row[self._COL_LCL_CAF - 1],
            "lcl_thc_currency": self._normalize_text(row[self._COL_LCL_THC_CCY - 1]),
            "lcl_thc_raw": self._normalize_text(row[self._COL_LCL_THC - 1]),
            "cfs_currency": self._normalize_text(row[self._COL_CFS_CCY - 1]),
            "cfs_value": row[self._COL_CFS - 1],
            "bl_currency": self._normalize_text(row[self._COL_BL_CCY - 1]),
            "bl_value": row[self._COL_BL - 1],
            "lcl_other1_raw": self._extract_text_keep_newline(row[self._COL_LCL_OTHER1 - 1]),
            "lcl_other2_raw": self._extract_text_keep_newline(row[self._COL_LCL_OTHER2 - 1]),
        }

    def _build_column_index_map(
        self,
        *,
        row: list[Any],
        row_index: int,
        level: str,
        mode: str,
    ) -> dict[int, Any]:
        """业务需求 §11.1 + §10 AC-NGB-20。

        - Lv.1 行：把所有非空 cell 写入 map（含 R/S/T/AO/AP 主运费数值）
        - Lv.2 / Lv.3 行：返回空 dict，让 Writer 端模板原 ROUNDUP / =A2 等公式自动联动
        """
        if level != "Lv.1":
            return {}
        column_map: dict[int, Any] = {}
        for col in range(1, self._MAX_COL + 1):
            value = row[col - 1]
            if value is None:
                continue
            if isinstance(value, str) and value == "":
                continue
            column_map[col] = value
        return column_map

    def _resolve_batch_range(
        self, date_pairs: list[tuple[date | None, date | None]]
    ) -> tuple[date | None, date | None, str | None]:
        valid_pairs = [(f, t) for f, t in date_pairs if f is not None and t is not None]
        if not valid_pairs:
            return None, None, None
        unique = set(valid_pairs)
        if len(unique) == 1:
            f, t = next(iter(unique))
            return f, t, None
        froms = [f for f, _ in valid_pairs]
        tos = [t for _, t in valid_pairs]
        warning = (
            f"NGB workbook has {len(unique)} distinct (effective_from, effective_to) tuples; "
            "batch range fallback to min(from) / max(to)"
        )
        return min(froms), max(tos), warning

    def _to_decimal_or_none(self, value: Any) -> Decimal | None:
        if value is None:
            return None
        if isinstance(value, bool):
            return None
        if isinstance(value, Decimal):
            return value
        if isinstance(value, (int, float)):
            try:
                return Decimal(str(value))
            except (InvalidOperation, ValueError):
                return None
        if isinstance(value, str):
            stripped = value.strip()
            if not stripped:
                return None
            try:
                return Decimal(stripped)
            except (InvalidOperation, ValueError):
                return None
        return None

    def _normalize_text(self, value: Any) -> str | None:
        if value is None:
            return None
        text = self._WHITESPACE_RE.sub(" ", str(value)).strip()
        return text or None

    def _extract_text_keep_newline(self, value: Any) -> str | None:
        if value is None:
            return None
        text = str(value)
        # 折叠行内空白（不动换行）
        lines = self._MULTI_NEWLINE_RE.split(text)
        cleaned_lines = [self._INLINE_WS_RE.sub(" ", line).strip() for line in lines]
        result = "\n".join(cleaned_lines).strip()
        return result or None

    def _to_date(self, value: Any) -> date | None:
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        return None

    def _dedupe_warnings(self, warnings: Iterable[str]) -> list[str]:
        deduped: list[str] = []
        seen: set[str] = set()
        for warning in warnings:
            if warning in seen:
                continue
            deduped.append(warning)
            seen.add(warning)
        return deduped
