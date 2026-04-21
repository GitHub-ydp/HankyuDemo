from __future__ import annotations

from collections.abc import Iterable
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.services.rate_parser import _resolve_port, _safe_decimal
from app.services.step1_rates.entities import ParsedRateBatch, ParsedRateRecord, Step1FileType


class OceanAdapter:
    """Step1 Ocean parser for the current Sea Net Rate workbook."""

    key = "ocean"
    file_type = Step1FileType.ocean
    priority = 20

    _JP_SHEET = "JP N RATE FCL & LCL"
    _OTHER_FCL_SHEET = "FCL N RATE OF OTHER PORTS"
    _LCL_SHEET = "LCL N RATE"
    _WHITESPACE_RE = re.compile(r"\s+")
    _KNOWN_NON_NUMERIC_MARKERS = (
        "at cost",
        "collect",
        "included",
        "incl.",
        "subject to",
        "dest charge",
        "destination surcharge",
        "destination baf",
        "destination charges",
    )
    _LCL_FREIGHT_PLACEHOLDERS = frozenset({"-", "mbl cc"})

    def detect(self, path: Path, *, file_type_hint: Step1FileType | None = None) -> bool:
        if file_type_hint == self.file_type:
            return True
        normalized_name = path.name.lower()
        return "ocean" in normalized_name and "ngb" not in normalized_name

    def parse(self, path: Path, db: Session | None = None) -> ParsedRateBatch:
        workbook = load_workbook(path, data_only=True)
        warnings: list[str] = []
        records: list[ParsedRateRecord] = []
        sheet_summaries: list[dict[str, Any]] = []

        jp_records, jp_meta, jp_warnings = self._parse_fcl_sheet(
            workbook[self._JP_SHEET],
            db,
            surcharge_mode="jp",
            source_file=path.name,
        )
        records.extend(jp_records)
        warnings.extend(jp_warnings)
        sheet_summaries.append(jp_meta)

        other_records, other_meta, other_warnings = self._parse_fcl_sheet(
            workbook[self._OTHER_FCL_SHEET],
            db,
            surcharge_mode="other",
            source_file=path.name,
        )
        records.extend(other_records)
        warnings.extend(other_warnings)
        sheet_summaries.append(other_meta)

        lcl_records, lcl_meta, lcl_warnings = self._parse_lcl_sheet(
            workbook[self._LCL_SHEET],
            db,
            source_file=path.name,
        )
        records.extend(lcl_records)
        warnings.extend(lcl_warnings)
        sheet_summaries.append(lcl_meta)

        effective_ranges = {
            (
                meta.get("effective_from").isoformat() if meta.get("effective_from") else None,
                meta.get("effective_to").isoformat() if meta.get("effective_to") else None,
            )
            for meta in sheet_summaries
        }
        if len(effective_ranges) > 1:
            warnings.append(
                "Ocean workbook sheet effective date ranges differ; batch-level dates should be reviewed."
            )

        batch_effective_from = sheet_summaries[0].get("effective_from")
        batch_effective_to = sheet_summaries[0].get("effective_to")

        return ParsedRateBatch(
            file_type=self.file_type,
            source_file=path.name,
            effective_from=batch_effective_from,
            effective_to=batch_effective_to,
            records=records,
            warnings=self._dedupe_warnings(warnings),
            adapter_key=self.key,
            metadata={
                "file_name": path.name,
                "source_type": "excel",
                "carrier_code": "OCEAN_STEP1",
                "sheets": [
                    {
                        "sheet_name": meta["sheet_name"],
                        "total_rows": meta["total_rows"],
                        "effective_from": meta["effective_from"],
                        "effective_to": meta["effective_to"],
                    }
                    for meta in sheet_summaries
                ],
            },
        )

    def _parse_fcl_sheet(
        self,
        worksheet,
        db: Session | None,
        *,
        surcharge_mode: str,
        source_file: str,
    ) -> tuple[list[ParsedRateRecord], dict[str, Any], list[str]]:
        effective_from, effective_to = self._read_effective_range(worksheet)
        warnings: list[str] = []
        records: list[ParsedRateRecord] = []
        pending: dict[str, Any] | None = None
        current_destination: str | None = None
        current_layout: dict[str, int | None] | None = None
        current_origin_name = self._read_origin_name(worksheet)
        current_origin_port = self._resolve_port_ref(current_origin_name, db)

        for row_index in range(1, worksheet.max_row + 1):
            row = [worksheet.cell(row=row_index, column=column).value for column in range(1, 18)]
            if self._is_empty_row(row):
                continue

            section_origin = self._read_origin_name_from_row(row[0])
            if section_origin:
                if pending is not None:
                    self._append_pending_record(
                        records=records,
                        warnings=warnings,
                        pending=pending,
                        sheet_name=worksheet.title,
                    )
                    pending = None
                current_origin_name = section_origin
                current_origin_port = self._resolve_port_ref(section_origin, db)
                current_destination = None
                continue

            layout = self._build_fcl_column_map(row)
            if layout is not None:
                if pending is not None:
                    self._append_pending_record(
                        records=records,
                        warnings=warnings,
                        pending=pending,
                        sheet_name=worksheet.title,
                    )
                    pending = None
                current_layout = layout
                current_destination = None
                continue

            if current_layout is None:
                continue

            destination_cell = self._get_layout_value(row, current_layout, "destination")
            shipping_line_cell = self._get_layout_value(row, current_layout, "shipping_line")
            container_type = self._normalize_container_type(row[2])

            if container_type is None:
                continue

            destination_text = self._normalize_text(destination_cell)
            shipping_line_text = self._normalize_text(shipping_line_cell)
            destination = destination_text or current_destination
            if destination_text:
                current_destination = destination_text

            if container_type == "20ft":
                shipping_line = shipping_line_text
            else:
                shipping_line = shipping_line_text or (pending["carrier_name"] if pending else None)

            if not destination:
                warnings.append(f"{worksheet.title} row {row_index}: missing destination after merge expansion.")
                continue

            destination_port = self._resolve_port_ref(destination, db)
            base_payload = self._build_fcl_row_payload(
                row=row,
                row_index=row_index,
                sheet_name=worksheet.title,
                container_type=container_type,
                destination=destination,
                shipping_line=shipping_line,
                origin_port=current_origin_port,
                origin_name=current_origin_name,
                destination_port=destination_port,
                effective_from=effective_from,
                effective_to=effective_to,
                surcharge_mode=surcharge_mode,
                layout=current_layout,
                warnings=warnings,
                source_file=source_file,
            )

            if container_type == "20ft":
                if pending is not None:
                    self._append_pending_record(
                        records=records,
                        warnings=warnings,
                        pending=pending,
                        sheet_name=worksheet.title,
                    )
                pending = base_payload
                continue

            if pending is None:
                warnings.append(
                    f"{worksheet.title} row {row_index}: encountered {container_type.upper()} before 20FT pair."
                )
                self._append_orphan_40_record(
                    records=records,
                    warnings=warnings,
                    payload=base_payload,
                    sheet_name=worksheet.title,
                )
                continue

            if not self._same_pair_key(pending, base_payload):
                self._append_pending_record(
                    records=records,
                    warnings=warnings,
                    pending=pending,
                    sheet_name=worksheet.title,
                )
                warnings.append(
                    f"{worksheet.title} row {row_index}: FCL pair key changed before 40FT/40HQ merge completed."
                )
                self._append_orphan_40_record(
                    records=records,
                    warnings=warnings,
                    payload=base_payload,
                    sheet_name=worksheet.title,
                )
                pending = None
                continue

            self._merge_40_payload(pending, base_payload, container_type)

        if pending is not None:
            self._append_pending_record(
                records=records,
                warnings=warnings,
                pending=pending,
                sheet_name=worksheet.title,
            )

        meta = {
            "sheet_name": worksheet.title,
            "total_rows": len(records),
            "effective_from": effective_from,
            "effective_to": effective_to,
        }
        return records, meta, warnings

    def _parse_lcl_sheet(
        self,
        worksheet,
        db: Session | None,
        *,
        source_file: str,
    ) -> tuple[list[ParsedRateRecord], dict[str, Any], list[str]]:
        effective_from, effective_to = self._read_effective_range(worksheet)
        origin_name = self._read_origin_name(worksheet)
        origin_port = self._resolve_port_ref(origin_name, db)
        warnings: list[str] = []
        records: list[ParsedRateRecord] = []

        for row_index in range(10, worksheet.max_row + 1):
            row = [worksheet.cell(row=row_index, column=column).value for column in range(1, 13)]
            destination = self._normalize_text(row[0])
            if not destination:
                continue

            destination_port = self._resolve_port_ref(destination, db)
            freight_raw = self._normalize_text(row[1])
            freight_per_cbm, freight_per_ton, freight_extras, freight_warning = self._parse_lcl_freight(
                freight_raw
            )
            if freight_warning:
                warnings.append(f"{worksheet.title} row {row_index}: {freight_warning}")

            ebs_value = self._compose_dual_charge(row[3], row[4])
            cic_value = self._compose_dual_charge(row[5], row[6])
            ams_value = self._normalize_text(row[7])

            extras = {
                "sheet_name": worksheet.title,
                "row_index": row_index,
                "freight_raw": freight_raw,
                "lss_raw": self._normalize_text(row[2]),
                "ebs_raw": ebs_value,
                "cic_raw": cic_value,
                "ams_aci_ens_raw": ams_value,
            }
            extras.update(freight_extras)

            records.append(
                ParsedRateRecord(
                    record_kind="lcl",
                    carrier_name="LCL",
                    origin_port_id=origin_port["id"],
                    origin_port_name=origin_port["name"],
                    destination_port_id=destination_port["id"],
                    destination_port_name=destination_port["name"],
                    freight_per_cbm=freight_per_cbm,
                    freight_per_ton=freight_per_ton,
                    baf=None,
                    currency="USD",
                    valid_from=effective_from,
                    valid_to=effective_to,
                    sailing_day=self._normalize_text(row[8]),
                    via=self._normalize_text(row[9]),
                    transit_time_text=self._normalize_text(row[10]),
                    remarks=self._normalize_text(row[11]),
                    source_type="excel",
                    source_file=source_file,
                    extras=extras,
                )
            )

        meta = {
            "sheet_name": worksheet.title,
            "total_rows": len(records),
            "effective_from": effective_from,
            "effective_to": effective_to,
        }
        return records, meta, warnings

    def _build_fcl_row_payload(
        self,
        *,
        row: list[Any],
        row_index: int,
        sheet_name: str,
        container_type: str,
        destination: str,
        shipping_line: str,
        origin_port: dict[str, Any],
        origin_name: str | None,
        destination_port: dict[str, Any],
        effective_from: date | None,
        effective_to: date | None,
        surcharge_mode: str,
        layout: dict[str, int | None],
        warnings: list[str],
        source_file: str,
    ) -> dict[str, Any]:
        freight = _safe_decimal(self._get_layout_value(row, layout, "freight"))
        usd_surcharge_1 = self._get_layout_value(row, layout, "charge_1")
        usd_surcharge_2 = self._get_layout_value(row, layout, "charge_2")
        usd_surcharge_3 = self._get_layout_value(row, layout, "charge_3")
        usd_surcharge_4 = self._get_layout_value(row, layout, "charge_4")
        booking_charge = _safe_decimal(self._get_layout_value(row, layout, "booking_charge"))
        thc = _safe_decimal(self._get_layout_value(row, layout, "thc"))
        doc = _safe_decimal(self._get_layout_value(row, layout, "doc"))
        isps = _safe_decimal(self._get_layout_value(row, layout, "isps"))
        equipment_mgmt = _safe_decimal(self._get_layout_value(row, layout, "equipment_mgmt"))
        remarks = self._normalize_text(self._get_layout_value(row, layout, "remarks"))

        extras = {
            "sheet_name": sheet_name,
            "row_index": row_index,
            "sheet_mode": surcharge_mode,
            "origin_name": origin_name,
            "raw_usd_col_1": self._normalize_text(usd_surcharge_1),
            "raw_usd_col_2": self._normalize_text(usd_surcharge_2),
            "raw_usd_col_3": self._normalize_text(usd_surcharge_3),
            "raw_usd_col_4": self._normalize_text(usd_surcharge_4),
        }

        lss_cic = None
        baf = None
        ebs = None
        yas_caf = None
        lss_20 = None
        lss_40 = None
        baf_20 = None
        baf_40 = None

        value_1, raw_1, known_1 = self._classify_charge_value(usd_surcharge_1)
        value_2, raw_2, known_2 = self._classify_charge_value(usd_surcharge_2)
        value_3, raw_3, known_3 = self._classify_charge_value(usd_surcharge_3)
        value_4, raw_4, known_4 = self._classify_charge_value(usd_surcharge_4)

        extras["raw_usd_col_1"] = raw_1 or extras["raw_usd_col_1"]
        extras["raw_usd_col_2"] = raw_2 or extras["raw_usd_col_2"]
        extras["raw_usd_col_3"] = raw_3 or extras["raw_usd_col_3"]
        extras["raw_usd_col_4"] = raw_4 or extras["raw_usd_col_4"]

        if surcharge_mode == "jp":
            lss_cic = value_1
            baf = value_2
            ebs = value_3
            yas_caf = value_4
            if container_type == "20ft":
                lss_20 = lss_cic
                baf_20 = baf
            else:
                lss_40 = lss_cic
                baf_40 = baf
        else:
            lss_cic = value_1
            baf = value_2
            yas_caf = value_4
            if raw_3:
                extras["cic_raw"] = raw_3
            if raw_4:
                extras["caf_raw"] = raw_4
            if raw_3 and not known_3:
                warnings.append(
                    f"{sheet_name} row {row_index}: CIC contains non-numeric text '{raw_3}'."
                )
            if raw_4 and not known_4:
                warnings.append(
                    f"{sheet_name} row {row_index}: CAF contains non-numeric text '{raw_4}'."
                )
            if container_type == "20ft":
                lss_20 = lss_cic
                baf_20 = baf
            else:
                lss_40 = lss_cic
                baf_40 = baf

        if raw_1 and not known_1:
            warnings.append(
                f"{sheet_name} row {row_index}: surcharge column 1 contains non-numeric text '{raw_1}'."
            )
        if raw_2 and not known_2:
            warnings.append(
                f"{sheet_name} row {row_index}: surcharge column 2 contains non-numeric text '{raw_2}'."
            )
        if surcharge_mode == "jp" and raw_3 and not known_3:
            warnings.append(
                f"{sheet_name} row {row_index}: EBS contains non-numeric text '{raw_3}'."
            )
        if surcharge_mode == "jp" and raw_4 and not known_4:
            warnings.append(
                f"{sheet_name} row {row_index}: YAS/CAF contains non-numeric text '{raw_4}'."
            )

        return {
            "record_kind": "fcl",
            "carrier_name": shipping_line,
            "origin_port_id": origin_port["id"],
            "origin_port_name": origin_port["name"],
            "destination_port_id": destination_port["id"],
            "destination_port_name": destination_port["name"],
            "container_20gp": freight if container_type == "20ft" else None,
            "container_40gp": freight if container_type in {"40ft", "40ft_40hq"} else None,
            "container_40hq": freight if container_type in {"40hq", "40ft_40hq"} else None,
            "lss_cic": lss_cic,
            "baf": baf,
            "ebs": ebs,
            "yas_caf": yas_caf,
            "lss_20": lss_20,
            "lss_40": lss_40,
            "baf_20": baf_20,
            "baf_40": baf_40,
            "currency": "USD",
            "valid_from": effective_from,
            "valid_to": effective_to,
            "sailing_day": self._normalize_text(self._get_layout_value(row, layout, "sailing_day")),
            "via": self._normalize_text(self._get_layout_value(row, layout, "via")),
            "transit_time_text": self._normalize_text(
                self._get_layout_value(row, layout, "transit_time")
            ),
            "booking_charge": booking_charge,
            "thc": thc,
            "doc": doc,
            "isps": isps,
            "equipment_mgmt": equipment_mgmt,
            "remarks": remarks,
            "source_type": "excel",
            "source_file": source_file,
            "extras": extras,
            "_saw_40_row": False,
            "_is_orphan_block": shipping_line is None,
        }

    def _append_pending_record(
        self,
        *,
        records: list[ParsedRateRecord],
        warnings: list[str],
        pending: dict[str, Any],
        sheet_name: str,
    ) -> None:
        if not pending["_saw_40_row"]:
            warnings.append(
                f"{sheet_name} row {pending['extras']['row_index']}: 20FT row was not paired with a 40FT/40HQ row."
            )
        if pending["_is_orphan_block"]:
            warnings.append(
                f"{sheet_name} row {pending['extras']['row_index']}: carrier is blank for this destination block; kept as orphan block."
            )
        if pending["container_40gp"] is None and pending["container_40hq"] is not None:
            pending["container_40gp"] = pending["container_40hq"]
        if pending["container_40hq"] is None and pending["container_40gp"] is not None:
            pending["container_40hq"] = pending["container_40gp"]
        records.append(ParsedRateRecord(**self._materialize_record_payload(pending)))

    def _append_orphan_40_record(
        self,
        *,
        records: list[ParsedRateRecord],
        warnings: list[str],
        payload: dict[str, Any],
        sheet_name: str,
    ) -> None:
        orphan_payload = dict(payload)
        orphan_payload["container_20gp"] = None
        if orphan_payload["container_40gp"] is None and orphan_payload["container_40hq"] is not None:
            orphan_payload["container_40gp"] = orphan_payload["container_40hq"]
        if orphan_payload["container_40hq"] is None and orphan_payload["container_40gp"] is not None:
            orphan_payload["container_40hq"] = orphan_payload["container_40gp"]
        orphan_payload["_saw_40_row"] = True
        orphan_payload["_is_orphan_block"] = True
        warnings.append(
            f"{sheet_name} row {payload['extras']['row_index']}: stored orphan 40FT/40HQ row without matching 20FT row."
        )
        records.append(ParsedRateRecord(**self._materialize_record_payload(orphan_payload)))

    def _merge_40_payload(
        self,
        pending: dict[str, Any],
        payload: dict[str, Any],
        container_type: str,
    ) -> None:
        pending["_saw_40_row"] = True
        if container_type == "40ft_40hq":
            pending["container_40gp"] = payload["container_40gp"]
            pending["container_40hq"] = payload["container_40hq"]
        elif container_type == "40ft":
            pending["container_40gp"] = payload["container_40gp"]
        elif container_type == "40hq":
            pending["container_40hq"] = payload["container_40hq"]

        if payload["lss_40"] is not None:
            pending["lss_40"] = payload["lss_40"]
        if payload["baf_40"] is not None:
            pending["baf_40"] = payload["baf_40"]
        if payload["booking_charge"] is not None:
            pending["booking_charge"] = payload["booking_charge"]
        if payload["thc"] is not None:
            pending["thc"] = payload["thc"]
        if payload["doc"] is not None:
            pending["doc"] = payload["doc"]
        if payload["isps"] is not None:
            pending["isps"] = payload["isps"]
        if payload["equipment_mgmt"] is not None:
            pending["equipment_mgmt"] = payload["equipment_mgmt"]
        if payload["remarks"]:
            pending["remarks"] = payload["remarks"]

    def _materialize_record_payload(self, payload: dict[str, Any]) -> dict[str, Any]:
        return {key: value for key, value in payload.items() if not key.startswith("_")}

    def _same_pair_key(self, left: dict[str, Any], right: dict[str, Any]) -> bool:
        return (
            left["origin_port_name"] == right["origin_port_name"]
            and left["destination_port_name"] == right["destination_port_name"]
            and left["carrier_name"] == right["carrier_name"]
        )

    def _build_fcl_column_map(self, row: list[Any]) -> dict[str, int | None] | None:
        normalized = [self._normalize_header_text(value) for value in row]
        if "to" not in normalized or "shipping line" not in normalized:
            return None

        layout: dict[str, int | None] = {
            "destination": None,
            "shipping_line": None,
            "freight": None,
            "charge_1": None,
            "charge_2": None,
            "charge_3": None,
            "charge_4": None,
            "sailing_day": None,
            "via": None,
            "transit_time": None,
            "booking_charge": None,
            "thc": None,
            "doc": None,
            "isps": None,
            "equipment_mgmt": None,
            "remarks": None,
        }

        for index, cell in enumerate(normalized):
            if cell == "to":
                layout["destination"] = index
            elif cell == "shipping line":
                layout["shipping_line"] = index
            elif cell.startswith("freight"):
                layout["freight"] = index
            elif cell.startswith("lss+cic") or cell == "lss":
                layout["charge_1"] = index
            elif cell == "baf":
                layout["charge_2"] = index
            elif cell == "ebs" or cell == "cic":
                layout["charge_3"] = index
            elif cell == "caf" or "yas/caf" in cell or cell == "yas caf":
                layout["charge_4"] = index
            elif cell == "sailing day":
                layout["sailing_day"] = index
            elif cell == "via":
                layout["via"] = index
            elif cell == "transit time":
                layout["transit_time"] = index
            elif cell.startswith("booking charge"):
                layout["booking_charge"] = index
            elif cell.startswith("thc"):
                layout["thc"] = index
            elif cell.startswith("doc"):
                layout["doc"] = index
            elif cell == "isps":
                layout["isps"] = index
            elif "equipment management fee" in cell or cell == "emf":
                layout["equipment_mgmt"] = index
            elif cell == "rmks":
                layout["remarks"] = index

        required_fields = ("destination", "shipping_line", "freight")
        if any(layout[field] is None for field in required_fields):
            return None
        return layout

    def _get_layout_value(
        self,
        row: list[Any],
        layout: dict[str, int | None],
        field: str,
    ) -> Any:
        index = layout.get(field)
        if index is None or index >= len(row):
            return None
        return row[index]

    def _read_effective_range(self, worksheet) -> tuple[date | None, date | None]:
        return (
            self._to_date(worksheet["B3"].value),
            self._to_date(worksheet["D3"].value),
        )

    def _read_origin_name(self, worksheet) -> str:
        raw = self._normalize_text(worksheet["A7"].value) or "Shanghai"
        if ":" in raw:
            return raw.split(":", 1)[1].strip()
        return raw

    def _read_origin_name_from_row(self, value: Any) -> str | None:
        raw = self._normalize_text(value)
        if not raw or not raw.lower().startswith("from:"):
            return None
        return raw.split(":", 1)[1].strip() or None

    def _resolve_port_ref(self, port_name: str | None, db: Session | None) -> dict[str, Any]:
        if not port_name:
            return {"id": None, "name": None}
        if db is None:
            return {"id": None, "name": port_name}
        port = _resolve_port(port_name, db)
        if port is None:
            return {"id": None, "name": port_name}
        return {"id": port.id, "name": f"{port.name_en}/{port.name_cn or port.name_en}"}

    def _parse_lcl_freight(
        self,
        freight_raw: str | None,
    ) -> tuple[Decimal | None, Decimal | None, dict[str, Any], str | None]:
        extras: dict[str, Any] = {}
        if not freight_raw:
            return None, None, extras, None

        if freight_raw.strip().lower() in self._LCL_FREIGHT_PLACEHOLDERS:
            extras["freight_parse_status"] = "placeholder"
            return None, None, extras, None

        normalized = freight_raw.replace(" ", "").upper()
        if "/CBM" in normalized or "/TON" in normalized:
            per_cbm = self._extract_unit_decimal(normalized, "CBM")
            per_ton = self._extract_unit_decimal(normalized, "TON")
            # 组合分支（如 "0/CBM, 0/TON"）只发一条 zero warning，避免每元一条重复刷屏
            zero_hit = False
            if per_cbm is not None and per_cbm == Decimal("0"):
                per_cbm = None
                zero_hit = True
            if per_ton is not None and per_ton == Decimal("0"):
                per_ton = None
                zero_hit = True
            if zero_hit:
                return per_cbm, per_ton, extras, f"zero freight rate ignored: {freight_raw}"
            return per_cbm, per_ton, extras, None

        if normalized.endswith("/RT"):
            value = _safe_decimal(normalized[:-3])
            extras["freight_unit"] = "RT"
            if value is not None and value == Decimal("0"):
                return None, None, extras, f"zero freight rate ignored: {freight_raw}"
            return value, value, extras, None

        if normalized.endswith("/CBM"):
            value = _safe_decimal(normalized[:-4])
            extras["freight_unit"] = "CBM"
            if value is not None and value == Decimal("0"):
                return None, None, extras, f"zero freight rate ignored: {freight_raw}"
            return value, None, extras, None

        if normalized.endswith("/TON"):
            value = _safe_decimal(normalized[:-4])
            extras["freight_unit"] = "TON"
            if value is not None and value == Decimal("0"):
                return None, None, extras, f"zero freight rate ignored: {freight_raw}"
            return None, value, extras, None

        extras["freight_parse_status"] = "raw_only"
        return None, None, extras, f"non-standard freight value '{freight_raw}' kept as raw text"

    def _extract_unit_decimal(self, freight_raw: str, unit: str) -> Decimal | None:
        for segment in freight_raw.split(","):
            if f"/{unit}" in segment:
                return _safe_decimal(segment.split(f"/{unit}", 1)[0])
        return None

    def _compose_dual_charge(self, primary: Any, secondary: Any) -> str | None:
        primary_text = self._normalize_text(primary)
        secondary_text = self._normalize_text(secondary)
        if primary_text and secondary_text:
            return f"USD:{primary_text}; CNY:{secondary_text}"
        if primary_text:
            return primary_text
        if secondary_text:
            return f"CNY:{secondary_text}"
        return None

    def _normalize_container_type(self, value: Any) -> str | None:
        text = self._normalize_text(value)
        if not text:
            return None
        upper = text.upper().replace(" ", "")
        if upper in {"20FT", "20GP", "20'"} or upper.startswith("20"):
            return "20ft"
        if upper in {"40FT", "40GP"}:
            return "40ft"
        if upper == "40HQ":
            return "40hq"
        if "40FT/40HQ" in upper or ("40FT" in upper and "40HQ" in upper):
            return "40ft_40hq"
        if upper.startswith("40"):
            return "40ft_40hq"
        return None

    def _normalize_text(self, value: Any) -> str | None:
        if value is None:
            return None
        text = self._WHITESPACE_RE.sub(" ", str(value)).strip()
        return text or None

    def _normalize_header_text(self, value: Any) -> str:
        text = self._normalize_text(value)
        if not text:
            return ""
        return re.sub(r"[^a-z0-9/+]+", " ", text.lower()).strip()

    def _classify_charge_value(self, value: Any) -> tuple[Decimal | None, str | None, bool]:
        raw_text = self._normalize_text(value)
        if raw_text is None:
            return None, None, False
        decimal_value = _safe_decimal(raw_text)
        if decimal_value is not None:
            return decimal_value, None, False
        labeled_decimal = self._extract_labeled_decimal(raw_text)
        if labeled_decimal is not None:
            return labeled_decimal, raw_text, True
        normalized = raw_text.lower()
        known = any(marker in normalized for marker in self._KNOWN_NON_NUMERIC_MARKERS)
        return None, raw_text, known

    def _extract_labeled_decimal(self, raw_text: str) -> Decimal | None:
        match = re.match(r"^[a-z/ +]+:\s*(?:usd)?\s*([0-9]+(?:\.[0-9]+)?)", raw_text.lower())
        if match is None:
            return None
        return _safe_decimal(match.group(1))

    def _is_empty_row(self, row: Iterable[Any]) -> bool:
        return all(value is None or str(value).strip() == "" for value in row)

    def _to_date(self, value: Any) -> date | None:
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        return None

    def _dedupe_warnings(self, warnings: Iterable[str]) -> list[str]:
        deduped: list[str] = []
        seen: set[str] = set()
        for warning in warnings:
            if warning in seen:
                continue
            deduped.append(warning)
            seen.add(warning)
        return deduped
