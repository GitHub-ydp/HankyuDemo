from __future__ import annotations

from collections.abc import Iterable
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.services.step1_rates.entities import ParsedRateBatch, ParsedRateRecord, Step1FileType


class AirAdapter:
    """Step1 Air parser for the weekly market price + Surcharges workbook."""

    key = "air"
    file_type = Step1FileType.air
    priority = 10

    _WEEKLY_SHEET_RE = re.compile(
        r"^([A-Z][a-z]{2})\s+(\d{1,2})\s+to\s+([A-Z][a-z]{2})\s+(\d{1,2})$"
    )
    _MONTH_ABBR = {
        "Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "May": 5, "Jun": 6,
        "Jul": 7, "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12,
    }
    _SURCHARGES_EXPECTED_HEADER = [
        "area", "from", "airlines", "effective date",
        "myc min", "myc fee/kg", "msc min", "msc fee/kg",
        "destination", "remarks",
    ]
    _YEAR_FROM_HEADER_RE = re.compile(r"(\d{4})\s*[/\-\.]\s*\d{1,2}\s*[/\-\.]\s*\d{1,2}")
    _MUST_GO_RE = re.compile(r"must\s*go\s*([0-9]+(?:\.[0-9]+)?)", re.IGNORECASE)
    _AIRLINE_CODE_RE = re.compile(r"(?<![A-Za-z])[A-Z]{2}(?![A-Za-z])")
    _INLINE_WHITESPACE_RE = re.compile(r"[ \t]+")
    _DEFAULT_ORIGIN = "PVG"
    _DEFAULT_CURRENCY = "CNY"

    def detect(self, path: Path, *, file_type_hint: Step1FileType | None = None) -> bool:
        if file_type_hint == self.file_type:
            return True
        name = path.name.lower()
        if "ocean" in name or "ngb" in name:
            return False
        return "air" in name or "market price" in name

    def parse(self, path: Path, db: Session | None = None) -> ParsedRateBatch:
        workbook = load_workbook(path, data_only=True)
        warnings: list[str] = []
        records: list[ParsedRateRecord] = []
        weekly_meta: list[dict[str, Any]] = []

        surcharges_sheet = self._find_surcharges_sheet(workbook)
        if surcharges_sheet is None:
            weekly_currency = self._DEFAULT_CURRENCY
            weekly_currency_source = "fallback_no_surcharges"
            weekly_currency_warning = (
                "weekly currency falls back to CNY: Surcharges sheet missing"
            )
        else:
            weekly_currency, surcharges_currency_source, _ = self._read_surcharges_currency(
                workbook[surcharges_sheet]
            )
            if surcharges_currency_source == "F2":
                weekly_currency_source = "from_surcharges_F2"
                weekly_currency_warning = None
            else:
                weekly_currency = self._DEFAULT_CURRENCY
                weekly_currency_source = "fallback_no_F2_declaration"
                weekly_currency_warning = (
                    "weekly currency falls back to CNY: "
                    "Surcharges F2 has no CURRENCY declaration"
                )

        weekly_sheet_names = self._detect_weekly_sheet_names(workbook)
        if not weekly_sheet_names:
            warnings.append("no weekly sheet matched pattern 'Mon dd to Mon dd'")

        for sheet_name in weekly_sheet_names:
            worksheet = workbook[sheet_name]
            sheet_records, sheet_meta, sheet_warnings = self._parse_weekly_sheet(
                worksheet,
                source_file=path.name,
                currency=weekly_currency,
                currency_source=weekly_currency_source,
            )
            records.extend(sheet_records)
            warnings.extend(sheet_warnings)
            weekly_meta.append(sheet_meta)

        if weekly_sheet_names and weekly_currency_warning:
            warnings.append(weekly_currency_warning)

        batch_effective_from: date | None = None
        batch_effective_to: date | None = None
        if weekly_meta:
            dated_meta = [m for m in weekly_meta if m.get("effective_week_start") is not None]
            if dated_meta:
                latest = max(dated_meta, key=lambda m: m["effective_week_start"])
            else:
                latest = weekly_meta[-1]
            batch_effective_from = latest.get("effective_week_start")
            batch_effective_to = latest.get("effective_week_end")

            if len(weekly_sheet_names) >= 2:
                warnings.append(
                    "workbook contains {n} weekly sheets: {names}; "
                    "batch effective range uses latest week {start} - {end}".format(
                        n=len(weekly_sheet_names),
                        names=weekly_sheet_names,
                        start=batch_effective_from.isoformat() if batch_effective_from else None,
                        end=batch_effective_to.isoformat() if batch_effective_to else None,
                    )
                )

        if surcharges_sheet is None:
            warnings.append("Surcharges sheet missing, skipped airline surcharge parsing")
        else:
            surcharge_records, surcharge_warnings = self._parse_surcharges_sheet(
                workbook[surcharges_sheet],
                source_file=path.name,
            )
            records.extend(surcharge_records)
            warnings.extend(surcharge_warnings)

        return ParsedRateBatch(
            file_type=self.file_type,
            source_file=path.name,
            effective_from=batch_effective_from,
            effective_to=batch_effective_to,
            records=records,
            warnings=self._dedupe_warnings(warnings),
            adapter_key=self.key,
            metadata={
                "file_name": path.name,
                "source_type": "excel",
                "parser_version": "air_v1",
                "air_origin_assumption": "PVG (no origin column in weekly sheet)",
                "weekly_sheets": [
                    {
                        "sheet_name": meta["sheet_name"],
                        "total_rows": meta["total_rows"],
                        "effective_week_start": meta["effective_week_start"],
                        "effective_week_end": meta["effective_week_end"],
                    }
                    for meta in weekly_meta
                ],
                "surcharges_sheet": surcharges_sheet,
            },
        )

    def _detect_weekly_sheet_names(self, workbook) -> list[str]:
        return [name for name in workbook.sheetnames if self._WEEKLY_SHEET_RE.match(name)]

    def _find_surcharges_sheet(self, workbook) -> str | None:
        for name in workbook.sheetnames:
            if name.strip().lower() == "surcharges":
                return name
        return None

    def _parse_weekly_sheet(
        self,
        worksheet,
        *,
        source_file: str,
        currency: str,
        currency_source: str,
    ) -> tuple[list[ParsedRateRecord], dict[str, Any], list[str]]:
        sheet_name = worksheet.title
        warnings: list[str] = []
        records: list[ParsedRateRecord] = []

        week_start, week_end, week_warning = self._parse_week_range_from_sheet_name(
            sheet_name, worksheet
        )
        if week_warning:
            warnings.append(week_warning)

        header_a1 = self._normalize_text(worksheet.cell(1, 1).value)
        if header_a1 != "Destinations":
            warnings.append(
                f"{sheet_name} header mismatch at A1; expected 'Destinations', got '{header_a1}'"
            )

        consecutive_empty = 0
        max_row = min(worksheet.max_row or 0, 100)

        for row_index in range(2, max_row + 1):
            row = [worksheet.cell(row=row_index, column=col).value for col in range(1, 12)]
            destination_raw = row[0]
            service_raw = row[1]
            if destination_raw is None and service_raw is None:
                consecutive_empty += 1
                if consecutive_empty >= 2:
                    break
                continue
            consecutive_empty = 0

            destination_text = self._normalize_text(destination_raw) or ""
            service_text = self._normalize_service_text(service_raw)
            remark_text = self._normalize_service_text(row[9])

            price_days: list[Decimal | None] = []
            price_day_missing: list[int] = []
            price_raw_extras: dict[str, Any] = {}
            for idx in range(7):
                day_no = idx + 1
                raw_value = row[2 + idx]
                if raw_value is None or (isinstance(raw_value, str) and raw_value.strip() == ""):
                    price_days.append(None)
                    price_day_missing.append(day_no)
                    warnings.append(
                        f"{sheet_name} row {row_index}: price_day{day_no} is empty"
                    )
                    continue
                decimal_value = self._to_decimal(raw_value)
                if decimal_value is None:
                    price_days.append(None)
                    raw_text = self._normalize_text(raw_value)
                    price_raw_extras[f"price_day{day_no}_raw"] = raw_text
                    warnings.append(
                        f"{sheet_name} row {row_index}: price_day{day_no} "
                        f"contains non-numeric '{raw_text}'"
                    )
                else:
                    price_days.append(decimal_value)

            destination_airports, density_hint = self._parse_destination(destination_text)
            airline_codes = self._extract_airline_codes(service_text or "")
            remark_lower = (remark_text or "").lower()
            must_go_match = self._MUST_GO_RE.search(remark_text or "")
            must_go_value = (
                self._to_decimal(must_go_match.group(1)) if must_go_match else None
            )

            extras: dict[str, Any] = {
                "sheet_name": sheet_name,
                "row_index": row_index,
                "week_label": sheet_name,
                "raw_destination": destination_raw,
                "raw_service": service_raw,
                "raw_remark": row[9],
                "destination_airports": destination_airports,
                "density_hint": density_hint,
                "airline_codes": airline_codes,
                "has_must_go": "must go" in remark_lower,
                "must_go_value": must_go_value,
                "is_case_by_case": "case by case" in remark_lower,
                "price_day_missing": price_day_missing,
                "currency_source": currency_source,
                "origin_source": "default_air_PVG",
            }
            extras.update(price_raw_extras)

            records.append(
                ParsedRateRecord(
                    record_kind="air_weekly",
                    carrier_name=None,
                    carrier_id=None,
                    airline_code="/".join(airline_codes) if airline_codes else None,
                    service_desc=service_text,
                    origin_port_id=None,
                    origin_port_name=self._DEFAULT_ORIGIN,
                    destination_port_id=None,
                    destination_port_name=destination_text or None,
                    price_day1=price_days[0],
                    price_day2=price_days[1],
                    price_day3=price_days[2],
                    price_day4=price_days[3],
                    price_day5=price_days[4],
                    price_day6=price_days[5],
                    price_day7=price_days[6],
                    effective_week_start=week_start,
                    effective_week_end=week_end,
                    valid_from=week_start,
                    valid_to=week_end,
                    currency=currency,
                    remarks=remark_text,
                    source_type="excel",
                    source_file=source_file,
                    extras=extras,
                )
            )

        meta = {
            "sheet_name": sheet_name,
            "total_rows": len(records),
            "effective_week_start": week_start,
            "effective_week_end": week_end,
        }
        return records, meta, warnings

    def _parse_surcharges_sheet(
        self,
        worksheet,
        *,
        source_file: str,
    ) -> tuple[list[ParsedRateRecord], list[str]]:
        sheet_name = worksheet.title
        warnings: list[str] = []
        records: list[ParsedRateRecord] = []

        currency, currency_source, currency_warning = self._read_surcharges_currency(worksheet)
        if currency_warning:
            warnings.append(currency_warning)

        header_row = [self._normalize_text(worksheet.cell(4, col).value) for col in range(2, 12)]
        header_keys = [h.lower() if h else "" for h in header_row]
        if header_keys != self._SURCHARGES_EXPECTED_HEADER:
            warnings.append(
                "Surcharges header mismatch at row 4; expected AREA/FROM/...; "
                f"got {header_row}"
            )

        from_merged_value = self._read_merged_value_for(worksheet, anchor_col=3)

        last_area: str | None = None
        consecutive_empty = 0
        max_row = min(worksheet.max_row or 0, 100)

        for row_index in range(5, max_row + 1):
            airlines_raw = worksheet.cell(row_index, 4).value
            airlines_text = self._normalize_text(airlines_raw)
            if airlines_text is None:
                consecutive_empty += 1
                if consecutive_empty >= 3:
                    break
                continue
            consecutive_empty = 0

            area_raw = worksheet.cell(row_index, 2).value
            area_text = self._normalize_text(area_raw)
            if area_text:
                last_area = area_text
            elif last_area is None:
                warnings.append(
                    f"Surcharges row {row_index}: AREA empty and no previous value to fill forward"
                )

            from_cell = worksheet.cell(row_index, 3).value
            from_value = from_cell if from_cell is not None else from_merged_value
            from_region, from_raw = self._clean_from_value(from_value)
            if not from_region:
                warnings.append(
                    f"Surcharges row {row_index}: FROM column empty after merge expansion"
                )

            effective_raw = worksheet.cell(row_index, 5).value
            valid_from, date_warning = self._to_date_lenient(effective_raw)
            if date_warning:
                warnings.append(
                    f"Surcharges row {row_index}: cannot parse effective date '{date_warning}'"
                )

            myc_min_value, myc_min_dash = self._to_decimal_or_dash(worksheet.cell(row_index, 6).value)
            myc_fee_value, myc_fee_dash = self._to_decimal_or_dash(worksheet.cell(row_index, 7).value)
            msc_min_value, msc_min_dash = self._to_decimal_or_dash(worksheet.cell(row_index, 8).value)
            msc_fee_value, msc_fee_dash = self._to_decimal_or_dash(worksheet.cell(row_index, 9).value)

            all_fees_dash = all([myc_min_dash, myc_fee_dash, msc_min_dash, msc_fee_dash])
            all_fees_empty = (
                not myc_min_dash and myc_min_value is None
                and not myc_fee_dash and myc_fee_value is None
                and not msc_min_dash and msc_min_value is None
                and not msc_fee_dash and msc_fee_value is None
            )
            if all_fees_empty:
                warnings.append(
                    f"Surcharges row {row_index} ({airlines_text}): all four fee fields empty"
                )

            destination_scope = self._normalize_text(worksheet.cell(row_index, 10).value)
            remarks_text = self._normalize_text(worksheet.cell(row_index, 11).value)

            airline_iata = self._extract_airline_iata(airlines_text)

            extras: dict[str, Any] = {
                "sheet_name": sheet_name,
                "row_index": row_index,
                "area": last_area,
                "from_region": from_region,
                "raw_from": from_raw,
                "airline_code_raw": airlines_text,
                "airline_iata": airline_iata,
                "effective_date_raw": effective_raw,
                "myc_min_value": myc_min_value,
                "myc_min_is_dash": myc_min_dash,
                "myc_fee_per_kg": myc_fee_value,
                "myc_fee_is_dash": myc_fee_dash,
                "msc_min_value": msc_min_value,
                "msc_min_is_dash": msc_min_dash,
                "msc_fee_per_kg": msc_fee_value,
                "msc_fee_is_dash": msc_fee_dash,
                "destination_scope": destination_scope,
                "all_fees_dash": all_fees_dash,
                "all_fees_empty": all_fees_empty,
                "currency_source": currency_source,
            }

            records.append(
                ParsedRateRecord(
                    record_kind="air_surcharge",
                    carrier_name=airlines_text,
                    carrier_id=None,
                    airline_code=airline_iata,
                    service_desc=None,
                    origin_port_id=None,
                    origin_port_name=self._DEFAULT_ORIGIN,
                    destination_port_id=None,
                    destination_port_name=None,
                    effective_week_start=None,
                    effective_week_end=None,
                    valid_from=valid_from,
                    valid_to=None,
                    currency=currency,
                    remarks=remarks_text,
                    source_type="excel",
                    source_file=source_file,
                    extras=extras,
                )
            )

        return records, warnings

    def _parse_week_range_from_sheet_name(
        self,
        sheet_name: str,
        worksheet,
    ) -> tuple[date | None, date | None, str | None]:
        match = self._WEEKLY_SHEET_RE.match(sheet_name)
        if match is None:
            return None, None, f"cannot parse week range from sheet name '{sheet_name}'"
        start_month = self._MONTH_ABBR.get(match.group(1))
        end_month = self._MONTH_ABBR.get(match.group(3))
        start_day = int(match.group(2))
        end_day = int(match.group(4))
        if start_month is None or end_month is None:
            return None, None, f"cannot parse week range from sheet name '{sheet_name}'"

        year = self._extract_year_from_header(worksheet.cell(1, 3).value) or date.today().year
        try:
            start = date(year, start_month, start_day)
            end = date(year, end_month, end_day)
        except ValueError:
            return None, None, f"cannot parse week range from sheet name '{sheet_name}'"
        if end < start:
            try:
                end = end.replace(year=year + 1)
            except ValueError:
                pass
        return start, end, None

    def _extract_year_from_header(self, value: Any) -> int | None:
        if isinstance(value, datetime):
            return value.year
        if isinstance(value, date):
            return value.year
        text = self._normalize_text(value)
        if not text:
            return None
        match = self._YEAR_FROM_HEADER_RE.search(text)
        if match is None:
            return None
        try:
            return int(match.group(1))
        except ValueError:
            return None

    def _parse_destination(self, destination_text: str) -> tuple[list[str], str | None]:
        if not destination_text:
            return [], None
        airports: list[str] = []
        density: str | None = None
        tokens = re.split(r"[\s/]+", destination_text)
        for token in tokens:
            cleaned = token.strip()
            if not cleaned:
                continue
            if len(cleaned) == 3 and cleaned.isalpha() and cleaned.isupper():
                airports.append(cleaned)
            elif cleaned.upper() in {"DENSE", "VOLUME"}:
                density = cleaned.upper()
        deduped_airports: list[str] = []
        for code in airports:
            if code not in deduped_airports:
                deduped_airports.append(code)
        return deduped_airports, density

    def _extract_airline_codes(self, service_desc: str) -> list[str]:
        codes = self._AIRLINE_CODE_RE.findall(service_desc)
        return sorted(set(codes))

    def _extract_airline_iata(self, airlines_text: str | None) -> str | None:
        if not airlines_text:
            return None
        prefix = re.split(r"[\s\-]", airlines_text, maxsplit=1)[0].strip()
        if not prefix:
            return None
        return prefix

    def _read_surcharges_currency(self, worksheet) -> tuple[str, str, str | None]:
        raw = worksheet["F2"].value
        text = self._normalize_text(raw)
        if text:
            match = re.search(r"CURRENCY\s*:\s*([A-Z]{3})", text, re.IGNORECASE)
            if match:
                return match.group(1).upper(), "F2", None
        return (
            self._DEFAULT_CURRENCY,
            "fallback",
            "Surcharges F2 currency tag missing or unparseable, fallback to CNY",
        )

    def _read_merged_value_for(self, worksheet, *, anchor_col: int) -> Any:
        for merged_range in worksheet.merged_cells.ranges:
            if merged_range.min_col == anchor_col == merged_range.max_col:
                anchor_value = worksheet.cell(
                    row=merged_range.min_row, column=merged_range.min_col
                ).value
                if anchor_value is not None:
                    return anchor_value
        return None

    def _clean_from_value(self, value: Any) -> tuple[str | None, Any]:
        if value is None:
            return None, None
        raw_text = str(value)
        first_segment = re.split(r"\n|\s{2,}", raw_text)[0]
        cleaned = self._normalize_text(first_segment)
        return cleaned, value

    def _to_decimal_or_dash(self, value: Any) -> tuple[Decimal | None, bool]:
        if value is None:
            return None, False
        if isinstance(value, str):
            stripped = value.strip()
            if stripped == "":
                return None, False
            if stripped == "-":
                return None, True
            decimal_value = self._to_decimal(stripped)
            return decimal_value, False
        decimal_value = self._to_decimal(value)
        return decimal_value, False

    def _to_decimal(self, value: Any) -> Decimal | None:
        if value is None:
            return None
        if isinstance(value, Decimal):
            return value
        if isinstance(value, bool):
            return None
        if isinstance(value, (int, float)):
            try:
                return Decimal(str(value))
            except (InvalidOperation, ValueError):
                return None
        if isinstance(value, str):
            stripped = value.strip()
            if stripped == "":
                return None
            try:
                return Decimal(stripped)
            except (InvalidOperation, ValueError):
                return None
        return None

    def _to_date_lenient(self, value: Any) -> tuple[date | None, str | None]:
        if value is None:
            return None, None
        if isinstance(value, datetime):
            return value.date(), None
        if isinstance(value, date):
            return value, None
        if isinstance(value, str):
            text = value.strip()
            if not text:
                return None, None
            for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
                try:
                    return datetime.strptime(text, fmt).date(), None
                except ValueError:
                    continue
            return None, text
        return None, str(value)

    def _normalize_text(self, value: Any) -> str | None:
        if value is None:
            return None
        text = re.sub(r"\s+", " ", str(value)).strip()
        return text or None

    def _normalize_service_text(self, value: Any) -> str | None:
        if value is None:
            return None
        text = str(value)
        text = self._INLINE_WHITESPACE_RE.sub(" ", text).strip()
        return text or None

    def _dedupe_warnings(self, warnings: Iterable[str]) -> list[str]:
        deduped: list[str] = []
        seen: set[str] = set()
        for warning in warnings:
            if warning in seen:
                continue
            deduped.append(warning)
            seen.add(warning)
        return deduped
