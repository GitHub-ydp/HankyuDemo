from __future__ import annotations

from collections.abc import Iterable
from dataclasses import dataclass
from datetime import date
from decimal import Decimal
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.services.rate_parser import (
    PORT_ALIAS_MAP,
    _TERMINAL_CONNECTORS,
    _TERMINAL_SUFFIXES,
    _safe_decimal,
)
from app.services.step1_rates.entities import ParsedRateBatch, ParsedRateRecord, Step1FileType


_STATE_SUFFIX_RE = re.compile(r",\s*[A-Z]{2}\s*$")
_PAREN_RE = re.compile(r"[（(].*?[）)]")
_LOCODE_RE = re.compile(r"^[A-Z]{5}$")


def _nvo_safe_decimal(value: Any) -> Decimal | None:
    """扩展 _safe_decimal — 处理 '$4,160' 风格 + '' → None。"""
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        text = text.lstrip("$").replace(",", "").strip()
        if not text:
            return None
        return _safe_decimal(text)
    return _safe_decimal(value)


def _split_origins(s: str | None) -> list[str]:
    """'KRPUS,KRKAN' → ['KRPUS','KRKAN']；'SGSIN, THLCH' → 同样按 , 拆。"""
    if s is None:
        return []
    text = str(s).strip()
    if not text:
        return []
    return [part.strip() for part in text.split(",") if part.strip()]


def _clean_destination(name_raw: str | None) -> str | None:
    """destination：去 ', ST'、去括号、查 alias → LOCODE 或纯英文。

    多港描述（"Los Angeles, Long Beach"）取第一段；adapter 不接 db。
    """
    if name_raw is None:
        return None
    text = str(name_raw).strip()
    if not text:
        return None

    # 多港描述（含 / 或 ,）取第一段（注意：先按 / 切再按 , 切，因为
    # "Cape Town / Durban / Coega" 与 "Los Angeles, Long Beach" 都属于多港）
    head = text
    if "/" in head:
        head = head.split("/")[0].strip()
    if "," in head:
        # 注意：可能是 "Norfolk, VA"（state 缩写）或 "Los Angeles, Long Beach"
        # 先剥 state 后缀，否则两段都被切掉就只剩第一个
        head_no_state = _STATE_SUFFIX_RE.sub("", head).strip()
        if head_no_state == head:
            # 没有 state 后缀，是多港 → 取第一段
            head = head.split(",")[0].strip()
        else:
            head = head_no_state
    head = _PAREN_RE.sub("", head).strip()
    if not head:
        return None

    # 终端后缀清理（沿用 KMTC _clean_and_resolve_port）
    for suf in _TERMINAL_SUFFIXES:
        if head.upper().endswith(suf):
            head = head[: -len(suf)].strip()
            break
    for conn in _TERMINAL_CONNECTORS:
        if conn in head:
            head = head.split(conn)[0].strip()
            break
    if not head:
        return None

    locode = PORT_ALIAS_MAP.get(head.lower())
    if locode:
        return locode
    return head


def _clean_origin_locode(name_raw: str | None) -> str | None:
    """origin：先查 alias（PUSAN/MANILA 等英文全大写港名要走 alias 命中 LOCODE），
    未命中且形态符合 LOCODE 则直接返回 LOCODE（如 KRPUS、INHZA）；
    否则走 _clean_destination 同套清洗。
    """
    if name_raw is None:
        return None
    text = str(name_raw).strip()
    if not text:
        return None
    locode = PORT_ALIAS_MAP.get(text.lower())
    if locode:
        return locode
    if _LOCODE_RE.match(text):
        return text
    return _clean_destination(text)


@dataclass(frozen=True)
class _Section:
    sheet_name: str
    header_row: int
    end_row: int
    layout: dict[str, int]
    section_label: str | None
    section_kind: str  # 'main_us' / 'main_ca' / 'ipi' / 'africa'


class NvoFakAdapter:
    """Step1 NVO FAK parser (4 sheets, multi-section headers, USD)."""

    key: str = "nvo_fak"
    file_type: Step1FileType = Step1FileType.ocean
    priority: int = 5

    _CARRIER_NAME: str = "NVO_FAK"
    _CURRENCY: str = "USD"
    _PARSER_VERSION: str = "nvo_fak_v1"

    _SHEET_TPE: str = "TPE"
    _SHEET_WPE: str = "WPE"
    _SHEET_HAWAII: str = "Hawaii"
    _SHEET_ARBITRARY: str = "Arbitrary"

    _DETECT_NAME_KEYWORDS_LOWER: tuple[str, ...] = ("nvo", "fak")
    _DETECT_SHEET_KEYWORDS: tuple[str, ...] = ("TPE", "WPE", "Hawaii")

    _MAX_COL: int = 12

    _EFFECTIVE_RE = re.compile(
        r"[Ee]ffective\s+(?:from\s+)?(\d{1,2})[/\-](\d{1,2})\s+to\s+(\d{1,2})[/\-](\d{1,2})"
    )
    _YEAR_RE = re.compile(r"20\d{2}")
    _BASE_PORTS_RE = re.compile(r"All Base Ports[:\s]+(.*)")
    _WHITESPACE_RE = re.compile(r"\s+")

    # ------------------------------------------------------------------
    # detect / parse
    # ------------------------------------------------------------------

    def detect(self, path: Path, *, file_type_hint: Step1FileType | None = None) -> bool:
        name_lower = path.name.lower()
        if any(kw in name_lower for kw in self._DETECT_NAME_KEYWORDS_LOWER):
            return True
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            try:
                sheetnames = list(wb.sheetnames)
            finally:
                wb.close()
        except Exception:
            return False
        for sname in sheetnames:
            for kw in self._DETECT_SHEET_KEYWORDS:
                if kw == sname:
                    return True
        return False

    def parse(self, path: Path, db: Session | None = None) -> ParsedRateBatch:
        all_records: list[ParsedRateRecord] = []
        all_warnings: list[str] = []
        sheets_summary: list[dict[str, Any]] = []
        section_notes: list[dict[str, Any]] = []
        unseeded_origin_locodes: set[str] = set()

        wb = load_workbook(path, data_only=True)

        base_ports: list[str] = []
        if self._SHEET_TPE in wb.sheetnames:
            base_ports = self._extract_base_ports(wb[self._SHEET_TPE])

        for sname in wb.sheetnames:
            ws = wb[sname]
            if sname.lower() == self._SHEET_ARBITRARY.lower():
                all_warnings.append(
                    f"sheet '{sname}' skipped (inland arbitrary fees, not freight rate)"
                )
                sheets_summary.append(
                    {
                        "sheet_name": sname,
                        "skipped": True,
                        "skip_reason": "inland arbitrary fees, not freight rate",
                    }
                )
                continue
            if sname == self._SHEET_HAWAII:
                eff_from, eff_to = self._extract_effective_dates(ws, path)
                recs, warns, summary = self._parse_hawaii_sheet(
                    ws, eff_from, eff_to, path.name
                )
                all_records.extend(recs)
                all_warnings.extend(warns)
                sheets_summary.append(summary)
                continue
            if sname in (self._SHEET_TPE, self._SHEET_WPE):
                eff_from, eff_to = self._extract_effective_dates(ws, path)
                recs, warns, summary, notes, unseeded = self._parse_main_sheet(
                    ws, sname, eff_from, eff_to, path.name
                )
                all_records.extend(recs)
                all_warnings.extend(warns)
                sheets_summary.append(summary)
                section_notes.extend(notes)
                unseeded_origin_locodes.update(unseeded)
                continue
            all_warnings.append(f"unexpected sheet '{sname}' skipped")

        eff_from_values = [
            s["effective_from"] for s in sheets_summary if s.get("effective_from")
        ]
        eff_to_values = [s["effective_to"] for s in sheets_summary if s.get("effective_to")]
        batch_eff_from = min(eff_from_values) if eff_from_values else None
        batch_eff_to = max(eff_to_values) if eff_to_values else None

        if unseeded_origin_locodes:
            all_warnings.append(
                "unseeded origin LOCODEs (will fail activation port lookup): "
                + ", ".join(sorted(unseeded_origin_locodes))
            )

        kind_distribution = {"ocean_ngb_fcl": len(all_records)}
        metadata = {
            "file_name": path.name,
            "source_type": "excel",
            "carrier_code": self._CARRIER_NAME,
            "parser_version": self._PARSER_VERSION,
            "adapter_key": self.key,
            "base_ports": base_ports,
            "sheets": sheets_summary,
            "record_kind_distribution": kind_distribution,
            "unseeded_origin_locodes": sorted(unseeded_origin_locodes),
            "section_notes": section_notes,
        }

        return ParsedRateBatch(
            file_type=self.file_type,
            source_file=path.name,
            effective_from=batch_eff_from,
            effective_to=batch_eff_to,
            records=all_records,
            warnings=self._dedupe_warnings(all_warnings),
            adapter_key=self.key,
            metadata=metadata,
        )

    # ------------------------------------------------------------------
    # main sheet (TPE / WPE)
    # ------------------------------------------------------------------

    def _parse_main_sheet(
        self,
        ws,
        sheet_name: str,
        eff_from: date | None,
        eff_to: date | None,
        source_file: str,
    ) -> tuple[
        list[ParsedRateRecord],
        list[str],
        dict[str, Any],
        list[dict[str, Any]],
        set[str],
    ]:
        records: list[ParsedRateRecord] = []
        warnings: list[str] = []
        notes: list[dict[str, Any]] = []
        unseeded: set[str] = set()

        ipi_count = 0
        africa_count = 0

        for sec in self._iter_sections(ws, sheet_name):
            if sec.section_kind == "ipi":
                # IPI 段不出 record，仅计数
                row_count = 0
                for ri in range(sec.header_row + 1, sec.end_row + 1):
                    row = [
                        ws.cell(row=ri, column=c).value
                        for c in range(1, self._MAX_COL + 1)
                    ]
                    if self._is_empty_row(row):
                        continue
                    a_text = self._normalize_text(row[0])
                    if not a_text:
                        continue
                    if self._is_terminator_text(a_text):
                        break
                    row_count += 1
                ipi_count += row_count
                notes.append(
                    {
                        "sheet": sheet_name,
                        "section_label": sec.section_label,
                        "section_kind": "ipi",
                        "row_count": row_count,
                    }
                )
                continue
            if sec.section_kind == "africa":
                row_count = 0
                for ri in range(sec.header_row + 1, sec.end_row + 1):
                    row = [
                        ws.cell(row=ri, column=c).value
                        for c in range(1, self._MAX_COL + 1)
                    ]
                    if self._is_empty_row(row):
                        continue
                    a_text = self._normalize_text(row[0])
                    if not a_text:
                        continue
                    if self._is_terminator_text(a_text):
                        break
                    row_count += 1
                africa_count += row_count
                notes.append(
                    {
                        "sheet": sheet_name,
                        "section_label": sec.section_label,
                        "section_kind": "africa",
                        "row_count": row_count,
                    }
                )
                continue

            # main_us / main_ca
            for ri in range(sec.header_row + 1, sec.end_row + 1):
                row = [
                    ws.cell(row=ri, column=c).value
                    for c in range(1, self._MAX_COL + 1)
                ]
                if self._is_empty_row(row):
                    continue
                a_text = self._normalize_text(row[0])
                if not a_text:
                    continue
                if self._is_terminator_text(a_text):
                    break
                if self._should_skip_main(row, sec.layout):
                    continue
                built, row_unseeded, row_warnings = self._build_records_from_main_row(
                    row=row,
                    row_index=ri,
                    layout=sec.layout,
                    sheet_name=sheet_name,
                    section_kind=sec.section_kind,
                    section_label=sec.section_label,
                    eff_from=eff_from,
                    eff_to=eff_to,
                    source_file=source_file,
                )
                records.extend(built)
                unseeded.update(row_unseeded)
                warnings.extend(row_warnings)

        summary = {
            "sheet_name": sheet_name,
            "total_rows": len(records),
            "ipi_addon_count": ipi_count,
            "africa_count": africa_count,
            "effective_from": eff_from,
            "effective_to": eff_to,
        }
        return records, warnings, summary, notes, unseeded

    def _build_records_from_main_row(
        self,
        *,
        row: list[Any],
        row_index: int,
        layout: dict[str, int],
        sheet_name: str,
        section_kind: str,
        section_label: str | None,
        eff_from: date | None,
        eff_to: date | None,
        source_file: str,
    ) -> tuple[list[ParsedRateRecord], set[str], list[str]]:
        records: list[ParsedRateRecord] = []
        unseeded: set[str] = set()
        warnings: list[str] = []

        origin_idx = layout.get("origin", 1) - 1
        pod_idx = layout.get("pod", 2) - 1 if "pod" in layout else None
        dest_idx = layout.get("dest", 3) - 1 if "dest" in layout else None
        coast_idx = layout.get("coast", 4) - 1 if "coast" in layout else None
        service_idx = layout.get("service", 5) - 1 if "service" in layout else None
        c20_idx = layout["c20"] - 1
        c40_idx = layout.get("c40") - 1 if "c40" in layout else None
        hc_idx = layout.get("hc") - 1 if "hc" in layout else None
        c45_idx = layout.get("c45") - 1 if "c45" in layout else None
        rad_idx = layout.get("rad") - 1 if "rad" in layout else None

        origin_raw_text = self._normalize_text(row[origin_idx])
        if not origin_raw_text:
            return records, unseeded, warnings

        # "All Base Ports" 行 — 不展开成具体 origin，跳过
        if "all base ports" in origin_raw_text.lower():
            return records, unseeded, warnings

        pod_raw = (
            self._normalize_text(row[pod_idx]) if pod_idx is not None else None
        )
        dest_raw = (
            self._normalize_text(row[dest_idx]) if dest_idx is not None else None
        )
        # destination 优先 dest 列，回退 pod
        dest_source = dest_raw or pod_raw
        dest_clean = _clean_destination(dest_source)
        if not dest_clean:
            warnings.append(
                f"{sheet_name} row {row_index}: missing destination; raw='{dest_source}'"
            )
            return records, unseeded, warnings

        # 多港 destination 提示（仅当原文含 , 或 / 且非 state 后缀）
        if dest_source:
            stripped_state = _STATE_SUFFIX_RE.sub("", dest_source).strip()
            if "/" in dest_source or ("," in stripped_state):
                warnings.append(
                    f"{sheet_name} row {row_index}: multi-port destination "
                    f"'{dest_source}' simplified to '{dest_clean}'"
                )

        coast_text = (
            self._normalize_text(row[coast_idx]) if coast_idx is not None else None
        )
        service_text = (
            self._normalize_text(row[service_idx]) if service_idx is not None else None
        )

        c20 = _nvo_safe_decimal(row[c20_idx])
        c40 = _nvo_safe_decimal(row[c40_idx]) if c40_idx is not None else None
        hc = _nvo_safe_decimal(row[hc_idx]) if hc_idx is not None else None
        c45 = _nvo_safe_decimal(row[c45_idx]) if c45_idx is not None else None
        rad_raw_text = self._raw_text(row[rad_idx]) if rad_idx is not None else None

        if c20 is None and c40 is None and hc is None and c45 is None:
            return records, unseeded, warnings

        origin_parts = _split_origins(origin_raw_text)
        if not origin_parts:
            return records, unseeded, warnings

        for part in origin_parts:
            origin_clean = _clean_origin_locode(part)
            if not origin_clean:
                continue
            # 5 字符 LOCODE 但不在 PORT_ALIAS_MAP 反向命中（无别名指向它）
            # 这种场景下 activator 走 LOCODE 精确匹配；如 seed 缺则软失败 skip
            if _LOCODE_RE.match(origin_clean):
                if origin_clean not in PORT_ALIAS_MAP.values():
                    unseeded.add(origin_clean)

            extras: dict[str, Any] = {
                "sheet_name": sheet_name,
                "row_index": row_index,
                "section_label": section_label,
                "section_kind": section_kind,
                "origin_raw": origin_raw_text,
                "pod_raw": pod_raw,
                "destination_raw": dest_raw,
                "coast": coast_text,
                "rad_raw": rad_raw_text,
                "container_20gp_raw": self._raw_text(row[c20_idx]),
                "container_40gp_raw": self._raw_text(row[c40_idx])
                if c40_idx is not None
                else None,
                "container_40hq_raw": self._raw_text(row[hc_idx])
                if hc_idx is not None
                else None,
                "container_45_raw": self._raw_text(row[c45_idx])
                if c45_idx is not None
                else None,
            }

            record = ParsedRateRecord(
                record_kind="ocean_ngb_fcl",
                carrier_name=self._CARRIER_NAME,
                origin_port_id=None,
                origin_port_name=origin_clean,
                destination_port_id=None,
                destination_port_name=dest_clean,
                service_code=service_text,
                container_20gp=c20,
                container_40gp=c40,
                container_40hq=hc,
                container_45=c45,
                currency=self._CURRENCY,
                valid_from=eff_from,
                valid_to=eff_to,
                transit_days=None,
                is_direct=True,
                remarks=None,
                source_type="excel",
                source_file=source_file,
                extras=extras,
            )
            records.append(record)

        return records, unseeded, warnings

    # ------------------------------------------------------------------
    # Hawaii sheet
    # ------------------------------------------------------------------

    def _parse_hawaii_sheet(
        self,
        ws,
        eff_from: date | None,
        eff_to: date | None,
        source_file: str,
    ) -> tuple[list[ParsedRateRecord], list[str], dict[str, Any]]:
        records: list[ParsedRateRecord] = []
        warnings: list[str] = []

        # R5/R6 双行表头写死；数据从 R7 开始
        max_row = ws.max_row or 0
        for ri in range(7, max_row + 1):
            row = [ws.cell(row=ri, column=c).value for c in range(1, 9)]
            if self._is_empty_row(row):
                continue
            a_text = self._normalize_text(row[0])
            if a_text and self._is_hawaii_terminator(a_text):
                break
            # ORIGIN 列（col 2）必须有值
            origin_text = self._normalize_text(row[1])
            if not origin_text:
                continue
            record = self._build_hawaii_record(
                row=row,
                row_index=ri,
                eff_from=eff_from,
                eff_to=eff_to,
                source_file=source_file,
                warnings=warnings,
            )
            if record is not None:
                records.append(record)

        summary = {
            "sheet_name": self._SHEET_HAWAII,
            "total_rows": len(records),
            "ipi_addon_count": 0,
            "africa_count": 0,
            "effective_from": eff_from,
            "effective_to": eff_to,
        }
        return records, warnings, summary

    def _build_hawaii_record(
        self,
        *,
        row: list[Any],
        row_index: int,
        eff_from: date | None,
        eff_to: date | None,
        source_file: str,
        warnings: list[str],
    ) -> ParsedRateRecord | None:
        country_text = self._normalize_text(row[0])
        origin_text = self._normalize_text(row[1])
        dest_text = self._normalize_text(row[2])
        if not origin_text:
            return None

        c20 = _nvo_safe_decimal(row[3])
        c40 = _nvo_safe_decimal(row[4])
        hc = _nvo_safe_decimal(row[5])
        c45 = _nvo_safe_decimal(row[6])
        if c20 is None and c40 is None and hc is None and c45 is None:
            return None

        origin_clean = _clean_origin_locode(origin_text)
        dest_clean = _clean_destination(dest_text)
        if dest_text and dest_text.strip().upper() == "CY":
            warnings.append(
                f"Hawaii row {row_index}: destination 'CY' (Container Yard) is non-specific"
            )

        extras: dict[str, Any] = {
            "sheet_name": self._SHEET_HAWAII,
            "row_index": row_index,
            "section_label": None,
            "section_kind": "hawaii",
            "country": country_text,
            "origin_raw": origin_text,
            "destination_raw": dest_text,
            "container_20gp_raw": self._raw_text(row[3]),
            "container_40gp_raw": self._raw_text(row[4]),
            "container_40hq_raw": self._raw_text(row[5]),
            "container_45_raw": self._raw_text(row[6]),
        }

        return ParsedRateRecord(
            record_kind="ocean_ngb_fcl",
            carrier_name=self._CARRIER_NAME,
            origin_port_id=None,
            origin_port_name=origin_clean,
            destination_port_id=None,
            destination_port_name=dest_clean,
            service_code=None,
            container_20gp=c20,
            container_40gp=c40,
            container_40hq=hc,
            container_45=c45,
            currency=self._CURRENCY,
            valid_from=eff_from,
            valid_to=eff_to,
            transit_days=None,
            is_direct=True,
            remarks=None,
            source_type="excel",
            source_file=source_file,
            extras=extras,
        )

    # ------------------------------------------------------------------
    # section iteration / detection
    # ------------------------------------------------------------------

    def _iter_sections(self, ws, sheet_name: str) -> Iterable[_Section]:
        max_row = ws.max_row or 0
        candidates: list[tuple[int, dict[str, int], str | None, str]] = []

        for ri in range(1, max_row + 1):
            if not self._is_section_header(ws, ri):
                continue
            layout = self._locate_columns_at_header_row(ws, ri)
            if layout is None or "c20" not in layout:
                # 不识别为有效段头（如 Hawaii sheet 走专用通道）
                continue
            label = self._extract_section_label(ws, ri)
            kind = self._classify_section(label, layout)
            candidates.append((ri, layout, label, kind))

        sections: list[_Section] = []
        for idx, (header_row, layout, label, kind) in enumerate(candidates):
            end_row = (
                candidates[idx + 1][0] - 1 if idx + 1 < len(candidates) else max_row
            )
            sections.append(
                _Section(
                    sheet_name=sheet_name,
                    header_row=header_row,
                    end_row=end_row,
                    layout=layout,
                    section_label=label,
                    section_kind=kind,
                )
            )
        return sections

    def _is_section_header(self, ws, row_index: int) -> bool:
        a = ws.cell(row=row_index, column=1).value
        b = ws.cell(row=row_index, column=2).value
        a_text = self._normalize_text(a)
        b_text = self._normalize_text(b)
        if a_text and a_text.strip().lower() == "origin":
            return True
        if a_text and "location" in a_text.lower() and b_text and "via" in b_text.lower():
            return True
        return False

    def _locate_columns_at_header_row(self, ws, header_row: int) -> dict[str, int] | None:
        layout: dict[str, int] = {}
        for ci in range(1, self._MAX_COL + 1):
            value = ws.cell(row=header_row, column=ci).value
            if value is None:
                continue
            text = str(value).strip()
            tl = text.lower()
            if not tl:
                continue
            if "origin" == tl and "origin" not in layout:
                layout["origin"] = ci
            elif (
                ("port of discharge" in tl or tl == "via" or tl == "discharge")
                and "pod" not in layout
            ):
                # 'via' 在 Canada 段是第 2 列（pod 等价）
                layout["pod"] = ci
            elif "destination" in tl and "dest" not in layout:
                layout["dest"] = ci
            elif "coast" in tl and "coast" not in layout:
                layout["coast"] = ci
            elif "canal" in tl and "coast" not in layout:
                # Africa 段表头有 'Canal' 占位
                layout["coast"] = ci
            elif "service" in tl and "service" not in layout:
                layout["service"] = ci
            elif tl in ("location", "loc.") and "origin" not in layout:
                # IPI 段把 'Location' 当 origin 占位（用于段头识别，不实际取数）
                layout["origin"] = ci
            elif (
                "20" in tl
                and "rad" not in tl
                and "40" not in tl
                and "rate" not in tl
                and "c20" not in layout
            ):
                layout["c20"] = ci
            elif "40h" in tl.replace(" ", "") and "hc" not in layout:
                layout["hc"] = ci
            elif "hc" in tl and "hc" not in layout:
                layout["hc"] = ci
            elif "40" in tl and "c40" not in layout and "rate" not in tl:
                layout["c40"] = ci
            elif "45" in tl and "c45" not in layout and "rate" not in tl:
                layout["c45"] = ci
            elif "rad" == tl and "rad" not in layout:
                layout["rad"] = ci
            elif "rate 20" in tl and "c20" not in layout:
                layout["c20"] = ci
            elif "rate 40h" in tl and "hc" not in layout:
                layout["hc"] = ci
            elif "rate 40" in tl and "c40" not in layout:
                layout["c40"] = ci
            elif "rate 45" in tl and "c45" not in layout:
                layout["c45"] = ci

        return layout if layout else None

    def _extract_section_label(self, ws, header_row: int) -> str | None:
        # 段头上一行通常是 region label
        if header_row > 1:
            prev = ws.cell(row=header_row - 1, column=1).value
            text = self._normalize_text(prev)
            if text:
                return text
        return None

    def _classify_section(self, label: str | None, layout: dict[str, int]) -> str:
        label_lower = (label or "").lower()
        if "africa" in label_lower:
            return "africa"
        if "ipi" in label_lower or "add-on" in label_lower or "add on" in label_lower:
            return "ipi"
        # IPI 段表头是 Location/Via/Rate20...，没有 dest 列
        if "dest" not in layout:
            return "ipi"
        if "canada" in label_lower:
            return "main_ca"
        return "main_us"

    # ------------------------------------------------------------------
    # row classification
    # ------------------------------------------------------------------

    def _is_empty_row(self, row: list[Any]) -> bool:
        return all(
            value is None or (isinstance(value, str) and value.strip() == "")
            for value in row
        )

    def _should_skip_main(self, row: list[Any], layout: dict[str, int]) -> bool:
        c20_idx = layout["c20"] - 1
        if c20_idx >= len(row):
            return True
        c20_value = row[c20_idx]
        if not self._is_numeric(c20_value):
            return True
        return False

    def _is_terminator_text(self, text: str) -> bool:
        tl = text.lower()
        for kw in (
            "remark",
            "commodity",
            "applicable",
            "rates are",
            "for ",
            "t&c",
            "ipi / ripi",
            "ipi/ripi",
            "fak - africa",
            " canada  / inclusive of obs",
            "canada / inclusive of obs",
        ):
            if kw in tl:
                return True
        return False

    def _is_hawaii_terminator(self, text: str) -> bool:
        tl = text.lower()
        for kw in (
            "remark",
            "commodity",
            "applicable",
            "rates are",
            "for ",
            "destination add on",
            "fak contract filing",
        ):
            if kw in tl:
                return True
        return False

    # ------------------------------------------------------------------
    # helpers
    # ------------------------------------------------------------------

    def _extract_effective_dates(
        self, ws, path: Path
    ) -> tuple[date | None, date | None]:
        year: int | None = None
        m_date: re.Match[str] | None = None
        max_scan = min(5, ws.max_row or 0)
        for ri in range(1, max_scan + 1):
            cells = [
                ws.cell(row=ri, column=c).value
                for c in range(1, self._MAX_COL + 1)
                if ws.cell(row=ri, column=c).value is not None
            ]
            text = " ".join(str(v) for v in cells)
            ym = self._YEAR_RE.search(text)
            if ym and year is None:
                year = int(ym.group(0))
            md = self._EFFECTIVE_RE.search(text)
            if md and m_date is None:
                m_date = md
        if year is None:
            ym2 = self._YEAR_RE.search(path.name)
            if ym2:
                year = int(ym2.group(0))
        if m_date is None or year is None:
            return None, None
        try:
            from_month, from_day = int(m_date.group(1)), int(m_date.group(2))
            to_month, to_day = int(m_date.group(3)), int(m_date.group(4))
            return date(year, from_month, from_day), date(year, to_month, to_day)
        except ValueError:
            return None, None

    def _extract_base_ports(self, ws) -> list[str]:
        max_scan = min(5, ws.max_row or 0)
        for ri in range(1, max_scan + 1):
            cell = ws.cell(row=ri, column=1).value
            if cell is None:
                continue
            text = str(cell)
            m = self._BASE_PORTS_RE.search(text)
            if m:
                tail = m.group(1)
                return [p.strip() for p in tail.split(",") if p.strip()]
        return []

    def _normalize_text(self, value: Any) -> str | None:
        if value is None:
            return None
        text = self._WHITESPACE_RE.sub(" ", str(value)).strip()
        return text or None

    def _raw_text(self, value: Any) -> str | None:
        if value is None:
            return None
        if isinstance(value, str):
            stripped = value.strip()
            return stripped or None
        if isinstance(value, bool):
            return str(value)
        if isinstance(value, (int, float, Decimal)):
            return str(value)
        return self._normalize_text(value)

    def _is_numeric(self, value: Any) -> bool:
        if value is None:
            return False
        if isinstance(value, bool):
            return False
        if isinstance(value, (int, float)):
            if isinstance(value, float) and value != value:
                return False
            return True
        if isinstance(value, Decimal):
            return True
        if isinstance(value, str):
            text = value.strip().lstrip("$").replace(",", "")
            if not text:
                return False
            try:
                float(text)
                return True
            except ValueError:
                return False
        return False

    def _dedupe_warnings(self, warnings: Iterable[str]) -> list[str]:
        deduped: list[str] = []
        seen: set[str] = set()
        for warning in warnings:
            if warning in seen:
                continue
            deduped.append(warning)
            seen.add(warning)
        return deduped
