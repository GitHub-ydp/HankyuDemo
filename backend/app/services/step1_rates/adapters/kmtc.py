from __future__ import annotations

from collections.abc import Iterable
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.services.rate_parser import (
    PORT_ALIAS_MAP,
    _TERMINAL_CONNECTORS,
    _TERMINAL_SUFFIXES,
    _resolve_port,
    _safe_decimal,
)
from app.services.step1_rates.entities import ParsedRateBatch, ParsedRateRecord, Step1FileType


def _clean_and_resolve_port(name_raw: str | None) -> str | None:
    """把 KMTC 的港名清洗为 alias 命中时的 UN/LOCODE，或清洗后纯英文名。

    供 activator._resolve_port 精确/ilike 匹配。返回 None 仅当输入为空。
    """
    if not name_raw or not str(name_raw).strip():
        return None
    name = str(name_raw).strip()

    # 1. 取斜杠左半（中英合并 → 英文部分）
    en = name.split("/")[0].strip() if "/" in name else name

    # 2. 去掉括号（半角全角）
    en = re.sub(r"[（(].*?[）)]", "", en).strip()

    # 3. 去掉终端后缀（PAT/TCTB/UNITHAI/BMT/SCT 等）
    for suf in _TERMINAL_SUFFIXES:
        if en.upper().endswith(suf):
            en = en[: -len(suf)].strip()
            break

    # 4. 去掉连字符及之后内容（'BKK - BMT' → 'BKK'）
    for conn in _TERMINAL_CONNECTORS:
        if conn in en:
            en = en.split(conn)[0].strip()
            break

    if not en:
        return None

    # 5. alias 查表 → UN/LOCODE（命中即返回 5 字符大写）
    locode = PORT_ALIAS_MAP.get(en.lower())
    if locode:
        return locode

    # 6. fallback：返回清洗后纯英文（让 activator 走 ilike）
    return en


class KmtcAdapter:
    """Step1 KMTC parser for the Sea专刊 workbook (上海发, USD)."""

    key: str = "kmtc"
    file_type: Step1FileType = Step1FileType.ocean
    priority: int = 15

    _SHEET_NAME: str = "KMTC-专刊"
    _DEFAULT_ORIGIN_NAME: str = "CNSHA"
    _CARRIER_NAME: str = "KMTC"
    _CURRENCY: str = "USD"
    _MAX_COL: int = 15
    _PARSER_VERSION: str = "kmtc_v1"

    _HEADER_SCAN_LIMIT: int = 10

    _DETECT_NAME_KEYWORDS_LOWER: tuple[str, ...] = ("kmtc",)
    _DETECT_NAME_KEYWORDS_RAW: tuple[str, ...] = ("高丽", "高麗")
    _DETECT_SHEET_KEYWORDS: tuple[str, ...] = ("KMTC", "高丽", "高麗")

    _WHITESPACE_RE = re.compile(r"\s+")
    _LSS_REGION_RE = re.compile(r"LSS[：:]\s*USD?\s*(\d+\s*/\s*\d+)", re.IGNORECASE)
    _TRANSIT_DAYS_RE = re.compile(r"(\d+)\s*天")

    # ------------------------------------------------------------------
    # detect / parse
    # ------------------------------------------------------------------

    def detect(self, path: Path, *, file_type_hint: Step1FileType | None = None) -> bool:
        if file_type_hint == self.file_type:
            return True
        name_raw = path.name
        name_lower = name_raw.lower()
        if any(kw in name_lower for kw in self._DETECT_NAME_KEYWORDS_LOWER):
            return True
        if any(kw in name_raw for kw in self._DETECT_NAME_KEYWORDS_RAW):
            return True
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            try:
                sheetnames = list(wb.sheetnames)
            finally:
                wb.close()
        except Exception:
            return False
        for sheet_name in sheetnames:
            for kw in self._DETECT_SHEET_KEYWORDS:
                if kw in sheet_name:
                    return True
        return False

    def parse(self, path: Path, db: Session | None = None) -> ParsedRateBatch:
        warnings: list[str] = []
        records: list[ParsedRateRecord] = []
        region_lss_defaults: dict[str, str] = {}

        wb = load_workbook(path, data_only=True)
        if self._SHEET_NAME not in wb.sheetnames:
            warnings.append(
                f"KMTC workbook missing expected sheet '{self._SHEET_NAME}'; nothing to parse"
            )
            return self._empty_batch(path, warnings, region_lss_defaults)

        ws = wb[self._SHEET_NAME]
        main_row, sub_row = self._locate_headers(ws)
        if main_row is None or sub_row is None:
            warnings.append("KMTC workbook: header rows not located in first 10 rows")
            return self._empty_batch(path, warnings, region_lss_defaults)

        layout = self._build_column_layout(ws, main_row, sub_row)
        if layout is None:
            warnings.append("KMTC workbook: column layout could not be derived from header rows")
            return self._empty_batch(path, warnings, region_lss_defaults)

        max_row = ws.max_row or 0
        for row_index in range(sub_row + 1, max_row + 1):
            row = [ws.cell(row=row_index, column=c).value for c in range(1, self._MAX_COL + 1)]
            if self._is_empty_row(row):
                continue
            if self._is_region_header(row):
                region_text = self._normalize_text(row[0])
                if region_text:
                    extracted = self._extract_region_lss(region_text)
                    region_key = f"R{row_index}_{region_text[:20]}"
                    if extracted:
                        region_lss_defaults[region_key] = "; ".join(extracted)
                    else:
                        region_lss_defaults[region_key] = ""
                continue
            if self._is_remark_or_note(row, layout):
                continue
            record, row_warnings = self._build_record(
                row=row,
                row_index=row_index,
                layout=layout,
                source_file=path.name,
                db=db,
            )
            warnings.extend(row_warnings)
            if record is not None:
                records.append(record)

        effective_from = min(
            (r.valid_from for r in records if r.valid_from is not None),
            default=None,
        )
        if effective_from is None:
            warnings.append(
                "KMTC workbook: no valid effective dates extracted from L column"
            )

        sheet_summary = {
            "sheet_name": self._SHEET_NAME,
            "total_rows": len(records),
            "effective_from": effective_from,
            "effective_to": None,
        }
        kind_distribution = {"ocean_ngb_fcl": len(records)}
        metadata = {
            "file_name": path.name,
            "source_type": "excel",
            "carrier_code": self._CARRIER_NAME,
            "parser_version": self._PARSER_VERSION,
            "adapter_key": self.key,
            "kmtc_origin_assumption": "default origin = CNSHA (Shanghai) for all rows",
            "sheets": [sheet_summary],
            "region_lss_defaults": region_lss_defaults,
            "record_kind_distribution": kind_distribution,
        }

        return ParsedRateBatch(
            file_type=self.file_type,
            source_file=path.name,
            effective_from=effective_from,
            effective_to=None,
            records=records,
            warnings=self._dedupe_warnings(warnings),
            adapter_key=self.key,
            metadata=metadata,
        )

    # ------------------------------------------------------------------
    # header / layout
    # ------------------------------------------------------------------

    def _locate_headers(self, ws) -> tuple[int | None, int | None]:
        max_row = min(self._HEADER_SCAN_LIMIT, ws.max_row or 0)
        main_row: int | None = None
        sub_row: int | None = None
        for row_index in range(1, max_row + 1):
            row_text_lower = " ".join(
                str(ws.cell(row_index, c).value).lower()
                for c in range(1, self._MAX_COL + 1)
                if ws.cell(row_index, c).value is not None
            )
            if main_row is None:
                if "港口" in row_text_lower or (
                    "o/f" in row_text_lower and ("baf" in row_text_lower or "lss" in row_text_lower)
                ):
                    main_row = row_index
                    continue
            if main_row is not None and row_index > main_row:
                if any(
                    "20" in str(ws.cell(row_index, c).value or "")
                    or "40" in str(ws.cell(row_index, c).value or "")
                    for c in range(1, self._MAX_COL + 1)
                ):
                    sub_row = row_index
                    break
        return main_row, sub_row

    def _build_column_layout(self, ws, main_row: int, sub_row: int) -> dict[str, int] | None:
        main_cells = [ws.cell(main_row, c).value for c in range(1, self._MAX_COL + 1)]
        sub_cells = [ws.cell(sub_row, c).value for c in range(1, self._MAX_COL + 1)]

        port_col: int | None = None
        schedule_col: int | None = None
        company_col: int | None = None
        route_col: int | None = None
        date_col: int | None = None
        remark_col: int | None = None

        for ci, cell in enumerate(main_cells, start=1):
            if cell is None:
                continue
            text = str(cell).strip()
            tl = text.lower()
            if port_col is None and ("港口" in text or "port" in tl):
                port_col = ci
            elif schedule_col is None and ("船期" in text or "schedule" in tl):
                schedule_col = ci
            elif company_col is None and ("船公司" in text or "carrier" in tl):
                company_col = ci
            elif route_col is None and ("航线" in text or "route" in tl or "service" in tl):
                route_col = ci
            elif date_col is None and ("生效" in text or "effective" in tl):
                date_col = ci
            elif remark_col is None and ("備考" in text or "备注" in text or "remark" in tl):
                remark_col = ci

        slots_20: list[int] = []
        for ci, cell in enumerate(sub_cells, start=1):
            if cell is None:
                continue
            text_upper = str(cell).strip().upper()
            if "20" in text_upper:
                slots_20.append(ci)

        if not slots_20:
            return None

        of_20 = slots_20[0] if len(slots_20) >= 1 else None
        baf_20 = slots_20[1] if len(slots_20) >= 2 else None
        lss_20 = slots_20[2] if len(slots_20) >= 3 else None

        of_40 = of_20 + 1 if of_20 is not None else None
        of_hq = of_20 + 2 if of_20 is not None else None
        baf_40 = baf_20 + 1 if baf_20 is not None else None
        lss_40 = lss_20 + 1 if lss_20 is not None else None

        if of_20 is None or of_40 is None or of_hq is None:
            return None

        if port_col is None:
            port_col = 1

        layout: dict[str, int] = {
            "port": port_col,
            "of_20": of_20,
            "of_40": of_40,
            "of_hq": of_hq,
        }
        if schedule_col is not None:
            layout["schedule"] = schedule_col
        if company_col is not None:
            layout["company"] = company_col
        if route_col is not None:
            layout["route"] = route_col
        if baf_20 is not None:
            layout["baf_20"] = baf_20
        if baf_40 is not None:
            layout["baf_40"] = baf_40
        if lss_20 is not None:
            layout["lss_20"] = lss_20
        if lss_40 is not None:
            layout["lss_40"] = lss_40
        if date_col is not None:
            layout["date"] = date_col
        if remark_col is not None:
            layout["remark"] = remark_col
        return layout

    # ------------------------------------------------------------------
    # row classification
    # ------------------------------------------------------------------

    def _is_empty_row(self, row: list[Any]) -> bool:
        return all(value is None or (isinstance(value, str) and value.strip() == "") for value in row)

    def _is_region_header(self, row: list[Any]) -> bool:
        a_text = self._normalize_text(row[0])
        if not a_text:
            return False
        if "航线" not in a_text:
            return False
        if "：" not in a_text and ":" not in a_text:
            return False
        for value in row[1:]:
            if value is not None and not (isinstance(value, str) and value.strip() == ""):
                return False
        return True

    def _is_remark_or_note(self, row: list[Any], layout: dict[str, int]) -> bool:
        a_text = self._normalize_text(row[0])
        if not a_text:
            return False
        of_20_idx = layout["of_20"] - 1
        of_value = row[of_20_idx]
        if not self._is_numeric(of_value):
            return True
        return False

    def _extract_region_lss(self, text: str) -> list[str]:
        return [m.group(1).replace(" ", "") for m in self._LSS_REGION_RE.finditer(text)]

    # ------------------------------------------------------------------
    # record building
    # ------------------------------------------------------------------

    def _build_record(
        self,
        *,
        row: list[Any],
        row_index: int,
        layout: dict[str, int],
        source_file: str,
        db: Session | None,
    ) -> tuple[ParsedRateRecord | None, list[str]]:
        warnings: list[str] = []
        port_idx = layout["port"] - 1
        port_raw_value = row[port_idx]
        port_name_raw = self._normalize_text(port_raw_value)
        if not port_name_raw:
            warnings.append(f"KMTC row {row_index}: missing destination port name")
            return None, warnings

        if db is not None:
            resolved_port = _resolve_port(port_name_raw, db)
            if resolved_port is None:
                warnings.append(f"KMTC row {row_index}: 无法识别港口 '{port_name_raw}'")
                return None, warnings

        of_20_idx = layout["of_20"] - 1
        of_40_idx = layout["of_40"] - 1
        of_hq_idx = layout["of_hq"] - 1
        baf_20_idx = layout["baf_20"] - 1 if "baf_20" in layout else None
        baf_40_idx = layout["baf_40"] - 1 if "baf_40" in layout else None
        lss_20_idx = layout["lss_20"] - 1 if "lss_20" in layout else None
        lss_40_idx = layout["lss_40"] - 1 if "lss_40" in layout else None
        date_idx = layout["date"] - 1 if "date" in layout else None
        remark_idx = layout["remark"] - 1 if "remark" in layout else None
        schedule_idx = layout["schedule"] - 1 if "schedule" in layout else None
        company_idx = layout["company"] - 1 if "company" in layout else None
        route_idx = layout["route"] - 1 if "route" in layout else None

        container_20gp = _safe_decimal(row[of_20_idx])
        container_40gp = _safe_decimal(row[of_40_idx])
        container_40hq = _safe_decimal(row[of_hq_idx])
        baf_20 = _safe_decimal(row[baf_20_idx]) if baf_20_idx is not None else None
        baf_40 = _safe_decimal(row[baf_40_idx]) if baf_40_idx is not None else None
        lss_20 = _safe_decimal(row[lss_20_idx]) if lss_20_idx is not None else None
        lss_40 = _safe_decimal(row[lss_40_idx]) if lss_40_idx is not None else None

        date_value: Any = row[date_idx] if date_idx is not None else None
        valid_from = self._to_date(date_value)

        remark_value: Any = row[remark_idx] if remark_idx is not None else None
        remark_text = self._normalize_text(remark_value)
        transit_days, is_direct = self._parse_transit_days(remark_text)

        extras: dict[str, Any] = {
            "sheet_name": self._SHEET_NAME,
            "row_index": row_index,
            "schedule_text": self._normalize_text(row[schedule_idx]) if schedule_idx is not None else None,
            "shipping_line_text": self._normalize_text(row[company_idx]) if company_idx is not None else None,
            "route_code": self._normalize_text(row[route_idx]) if route_idx is not None else None,
            "container_20gp_raw": self._raw_text(row[of_20_idx]),
            "container_40gp_raw": self._raw_text(row[of_40_idx]),
            "container_40hq_raw": self._raw_text(row[of_hq_idx]),
            "baf_20_raw": self._raw_text(row[baf_20_idx]) if baf_20_idx is not None else None,
            "baf_40_raw": self._raw_text(row[baf_40_idx]) if baf_40_idx is not None else None,
            "lss_20_raw": self._raw_text(row[lss_20_idx]) if lss_20_idx is not None else None,
            "lss_40_raw": self._raw_text(row[lss_40_idx]) if lss_40_idx is not None else None,
            "valid_from_raw": self._raw_date_text(date_value),
            "remark_raw": remark_text,
            "destination_port_raw": port_name_raw,
        }

        destination_resolved = _clean_and_resolve_port(port_name_raw)
        record = ParsedRateRecord(
            record_kind="ocean_ngb_fcl",
            carrier_name=self._CARRIER_NAME,
            origin_port_id=None,
            origin_port_name=self._DEFAULT_ORIGIN_NAME,
            destination_port_id=None,
            destination_port_name=destination_resolved,
            container_20gp=container_20gp,
            container_40gp=container_40gp,
            container_40hq=container_40hq,
            baf_20=baf_20,
            baf_40=baf_40,
            lss_20=lss_20,
            lss_40=lss_40,
            currency=self._CURRENCY,
            valid_from=valid_from,
            valid_to=None,
            transit_days=transit_days,
            is_direct=is_direct,
            remarks=remark_text,
            source_type="excel",
            source_file=source_file,
            extras=extras,
        )
        return record, warnings

    # ------------------------------------------------------------------
    # helpers
    # ------------------------------------------------------------------

    def _to_date(self, value: Any) -> date | None:
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, str):
            text = value.strip()
            if not text:
                return None
            for fmt in ("%Y/%m/%d", "%Y-%m-%d"):
                try:
                    return datetime.strptime(text, fmt).date()
                except ValueError:
                    continue
        return None

    def _parse_transit_days(self, remark: str | None) -> tuple[int | None, bool]:
        if not remark:
            return None, True
        is_direct = "直达" in remark
        match = self._TRANSIT_DAYS_RE.search(remark)
        days: int | None = int(match.group(1)) if match else None
        if not is_direct and "中转" in remark:
            is_direct = False
        return days, is_direct

    def _normalize_text(self, value: Any) -> str | None:
        if value is None:
            return None
        text = self._WHITESPACE_RE.sub(" ", str(value)).strip()
        return text or None

    def _raw_text(self, value: Any) -> str | None:
        if value is None:
            return None
        if isinstance(value, str):
            stripped = value.strip()
            return stripped or None
        if isinstance(value, bool):
            return str(value)
        if isinstance(value, (int, float, Decimal)):
            return str(value)
        return self._normalize_text(value)

    def _raw_date_text(self, value: Any) -> str | None:
        if value is None:
            return None
        if isinstance(value, (datetime, date)):
            return value.isoformat()
        return self._normalize_text(value)

    def _is_numeric(self, value: Any) -> bool:
        if value is None:
            return False
        if isinstance(value, bool):
            return False
        if isinstance(value, (int, float)):
            if isinstance(value, float) and value != value:  # NaN
                return False
            return True
        if isinstance(value, Decimal):
            return True
        if isinstance(value, str):
            text = value.strip().replace(",", "")
            if not text:
                return False
            try:
                float(text)
                return True
            except ValueError:
                return False
        return False

    def _dedupe_warnings(self, warnings: Iterable[str]) -> list[str]:
        deduped: list[str] = []
        seen: set[str] = set()
        for warning in warnings:
            if warning in seen:
                continue
            deduped.append(warning)
            seen.add(warning)
        return deduped

    def _empty_batch(
        self,
        path: Path,
        warnings: list[str],
        region_lss_defaults: dict[str, str],
    ) -> ParsedRateBatch:
        sheet_summary = {
            "sheet_name": self._SHEET_NAME,
            "total_rows": 0,
            "effective_from": None,
            "effective_to": None,
        }
        return ParsedRateBatch(
            file_type=self.file_type,
            source_file=path.name,
            effective_from=None,
            effective_to=None,
            records=[],
            warnings=self._dedupe_warnings(warnings),
            adapter_key=self.key,
            metadata={
                "file_name": path.name,
                "source_type": "excel",
                "carrier_code": self._CARRIER_NAME,
                "parser_version": self._PARSER_VERSION,
                "adapter_key": self.key,
                "kmtc_origin_assumption": "default origin = CNSHA (Shanghai) for all rows",
                "sheets": [sheet_summary],
                "region_lss_defaults": region_lss_defaults,
                "record_kind_distribution": {"ocean_ngb_fcl": 0},
            },
        )
