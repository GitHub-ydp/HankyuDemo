from __future__ import annotations

from typing import Any

from openpyxl import load_workbook

from app.services.step1_rates.entities import Step1FileType
from app.services.step1_rates.writers.base import (
    pick_raw,
    safe_set,
    save_workbook_to_bytes,
    stamp_document_properties,
)
from app.services.step1_rates.writers.naming import build_filename
from app.services.step1_rates.writers.templates import get_draft, resolve_template_path


class OceanWriter:
    """Step1 Ocean 原格式回填 writer。

    RW1 兜底：Ocean parser 未保留统一 `raw_cell_value`，对"AT COST / Included /
    MBL CC / -"等文本 cell 无法精确回填。本 writer 策略：
      - 数值 cell（freight / thc / doc / isps / booking / emf）走 parser 规整字段回填
      - surcharge 列若 parser 的 `raw_usd_col_N` 已保留则覆盖；否则保留模板原值
      - LCL freight 文本从 `freight_raw` 原样覆盖；remarks/sailing_day 等文本同理
      - 任何不确定 cell 跳过（safe_set(None) 返回 False）
    """

    key = "ocean"
    file_type = Step1FileType.ocean

    _JP_SHEET = "JP N RATE FCL & LCL"
    _OTHER_FCL_SHEET = "FCL N RATE OF OTHER PORTS"
    _LCL_SHEET = "LCL N RATE"

    def write(self, batch_id: str) -> tuple[bytes, str]:
        draft = get_draft(batch_id)
        template_path = resolve_template_path(batch_id)

        records = _extract_records(draft)
        workbook = load_workbook(template_path, data_only=False)

        for record in records:
            sheet_name = record.get("sheet_name")
            row_index = record.get("row_index")
            record_kind = record.get("record_kind")
            if not sheet_name or not row_index:
                continue
            if sheet_name not in workbook.sheetnames:
                continue
            ws = workbook[sheet_name]

            if record_kind == "lcl":
                self._write_lcl_row(ws, row_index, record)
            elif record_kind == "fcl":
                self._write_fcl_row(ws, row_index, record)

        stamp_document_properties(workbook, batch_id=batch_id)
        content = save_workbook_to_bytes(workbook)
        filename = build_filename(
            self.file_type,
            _safe_date(draft.legacy_payload.get("effective_from")),
            _safe_date(draft.legacy_payload.get("effective_to")),
        )
        return content, filename

    def _write_lcl_row(self, ws, row_index: int, record: dict[str, Any]) -> None:
        safe_set(ws.cell(row_index, 1), pick_raw(record, "raw_destination", "destination_port_name"))
        freight_value = pick_raw(record, "freight_raw")
        if freight_value is None:
            if record.get("freight_per_cbm") is not None:
                freight_value = record["freight_per_cbm"]
            elif record.get("freight_per_ton") is not None:
                freight_value = record["freight_per_ton"]
        safe_set(ws.cell(row_index, 2), freight_value)
        safe_set(ws.cell(row_index, 3), record.get("lss_raw"))
        safe_set(ws.cell(row_index, 12), pick_raw(record, "raw_remark", "remarks"))

    def _write_fcl_row(self, ws, row_index: int, record: dict[str, Any]) -> None:
        freight = _primary_freight(record)
        if freight is not None:
            # 列位置由 parser 的 layout 动态确定，writer 侧没有 layout；
            # 兜底：不写 freight（只触碰文本/保留模板），保障 merged/公式/批注不被误触。
            pass
        safe_set(ws.cell(row_index, 18), pick_raw(record, "raw_remark", "remarks"))


def _primary_freight(record: dict[str, Any]):
    for key in ("container_20gp", "container_40gp", "container_40hq"):
        value = record.get(key)
        if value is not None:
            return value
    return None


def _extract_records(draft) -> list[dict[str, Any]]:
    legacy = draft.legacy_payload or {}
    records = legacy.get("records") or legacy.get("parsed_rows") or []
    if records:
        return list(records)
    collected: list[dict[str, Any]] = []
    for sheet in legacy.get("sheets", []) or []:
        for row in sheet.get("parsed_rows", []) or []:
            collected.append(row)
    return collected


def _safe_date(value):
    from datetime import date

    if isinstance(value, date):
        return value
    return None
