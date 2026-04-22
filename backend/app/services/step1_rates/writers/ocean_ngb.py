from __future__ import annotations

from typing import Any

from openpyxl import load_workbook

from app.services.step1_rates.entities import Step1FileType
from app.services.step1_rates.writers.base import (
    safe_set,
    save_workbook_to_bytes,
    stamp_document_properties,
)
from app.services.step1_rates.writers.naming import build_filename
from app.services.step1_rates.writers.templates import get_draft, resolve_template_path


class OceanNgbWriter:
    """Step1 Ocean-NGB 原格式回填 writer。

    Q-W4 纪律：
      - 只触碰 `Rate` sheet 数据 cell；`sample` / `Shipping line name` 不写
      - 每个写入走 `safe_set`，`is_formula_cell` 守卫让 1687 个 Lv.2/Lv.3 公式保留
      - 当前 Ocean-NGB parser 尚是 stub（records 为空），writer 事实上不写任何 cell，
        仅走 load → save 完成属性盖章；一旦 parser 交付，下面按 extras.column_index_map 驱动。
    """

    key = "ocean_ngb"
    file_type = Step1FileType.ocean_ngb

    _DATA_SHEET = "Rate"
    _SKIP_SHEETS = frozenset({"sample", "Shipping line name"})

    def write(self, batch_id: str) -> tuple[bytes, str]:
        draft = get_draft(batch_id)
        template_path = resolve_template_path(batch_id)

        records = _extract_records(draft)
        workbook = load_workbook(template_path, data_only=False)

        if self._DATA_SHEET in workbook.sheetnames:
            ws = workbook[self._DATA_SHEET]
            for record in records:
                sheet_name = record.get("sheet_name")
                if sheet_name in self._SKIP_SHEETS:
                    continue
                if sheet_name and sheet_name != self._DATA_SHEET:
                    continue
                row_index = record.get("row_index")
                if not row_index:
                    continue
                column_map = record.get("column_index_map") or {}
                for col_idx, value in column_map.items():
                    try:
                        col = int(col_idx)
                    except (TypeError, ValueError):
                        continue
                    safe_set(ws.cell(row_index, col), value)

        stamp_document_properties(workbook, batch_id=batch_id)
        content = save_workbook_to_bytes(workbook)
        filename = build_filename(
            self.file_type,
            _safe_date(draft.legacy_payload.get("effective_from")),
            _safe_date(draft.legacy_payload.get("effective_to")),
        )
        return content, filename


def _extract_records(draft) -> list[dict[str, Any]]:
    legacy = draft.legacy_payload or {}
    records = legacy.get("records") or legacy.get("parsed_rows") or []
    if records:
        return list(records)
    collected: list[dict[str, Any]] = []
    for sheet in legacy.get("sheets", []) or []:
        for row in sheet.get("parsed_rows", []) or []:
            collected.append(row)
    return collected


def _safe_date(value):
    from datetime import date

    if isinstance(value, date):
        return value
    return None
