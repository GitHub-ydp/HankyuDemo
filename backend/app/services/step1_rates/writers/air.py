from __future__ import annotations

from datetime import datetime
from typing import Any

from openpyxl import load_workbook

from app.services.step1_rates.entities import Step1FileType
from app.services.step1_rates.writers.base import (
    pick_raw,
    safe_set,
    save_workbook_to_bytes,
    stamp_document_properties,
)
from app.services.step1_rates.writers.naming import build_filename
from app.services.step1_rates.writers.templates import get_draft, resolve_template_path


class AirWriter:
    """Step1 Air 原格式回填 writer。"""

    key = "air"
    file_type = Step1FileType.air

    def write(self, batch_id: str) -> tuple[bytes, str]:
        draft = get_draft(batch_id)
        template_path = resolve_template_path(batch_id)

        records = _extract_records(draft)
        workbook = load_workbook(template_path, data_only=False)

        effective_week_start = _pick_batch_week_start(draft, records)

        weekly_records = [r for r in records if r.get("record_kind") == "air_weekly"]
        surcharge_records = [r for r in records if r.get("record_kind") == "air_surcharge"]

        self._write_weekly(workbook, weekly_records, effective_week_start)
        self._write_surcharges(workbook, surcharge_records)

        stamp_document_properties(workbook, batch_id=batch_id)

        content = save_workbook_to_bytes(workbook)
        filename = build_filename(
            self.file_type,
            _safe_date(draft.legacy_payload.get("effective_from")),
            _safe_date(draft.legacy_payload.get("effective_to")),
        )
        return content, filename

    def _write_weekly(
        self,
        workbook,
        records: list[dict[str, Any]],
        current_week_start,
    ) -> None:
        for record in records:
            sheet_name = record.get("sheet_name")
            row_index = record.get("row_index")
            if not sheet_name or not row_index:
                continue
            if sheet_name not in workbook.sheetnames:
                continue
            # Q-W2 默认：只写当前周；上周 sheet 数据区保持模板原样
            if (
                current_week_start is not None
                and record.get("effective_week_start") is not None
                and record["effective_week_start"] != current_week_start
            ):
                continue
            ws = workbook[sheet_name]

            safe_set(
                ws.cell(row_index, 1),
                pick_raw(record, "raw_destination", "destination_port_name"),
            )
            safe_set(
                ws.cell(row_index, 2),
                pick_raw(record, "raw_service", "service_desc"),
            )
            for day_no in range(1, 8):
                col = 2 + day_no
                raw_key = f"price_day{day_no}_raw"
                numeric_key = f"price_day{day_no}"
                value = pick_raw(record, raw_key, numeric_key)
                safe_set(ws.cell(row_index, col), value)
            safe_set(
                ws.cell(row_index, 10),
                pick_raw(record, "raw_remark", "remarks"),
            )

    def _write_surcharges(
        self,
        workbook,
        records: list[dict[str, Any]],
    ) -> None:
        if "Surcharges" not in workbook.sheetnames:
            return
        ws = workbook["Surcharges"]
        for record in records:
            row_index = record.get("row_index")
            if not row_index:
                continue
            # AREA(B) / FROM(C) 是合并区 anchor，跳过不重写（§5.1 / RW8）
            safe_set(ws.cell(row_index, 4), record.get("airline_code_raw"))

            effective_raw = record.get("effective_date_raw")
            if isinstance(effective_raw, (datetime,)):
                safe_set(ws.cell(row_index, 5), effective_raw)
            elif record.get("valid_from") is not None:
                safe_set(ws.cell(row_index, 5), record["valid_from"])
            else:
                safe_set(ws.cell(row_index, 5), effective_raw)

            fee_slots = [
                (6, "myc_min_value", "myc_min_is_dash"),
                (7, "myc_fee_per_kg", "myc_fee_is_dash"),
                (8, "msc_min_value", "msc_min_is_dash"),
                (9, "msc_fee_per_kg", "msc_fee_is_dash"),
            ]
            for col, value_key, dash_key in fee_slots:
                if record.get(dash_key):
                    safe_set(ws.cell(row_index, col), "-")
                else:
                    safe_set(ws.cell(row_index, col), record.get(value_key))

            safe_set(ws.cell(row_index, 10), record.get("destination_scope"))
            safe_set(
                ws.cell(row_index, 11),
                pick_raw(record, "raw_remark", "remarks"),
            )


def _extract_records(draft) -> list[dict[str, Any]]:
    """从 draft.legacy_payload 里取完整 records（含 extras）。"""
    legacy = draft.legacy_payload or {}
    records = legacy.get("records") or legacy.get("parsed_rows") or []
    if records:
        return list(records)
    collected: list[dict[str, Any]] = []
    for sheet in legacy.get("sheets", []) or []:
        for row in sheet.get("parsed_rows", []) or []:
            collected.append(row)
    return collected


def _pick_batch_week_start(draft, records: list[dict[str, Any]]):
    weekly = [r for r in records if r.get("record_kind") == "air_weekly"]
    # 默认当前周 = legacy effective_from（parser 已按当前周聚合）
    effective_from = _safe_date(draft.legacy_payload.get("effective_from"))
    if effective_from:
        return effective_from
    # fallback：取记录里最大的 effective_week_start（假定新周更靠后）
    starts = [
        r.get("effective_week_start")
        for r in weekly
        if r.get("effective_week_start") is not None
    ]
    return max(starts) if starts else None


def _safe_date(value):
    from datetime import date

    if isinstance(value, date):
        return value
    return None
