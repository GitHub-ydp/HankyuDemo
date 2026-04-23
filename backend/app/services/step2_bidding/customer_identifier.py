"""Step2 T-B8 customer_identifier 实现。

业务依据：docs/Step2_入札対応_T-B8_customer_identifier_业务需求_20260423.md
v1.0 双分类 only：customer_a vs unknown。
禁止为 Customer B / E / Nitori / 维度 A / 维度 C 留 stub。
"""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

# -------- 模块级常量 --------
_CUSTOMER_A = "customer_a"
_UNKNOWN = "unknown"

_SHEET_NAME_CUSTOMER_A = "見積りシート"      # 维度 B 等值匹配
_HEADER_ORIGIN = "発地"                       # 维度 D - B 列等值
_HEADER_DEST_KEYWORD = "着地"                 # 维度 D - C 列包含
_HEADER_CARRIER = "主要キャリアとルート"      # 维度 D - G 列等值

_COL_B = 2
_COL_C = 3
_COL_G = 7

_HEADER_SCAN_FIRST_ROW = 1
_HEADER_SCAN_LAST_ROW = 10                    # 业务需求 §2-D：1~10 行浮动


@dataclass(frozen=True, slots=True)
class IdentifierResult:
    """customer_identifier 唯一输出契约。字段语义见业务需求 §5。"""
    matched_customer: str
    matched_dimensions: tuple[str, ...]
    source: str
    confidence: str
    unmatched_reason: str | None
    warnings: tuple[str, ...] = field(default=())


def identify(xlsx_path: Path) -> IdentifierResult:
    """v1.0 双分类入口。详见架构任务单 §5。"""
    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    except (OSError, InvalidFileException, BadZipFile, KeyError) as e:
        return _make_unknown(
            reason="文件无法打开",
            warnings=(f"WBOPEN_FAIL: {type(e).__name__}",),
        )

    try:
        sheetnames = list(wb.sheetnames)
        if not sheetnames:
            return _make_unknown(reason="工作簿无 sheet", warnings=("EMPTY_WB",))

        normalized_names = [_normalize_sheet_name(n) for n in sheetnames]
        warnings_acc: list[str] = []

        # 维度 B：单 sheet 且 sheet 名等值
        dim_b = (
            len(sheetnames) == 1
            and normalized_names[0] == _SHEET_NAME_CUSTOMER_A
        )
        if (
            not dim_b
            and len(sheetnames) > 1
            and _SHEET_NAME_CUSTOMER_A in normalized_names
        ):
            warnings_acc.append(
                f"MULTI_SHEET: 工作簿含 {len(sheetnames)} 个 sheet，"
                f"含 '見積りシート' 但非单 sheet 模板"
            )

        # 维度 D：行 1~10 浮动扫描
        dim_d = False
        if dim_b:
            target_sheets = [sheetnames[0]]
        else:
            target_sheets = sheetnames
        dim_d_summary_parts: list[str] = []
        for sn in target_sheets:
            ws = wb[sn]
            last_scan = min(ws.max_row or 0, _HEADER_SCAN_LAST_ROW)
            hit_row = _scan_dim_d(ws, last_scan)
            dim_d_summary_parts.append(
                f"sheet={sn!r} 扫描行 1-{last_scan} {'命中' if hit_row else '未命中'}"
            )
            if hit_row:
                dim_d = True
                break

        # 组装结果
        dims: tuple[str, ...] = tuple(
            d for d, hit in (("B", dim_b), ("D", dim_d)) if hit
        )
        if dims:
            confidence = "high" if len(dims) == 2 else "medium"
            if dim_b and not dim_d:
                warnings_acc.append(
                    "HEADER_MISMATCH: sheet 名命中但表头三关键字未全中，解析可能失败"
                )
            if dim_d and not dim_b:
                warnings_acc.append(
                    f"SHEET_NAME_VARIANT: 表头命中但 sheet 名异常: {sheetnames!r}"
                )
            return IdentifierResult(
                matched_customer=_CUSTOMER_A,
                matched_dimensions=dims,
                source="auto",
                confidence=confidence,
                unmatched_reason=None,
                warnings=tuple(warnings_acc),
            )

        reason = _build_unmatched_reason(sheetnames, "; ".join(dim_d_summary_parts))
        return IdentifierResult(
            matched_customer=_UNKNOWN,
            matched_dimensions=(),
            source="auto",
            confidence="low",
            unmatched_reason=reason,
            warnings=tuple(warnings_acc),
        )
    finally:
        wb.close()


# -------- 私有辅助 --------


def _normalize_sheet_name(name: str | None) -> str:
    """业务需求 §4：strip 前后空白；不做大小写折叠；不做半角全角折叠。"""
    if name is None:
        return ""
    return str(name).strip()


def _normalize_header_cell(value: object) -> str:
    """业务需求 §2-D normalize 规则：strip 前后空白 + 去全角空格。

    不剥圆括号注解；不做大小写折叠 / 半角全角折叠。
    """
    if value is None:
        return ""
    return str(value).strip().replace("　", "")


def _row_matches_dim_d(ws, row_idx: int) -> bool:
    """同一行：B 列==発地 ∧ C 列⊇着地 ∧ G 列==主要キャリアとルート。"""
    b_val = _normalize_header_cell(ws.cell(row_idx, _COL_B).value)
    c_val = _normalize_header_cell(ws.cell(row_idx, _COL_C).value)
    g_val = _normalize_header_cell(ws.cell(row_idx, _COL_G).value)
    return (
        b_val == _HEADER_ORIGIN
        and _HEADER_DEST_KEYWORD in c_val
        and g_val == _HEADER_CARRIER
    )


def _scan_dim_d(ws, last_scan: int) -> int | None:
    """在指定 sheet 行 1..last_scan 内扫描，返回首个命中行号；未命中返回 None。"""
    for r in range(_HEADER_SCAN_FIRST_ROW, last_scan + 1):
        if _row_matches_dim_d(ws, r):
            return r
    return None


def _make_unknown(
    *, reason: str, warnings: tuple[str, ...] = ()
) -> IdentifierResult:
    """统一构造 unknown 结果。"""
    return IdentifierResult(
        matched_customer=_UNKNOWN,
        matched_dimensions=(),
        source="auto",
        confidence="low",
        unmatched_reason=reason,
        warnings=warnings,
    )


def _build_unmatched_reason(sheetnames: list[str], dim_d_summary: str) -> str:
    """unknown 时给营业看的业务可读原因。"""
    if len(sheetnames) == 1:
        sheet_part = f"sheet 名 {sheetnames[0]!r}，非 Customer A 模板"
    else:
        sheet_part = f"工作簿含 {len(sheetnames)} 个 sheet：{sheetnames!r}"
    return (
        f"{sheet_part}；"
        f"表头行 1-{_HEADER_SCAN_LAST_ROW} 未找到 "
        f"({_HEADER_ORIGIN}, {_HEADER_DEST_KEYWORD}, {_HEADER_CARRIER}) 三关键字组合"
        f"（{dim_d_summary}）"
    )
