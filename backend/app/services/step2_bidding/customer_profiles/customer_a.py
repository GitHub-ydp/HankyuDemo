"""Customer A（ミマキエンジニアリング）PKG 解析 + fill 实现。

黄金样本：资料/2026.04.02/Customer A (Air)/Customer A (Air)/2-①.xlsx
算法依据：架构任务单 §6.1-6.3（parse）、T-B7 架构任务单（fill）。
"""
from __future__ import annotations

import re
import shutil
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_CEILING
from pathlib import Path
from typing import Callable

from openpyxl import load_workbook

from app.services.step1_rates.writers.base import (
    is_formula_cell,
    safe_set,
    stamp_document_properties,
)
from app.services.step2_bidding.entities import (
    CostType,
    FillReport,
    ParsedPkg,
    PerRowReport,
    PkgRow,
    PkgSection,
    RowStatus,
)


_SHEET_NAME = "見積りシート"
_HEADER_ORIGIN_LABEL = "発地"
_HEADER_DEST_LABEL = "着地"
_HEADER_CARRIER_LABEL = "主要キャリアとルート"
_MAX_SCAN_ROWS = 60  # 黄金样本 41 行；留余量

# 列位置（Customer A 固定；不对其他客户复用，见 §6.3 说明）
_COL_ORIGIN = 2           # B
_COL_DESTINATION = 3      # C
_COL_VOLUME = 4           # D
_COL_PRICE = 5            # E
_COL_LEAD_TIME = 6        # F
_COL_CARRIER = 7          # G
_COL_REMARK = 8           # H

# 段码映射：按原文关键字 → section_code，长匹配优先
_ORIGIN_MAP: list[tuple[str, str]] = [
    ("インチョン", "ICN"),
    ("アムステルダム", "AMS"),
    ("オランダ", "AMS"),
    ("台北", "TPE"),
    ("台湾", "TPE"),
    ("成田", "NRT"),
    ("日本", "NRT"),
    ("上海", "PVG"),
    ("中国", "PVG"),
    ("韓国", "ICN"),
]

_DEST_MAP: list[tuple[str, str]] = [
    ("アトランタ", "ATL"),
    ("マイアミ", "MIA"),
    ("アムステルダム", "AMS"),
    ("サンパウロ", "GRU"),
    ("シドニー", "SYD"),
    ("台北", "TPE"),
    ("上海", "PVG"),
]

# 本轮 Customer A 定义 PVG 段为本地（发自上海）段。架构扩展点见 §14。
_LOCAL_SECTION_CODES = {"PVG"}

_EXAMPLE_MARKER = "記入例"
_CURRENCY_PATTERNS: list[tuple[str, str]] = [
    ("CNY", "CNY"),
    ("USD", "USD"),
    ("ユーロ", "EUR"),
    ("EUR", "EUR"),
    ("円", "JPY"),
    ("JPY", "JPY"),
]

# T-B7 常量
_DEFAULT_MARKUP_RATIO = Decimal("1.15")
_SR_FIXED_REMARK = "ALL-in"
_T_B7_WRITER_VERSION = "step2-customer_a-fill-0.1.0"
_SENTINEL_KEEP = object()
_ALL_KEEP: tuple[object, object, object, object] = (
    _SENTINEL_KEEP,
    _SENTINEL_KEEP,
    _SENTINEL_KEEP,
    _SENTINEL_KEEP,
)


@dataclass(slots=True)
class _HeaderLocation:
    row: int
    currency: str
    currency_header_raw: str


class CustomerAProfile:
    """Customer A 规则法解析器 + fill 回写器（无 AI）。"""

    customer_code = "customer_a"
    display_name = "ミマキエンジニアリング"
    priority = 100

    def __init__(self, markup_fn: Callable[[Decimal], Decimal] | None = None) -> None:
        """markup_fn：cost→sell 纯函数；仅在 variant='sr' 时使用。默认用 default_markup_fn。"""
        self._markup_fn: Callable[[Decimal], Decimal] = markup_fn or default_markup_fn

    # ---------- detect ----------

    def detect(self, path: Path, hint: str | None = None) -> bool:
        if hint == self.customer_code:
            return True
        try:
            wb = load_workbook(path, data_only=True, read_only=True)
        except Exception:
            return False
        try:
            if len(wb.sheetnames) != 1:
                return False
            ws = wb[wb.sheetnames[0]]
            if (ws.title or "").strip() != _SHEET_NAME:
                return False
            header_count = 0
            scan_limit = min(ws.max_row or 0, _MAX_SCAN_ROWS)
            for r in range(1, scan_limit + 1):
                b_val = _norm_text(ws.cell(r, _COL_ORIGIN).value)
                c_val = _norm_text(ws.cell(r, _COL_DESTINATION).value)
                g_val = _norm_text(ws.cell(r, _COL_CARRIER).value)
                if (
                    b_val == _HEADER_ORIGIN_LABEL
                    and _HEADER_DEST_LABEL in c_val
                    and g_val == _HEADER_CARRIER_LABEL
                ):
                    header_count += 1
            return header_count >= 2
        finally:
            wb.close()

    # ---------- fill (T-B7) ----------

    def fill(
        self,
        source_path: Path,
        parsed: ParsedPkg,
        row_reports: list[PerRowReport],
        variant: str,
        output_path: Path,
    ) -> FillReport:
        """回写 PVG 段 5 条 AIR_FREIGHT 行；产出独立 xlsx 文件。

        参见 T-B7 架构任务单 §4 数据流 / §5 决策表。
        一次调用产一个文件；variant∈{"cost","sr"}。
        """
        if variant not in ("cost", "sr"):
            raise ValueError(f"variant 必须为 cost / sr；实际：{variant!r}")

        shutil.copy2(source_path, output_path)

        wb = load_workbook(output_path, data_only=False, keep_vba=False)
        global_warnings: list[str] = []
        try:
            if _SHEET_NAME not in wb.sheetnames:
                raise ValueError(
                    f"Sheet {_SHEET_NAME!r} 缺失，无法回写"
                )
            ws = wb[_SHEET_NAME]

            pvg_rows = self._pvg_rowset(parsed)
            row_by_idx: dict[int, PkgRow] = {r.row_idx: r for r in parsed.rows}

            for report in row_reports:
                row = row_by_idx.get(report.row_idx)
                if row is None:
                    global_warnings.append(
                        f"row_reports 中 row_idx={report.row_idx} 在 parsed.rows 中不存在"
                    )
                    continue
                if report.row_idx not in pvg_rows:
                    continue

                targets = self._targets_for_status(
                    status=report.status,
                    variant=variant,
                    report=report,
                    row=row,
                )
                for col_idx, value in zip(
                    (_COL_PRICE, _COL_LEAD_TIME, _COL_CARRIER, _COL_REMARK),
                    targets,
                ):
                    if value is _SENTINEL_KEEP:
                        continue
                    cell = ws.cell(row.row_idx, col_idx)
                    if is_formula_cell(cell):
                        continue
                    safe_set(cell, value)

                if (
                    report.status == RowStatus.FILLED
                    and variant == "sr"
                    and row.client_constraint_text
                ):
                    global_warnings.append(
                        f"R{row.row_idx} 有客户约束文本 {row.client_constraint_text!r}，"
                        f"v1.0 未并入 H 列 'ALL-in'；审核页请人工处理"
                    )

            stamp_document_properties(
                wb, batch_id=f"{parsed.bid_id}:{variant}"
            )
            wb.save(output_path)
        finally:
            wb.close()

        return self._build_fill_report(
            parsed=parsed,
            row_reports=row_reports,
            variant=variant,
            output_path=output_path,
            warnings=global_warnings,
        )

    # ---------- T-B7 helpers ----------

    @staticmethod
    def _pvg_rowset(parsed: ParsedPkg) -> set[int]:
        """PVG 段（is_local_section=True）所有 row.row_idx 集合。"""
        local_section_indices = {
            s.section_index for s in parsed.sections if s.is_local_section
        }
        return {
            r.row_idx
            for r in parsed.rows
            if r.section_index in local_section_indices
        }

    def _targets_for_status(
        self,
        *,
        status: RowStatus,
        variant: str,
        report: PerRowReport,
        row: PkgRow,
    ) -> tuple[object, object, object, object]:
        """根据 RowStatus × variant 查决策表，返回 (E, F, G, H) 四元组。

        _SENTINEL_KEEP = 保留 2-①.xlsx 原值。
        决策表来源：T-B7 架构任务单 §5.2。
        """
        # 双保险：LOCAL_DELIVERY 行一律 KEEP（即便 status 误标）
        if row.cost_type == CostType.LOCAL_DELIVERY:
            return _ALL_KEEP

        if status == RowStatus.NON_LOCAL_LEG:
            return _ALL_KEEP
        if status == RowStatus.EXAMPLE:
            return _ALL_KEEP
        if status == RowStatus.LOCAL_DELIVERY_MANUAL:
            return _ALL_KEEP
        if status == RowStatus.ALREADY_FILLED:
            return _ALL_KEEP

        if status == RowStatus.NO_RATE:
            return ("", "", "", _SENTINEL_KEEP)

        if status == RowStatus.CONSTRAINT_BLOCK:
            constraint = report.remark_text or "; ".join(report.constraint_hits) or ""
            return ("", "", "", constraint)

        # FILLED 与 OVERRIDDEN（本轮按 FILLED 对待）
        if status in (RowStatus.FILLED, RowStatus.OVERRIDDEN):
            cost = report.cost_price
            if cost is None:
                return _ALL_KEEP
            lead = report.lead_time_text
            carrier = report.carrier_text
            if variant == "cost":
                return (cost, lead, carrier, _SENTINEL_KEEP)
            # variant == "sr"
            sell = self._markup_fn(cost)
            return (sell, lead, carrier, _SR_FIXED_REMARK)

        # 未知 status：保守不动
        return _ALL_KEEP

    def _build_fill_report(
        self,
        *,
        parsed: ParsedPkg,
        row_reports: list[PerRowReport],
        variant: str,
        output_path: Path,
        warnings: list[str],
    ) -> FillReport:
        filled_count = sum(
            1 for r in row_reports if r.status == RowStatus.FILLED
        )
        no_rate_count = sum(
            1 for r in row_reports if r.status == RowStatus.NO_RATE
        )
        skipped_count = len(row_reports) - filled_count - no_rate_count
        return FillReport(
            bid_id=parsed.bid_id,
            generated_at=datetime.utcnow(),
            row_reports=list(row_reports),
            filled_count=filled_count,
            no_rate_count=no_rate_count,
            skipped_count=skipped_count,
            cost_file_path=str(output_path) if variant == "cost" else "",
            sr_file_path=str(output_path) if variant == "sr" else "",
            global_warnings=warnings,
        )

    # ---------- parse ----------

    def parse(self, path: Path, bid_id: str, period: str) -> ParsedPkg:
        wb = load_workbook(path, data_only=True)
        try:
            if _SHEET_NAME not in wb.sheetnames:
                raise ValueError(
                    f"Customer A 解析失败：找不到 Sheet '{_SHEET_NAME}'，实际：{wb.sheetnames}"
                )
            ws = wb[_SHEET_NAME]
            warnings: list[str] = []

            # 若 period 未传（或空）从 B1 读取
            resolved_period = period or self._read_period_from_b1(ws) or ""

            max_row = min(ws.max_row or 0, _MAX_SCAN_ROWS)

            # Step1: 扫表头
            header_locs = self._scan_headers(ws, max_row)
            if len(header_locs) < 1:
                warnings.append("未识别到任何表头行（発地/着地 模式未命中）")

            # Step2: 对每个表头识别段起点（发地文本所在行）和数据行范围
            sections, rows = self._parse_sections_and_rows(
                ws, header_locs, max_row, warnings
            )

            return ParsedPkg(
                bid_id=bid_id,
                customer_code=self.customer_code,
                period=resolved_period,
                sheet_name=ws.title,
                source_file=str(path),
                sections=sections,
                rows=rows,
                warnings=warnings,
            )
        finally:
            wb.close()

    # ---------- helpers ----------

    @staticmethod
    def _read_period_from_b1(ws) -> str | None:
        val = ws.cell(1, _COL_ORIGIN).value
        if val is None:
            return None
        return str(val).strip() or None

    def _scan_headers(self, ws, max_row: int) -> list[_HeaderLocation]:
        locs: list[_HeaderLocation] = []
        for r in range(1, max_row + 1):
            b_val = _norm_text(ws.cell(r, _COL_ORIGIN).value)
            c_val = _norm_text(ws.cell(r, _COL_DESTINATION).value)
            if b_val != _HEADER_ORIGIN_LABEL or _HEADER_DEST_LABEL not in c_val:
                continue
            price_header = ws.cell(r, _COL_PRICE).value
            currency, raw = _parse_currency(price_header)
            locs.append(
                _HeaderLocation(row=r, currency=currency, currency_header_raw=raw)
            )
        return locs

    def _parse_sections_and_rows(
        self,
        ws,
        header_locs: list[_HeaderLocation],
        max_row: int,
        warnings: list[str],
    ) -> tuple[list[PkgSection], list[PkgRow]]:
        sections: list[PkgSection] = []
        rows: list[PkgRow] = []

        for idx, header in enumerate(header_locs):
            next_header_row = (
                header_locs[idx + 1].row if idx + 1 < len(header_locs) else max_row + 1
            )
            data_start = header.row + 1
            data_end = next_header_row - 1  # 含

            section_origin_text, section_origin_code = self._read_section_origin(
                ws, data_start, data_end
            )

            if section_origin_code is None:
                warnings.append(
                    f"段 {idx} (header R{header.row}) 未识别发地代码，原文={section_origin_text!r}"
                )
                section_code = f"SECTION_{idx}"
            else:
                section_code = section_origin_code

            section = PkgSection(
                section_index=idx,
                section_code=section_code,
                header_row=header.row,
                origin_text_raw=section_origin_text or "",
                origin_code=section_origin_code or "UNKNOWN",
                currency=header.currency,
                currency_header_raw=header.currency_header_raw,
                is_local_section=section_code in _LOCAL_SECTION_CODES,
            )
            sections.append(section)

            # 数据行 & 段级约束
            self._scan_section_body(
                ws, section, data_start, data_end, rows, warnings
            )

        return sections, rows

    @staticmethod
    def _read_section_origin(ws, data_start: int, data_end: int) -> tuple[str, str | None]:
        """在数据区首个非空 B 列值作为段发地原文，并映射代码。"""
        for r in range(data_start, data_end + 1):
            b_val = ws.cell(r, _COL_ORIGIN).value
            if b_val is None:
                continue
            text = str(b_val).strip()
            if not text:
                continue
            # 跳过段级约束（以 '※' 开头）
            if text.startswith("※"):
                continue
            code = _map_origin(text)
            return text, code
        return "", None

    def _scan_section_body(
        self,
        ws,
        section: PkgSection,
        data_start: int,
        data_end: int,
        rows_out: list[PkgRow],
        warnings: list[str],
    ) -> None:
        blank_streak = 0
        for r in range(data_start, data_end + 1):
            b_raw = ws.cell(r, _COL_ORIGIN).value
            c_raw = ws.cell(r, _COL_DESTINATION).value
            e_raw = ws.cell(r, _COL_PRICE).value

            b_text = _norm_text(b_raw)
            c_text = _norm_text(c_raw)

            # 段级约束：B 列非空，以 ※ 开头，且 C/E 均空 → 并入 section_level_remarks
            if (
                b_text.startswith("※")
                and not c_text
                and e_raw is None
            ):
                section.section_level_remarks.append(str(b_raw).strip())
                blank_streak = 0
                continue

            # 全空行
            if not b_text and not c_text and e_raw is None:
                blank_streak += 1
                if blank_streak >= 2:
                    break
                continue
            blank_streak = 0

            # 数据行必须有目的地文本
            if not c_text:
                continue

            row = self._build_pkg_row(ws, r, section, c_raw)
            rows_out.append(row)

    @staticmethod
    def _build_pkg_row(ws, r: int, section: PkgSection, c_raw) -> PkgRow:
        dest_text_raw = str(c_raw)
        dest_code = _map_destination(dest_text_raw)
        cost_type = _infer_cost_type(dest_text_raw)

        d_raw = ws.cell(r, _COL_VOLUME).value
        e_raw = ws.cell(r, _COL_PRICE).value
        f_raw = ws.cell(r, _COL_LEAD_TIME).value
        g_raw = ws.cell(r, _COL_CARRIER).value
        h_raw = ws.cell(r, _COL_REMARK).value

        g_text = _norm_text(g_raw)
        h_text = _norm_text(h_raw)
        is_example = (_EXAMPLE_MARKER in g_text) or (_EXAMPLE_MARKER in h_text)

        client_constraint = _strip_example(h_text)

        return PkgRow(
            row_idx=r,
            section_index=section.section_index,
            section_code=section.section_code,
            origin_code=section.origin_code,
            origin_text_raw=section.origin_text_raw,
            destination_text_raw=dest_text_raw,
            destination_code=dest_code,
            cost_type=cost_type,
            currency=section.currency,
            volume_desc=str(d_raw) if d_raw is not None else None,
            existing_price=_to_decimal(e_raw),
            existing_lead_time=str(f_raw) if f_raw is not None else None,
            existing_carrier=str(g_raw) if g_raw is not None else None,
            existing_remark=str(h_raw) if h_raw is not None else None,
            is_example=is_example,
            client_constraint_text=client_constraint or None,
        )


# ---------- module-level helpers ----------


def _norm_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _parse_currency(header_value: object) -> tuple[str, str]:
    """从 E 列表头原文解析币种。

    默认 CNY（匹配失败时），并把原文放入 currency_header_raw 便于审计。
    """
    raw = str(header_value) if header_value is not None else ""
    for keyword, code in _CURRENCY_PATTERNS:
        if keyword in raw:
            return code, raw
    return "CNY", raw


def _map_origin(text: str) -> str | None:
    for keyword, code in _ORIGIN_MAP:
        if keyword in text:
            return code
    return None


def _map_destination(text: str) -> str:
    for keyword, code in _DEST_MAP:
        if keyword in text:
            return code
    return "UNKNOWN"


def _infer_cost_type(dest_text: str) -> CostType:
    upper = dest_text.upper()
    if "LOCAL DELIVERY" in upper:
        return CostType.LOCAL_DELIVERY
    if "AIR FREIGHT" in upper:
        return CostType.AIR_FREIGHT
    # 默认 AIR_FREIGHT（见任务单 §6.2 规则）
    return CostType.AIR_FREIGHT


def _to_decimal(value: object) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        try:
            return Decimal(str(value))
        except InvalidOperation:
            return None
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped or stripped == "－" or stripped == "-":
            return None
        try:
            return Decimal(stripped)
        except InvalidOperation:
            return None
    return None


_EXAMPLE_BLOCK_RE = re.compile(r"※\s*記入例[\s\S]*", re.MULTILINE)


def _strip_example(text: str) -> str:
    if not text:
        return ""
    return _EXAMPLE_BLOCK_RE.sub("", text).strip()


def _ceil_int(value: Decimal) -> Decimal:
    """向上取整到整数。业务需求 §需求 7 V1：45×1.15=51.75 → 52。"""
    return value.to_integral_value(rounding=ROUND_CEILING)


def default_markup_fn(cost: Decimal) -> Decimal:
    """T-B6 未到位时的兜底加价：cost × 1.15，向上取整到整数。"""
    return _ceil_int(cost * _DEFAULT_MARKUP_RATIO)
