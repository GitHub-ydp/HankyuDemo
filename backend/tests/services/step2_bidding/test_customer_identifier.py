"""Step2 T-B8 customer_identifier 验收用例（V-T-B8-01..12 + R01）。

业务依据：docs/Step2_入札対応_T-B8_customer_identifier_业务需求_20260423.md
架构任务单：docs/Step2_入札対応_T-B8_customer_identifier_架构任务单_20260423.md §8

样本路径模式沿用 test_customer_a_parse.py:17-24（GOLDEN_SAMPLE）。
"""
from __future__ import annotations

import os
from pathlib import Path

import pytest
from openpyxl import Workbook

from app.services.step2_bidding import identify


_DATA_ROOT = Path(__file__).resolve().parents[4] / "资料" / "2026.04.02"

CUSTOMER_A_SAMPLE_1 = (
    _DATA_ROOT / "Customer A (Air)" / "Customer A (Air)" / "2-①.xlsx"
)
CUSTOMER_A_SAMPLE_2 = (
    _DATA_ROOT / "Customer A (Air)" / "Customer A (Air)" / "2-②.xlsx"
)
CUSTOMER_A_SAMPLE_4 = (
    _DATA_ROOT / "Customer A (Air)" / "Customer A (Air)" / "2-④.xlsx"
)
CUSTOMER_B_SAMPLE = (
    _DATA_ROOT / "Customer B (Ocean,LCL)" / "Customer B (Ocean,LCL)" / "2-①.xlsx"
)
CUSTOMER_E_SAMPLE = (
    _DATA_ROOT / "Customer E (Air & Ocean)" / "Customer E (Air & Ocean)" / "2-①.xlsx"
)
NITORI_SAMPLE = (
    _DATA_ROOT
    / "ニトリ様海上入札"
    / "ニトリ様海上入札"
    / "阪急阪神【to GLOBAL】2026年1月～3月_見積り書.xlsx"
)


def _require(path: Path) -> None:
    if not path.exists():
        pytest.skip(f"样本不可用：{path}")


# ---------- V-T-B8-01..03 真实 Customer A 样本 ----------


def test_v_t_b8_01_customer_a_sample_1():
    _require(CUSTOMER_A_SAMPLE_1)
    result = identify(CUSTOMER_A_SAMPLE_1)
    assert result.matched_customer == "customer_a"
    assert result.matched_dimensions == ("B", "D")
    assert result.confidence == "high"
    assert result.warnings == ()
    assert result.unmatched_reason is None
    assert result.source == "auto"


def test_v_t_b8_02_customer_a_sample_2():
    _require(CUSTOMER_A_SAMPLE_2)
    result = identify(CUSTOMER_A_SAMPLE_2)
    assert result.matched_customer == "customer_a"
    assert result.matched_dimensions == ("B", "D")
    assert result.confidence == "high"
    assert result.warnings == ()
    assert result.unmatched_reason is None


def test_v_t_b8_03_customer_a_sample_4():
    _require(CUSTOMER_A_SAMPLE_4)
    result = identify(CUSTOMER_A_SAMPLE_4)
    assert result.matched_customer == "customer_a"
    assert result.matched_dimensions == ("B", "D")
    assert result.confidence == "high"
    assert result.warnings == ()
    assert result.unmatched_reason is None


# ---------- V-T-B8-04..06 反例（B/E/Nitori → unknown） ----------


def test_v_t_b8_04_customer_b_unknown():
    _require(CUSTOMER_B_SAMPLE)
    result = identify(CUSTOMER_B_SAMPLE)
    assert result.matched_customer == "unknown"
    assert result.matched_dimensions == ()
    assert result.confidence == "low"
    assert result.unmatched_reason is not None
    assert "2025 LCL RATE" in result.unmatched_reason


def test_v_t_b8_05_customer_e_unknown():
    _require(CUSTOMER_E_SAMPLE)
    result = identify(CUSTOMER_E_SAMPLE)
    assert result.matched_customer == "unknown"
    assert result.matched_dimensions == ()
    assert result.confidence == "low"
    assert result.unmatched_reason is not None
    assert "AIR入力フォーム" in result.unmatched_reason
    # 不含 MULTI_SHEET（因为不含 見積りシート 这个 sheet 名）
    for w in result.warnings:
        assert not w.startswith("MULTI_SHEET")


def test_v_t_b8_06_nitori_unknown():
    _require(NITORI_SAMPLE)
    result = identify(NITORI_SAMPLE)
    assert result.matched_customer == "unknown"
    assert result.matched_dimensions == ()
    assert result.confidence == "low"


# ---------- V-T-B8-07 损坏文件 ----------


def test_v_t_b8_07_corrupted_xlsx(tmp_path):
    bad = tmp_path / "bad.xlsx"
    bad.write_bytes(os.urandom(1024))
    result = identify(bad)
    assert result.matched_customer == "unknown"
    assert result.matched_dimensions == ()
    assert result.confidence == "low"
    assert len(result.warnings) >= 1
    assert result.warnings[0].startswith("WBOPEN_FAIL:")


# ---------- V-T-B8-08 多 sheet 含 見積りシート（D 命中） ----------


def test_v_t_b8_08_multi_sheet_with_target(tmp_path):
    wb = Workbook()
    target_ws = wb.active
    target_ws.title = "見積りシート"
    target_ws.cell(3, 2, "発地")
    target_ws.cell(3, 3, "着地  (到着空港)")
    target_ws.cell(3, 7, "主要キャリアとルート")
    wb.create_sheet("Notes")
    path = tmp_path / "multi.xlsx"
    wb.save(path)

    result = identify(path)
    assert result.matched_customer == "customer_a"
    assert result.matched_dimensions == ("D",)
    assert result.confidence == "medium"
    assert any(w.startswith("MULTI_SHEET") for w in result.warnings)


# ---------- V-T-B8-09 浮动扫描第 5 行命中 ----------


def test_v_t_b8_09_header_in_row_5(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "見積りシート"
    for r in range(1, 5):
        ws.cell(r, 2, f"公告 R{r}")
    ws.cell(5, 2, "発地")
    ws.cell(5, 3, "着地 (到着空港)")
    ws.cell(5, 7, "主要キャリアとルート")
    path = tmp_path / "row5.xlsx"
    wb.save(path)

    result = identify(path)
    assert result.matched_customer == "customer_a"
    assert result.matched_dimensions == ("B", "D")
    assert result.confidence == "high"


# ---------- V-T-B8-10 仅 B 命中 ----------


def test_v_t_b8_10_only_dim_b(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "見積りシート"
    # 表头三关键字故意不齐：G 列空
    ws.cell(3, 2, "発地")
    ws.cell(3, 3, "着地")
    # G 列留空
    path = tmp_path / "only_b.xlsx"
    wb.save(path)

    result = identify(path)
    assert result.matched_customer == "customer_a"
    assert result.matched_dimensions == ("B",)
    assert result.confidence == "medium"
    assert any(w.startswith("HEADER_MISMATCH") for w in result.warnings)


# ---------- V-T-B8-11 仅 D 命中 ----------


def test_v_t_b8_11_only_dim_d(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Quote Sheet 2026-01"
    ws.cell(3, 2, "発地")
    ws.cell(3, 3, "着地 (到着空港)")
    ws.cell(3, 7, "主要キャリアとルート")
    path = tmp_path / "only_d.xlsx"
    wb.save(path)

    result = identify(path)
    assert result.matched_customer == "customer_a"
    assert result.matched_dimensions == ("D",)
    assert result.confidence == "medium"
    assert any(w.startswith("SHEET_NAME_VARIANT") for w in result.warnings)


# ---------- V-T-B8-12 sheet 名前后空白 ----------


def test_v_t_b8_12_sheet_name_with_whitespace(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "  見積りシート  "
    # D 不命中
    path = tmp_path / "ws.xlsx"
    wb.save(path)

    result = identify(path)
    assert result.matched_customer == "customer_a"
    assert result.matched_dimensions == ("B",)
    assert result.confidence == "medium"


# ---------- V-T-B8-R01 回归：升级 detect 不破坏 T-B4 ----------


def test_v_t_b8_r01_customer_a_detect_existing_pytest_pass():
    """语义同 T-B4 V-B01：升级 detect 后真实样本 detect 仍返 True。"""
    _require(CUSTOMER_A_SAMPLE_1)
    from app.services.step2_bidding.customer_profiles.customer_a import (
        CustomerAProfile,
    )

    assert CustomerAProfile().detect(CUSTOMER_A_SAMPLE_1) is True
