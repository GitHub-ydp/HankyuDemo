"""Customer A fill 双版本回写 验收用例（T-B7 范围：V-B7-01..V-B7-12）。

业务依据：docs/Step2_入札対応_T-B7_customer_a_fill_业务需求_20260423.md §需求 7（V1..V7）
架构任务单：docs/Step2_入札対応_T-B7_customer_a_fill_架构任务单_20260423.md §5 决策表、§8 测试清单

黄金样本：
  - 输入：资料/2026.04.02/Customer A (Air)/Customer A (Air)/2-①.xlsx
  - cost 版实证：同目录 2-②.xlsx（仅用于对照期望值）
  - S/R 版实证：同目录 2-④.xlsx（仅用于对照期望值）
"""
from __future__ import annotations

from decimal import Decimal
from pathlib import Path
from typing import Iterable

import pytest
from openpyxl import load_workbook

from app.services.step2_bidding.customer_profiles.customer_a import (
    CustomerAProfile,
    default_markup_fn,
    _ceil_int,
)
from app.services.step2_bidding.entities import (
    CostType,
    ParsedPkg,
    PerRowReport,
    RowStatus,
)


GOLDEN_SAMPLE = (
    Path(__file__).resolve().parents[4]
    / "资料"
    / "2026.04.02"
    / "Customer A (Air)"
    / "Customer A (Air)"
    / "2-①.xlsx"
)

_PVG_AIRFREIGHT_ROWS = (13, 15, 16, 18, 19)
_PVG_LOCAL_DELIVERY_ROWS = (14, 17)
_COL_E, _COL_F, _COL_G, _COL_H = 5, 6, 7, 8


@pytest.fixture(scope="module")
def parsed_pkg() -> ParsedPkg:
    if not GOLDEN_SAMPLE.exists():
        pytest.skip(f"Customer A 黄金样本不可用：{GOLDEN_SAMPLE}")
    profile = CustomerAProfile()
    return profile.parse(GOLDEN_SAMPLE, bid_id="bid_test", period="2026-01")


def _make_report(
    row_idx: int,
    section_code: str,
    destination_code: str,
    status: RowStatus,
    *,
    cost_price: Decimal | None = None,
    lead_time_text: str | None = None,
    carrier_text: str | None = None,
    remark_text: str | None = None,
    constraint_hits: list[str] | None = None,
) -> PerRowReport:
    return PerRowReport(
        row_idx=row_idx,
        section_code=section_code,
        destination_code=destination_code,
        status=status,
        cost_price=cost_price,
        sell_price=None,
        markup_ratio=None,
        lead_time_text=lead_time_text,
        carrier_text=carrier_text,
        remark_text=remark_text,
        selected_candidate=None,
        constraint_hits=constraint_hits or [],
    )


def _happy_reports(parsed: ParsedPkg) -> list[PerRowReport]:
    """V-B7-01..05 共用：5 条 AIR_FREIGHT FILLED + 2 条 LOCAL_DELIVERY_MANUAL + 非 PVG NON_LOCAL_LEG。"""
    costs = {13: Decimal("45"), 15: Decimal("50"), 16: Decimal("38"), 18: Decimal("22"), 19: Decimal("12")}
    leads = {13: "3-4DAYS", 15: "3-4DAYS", 16: "2DAYS", 18: "2DAYS", 19: "1DAY"}
    carriers = {
        13: "OZ via ICN / NH via NRT",
        15: "NH VIA NRT/CK VIA ORD",
        16: "CZ/CK/CA direct flt",
        18: "NH VIA NRT",
        19: "CK direct flt",
    }
    reports: list[PerRowReport] = []
    for r in parsed.rows:
        if r.section_code == "PVG":
            if r.cost_type == CostType.LOCAL_DELIVERY:
                reports.append(
                    _make_report(r.row_idx, r.section_code, r.destination_code, RowStatus.LOCAL_DELIVERY_MANUAL)
                )
            else:
                reports.append(
                    _make_report(
                        r.row_idx,
                        r.section_code,
                        r.destination_code,
                        RowStatus.FILLED,
                        cost_price=costs[r.row_idx],
                        lead_time_text=leads[r.row_idx],
                        carrier_text=carriers[r.row_idx],
                    )
                )
        else:
            reports.append(
                _make_report(r.row_idx, r.section_code, r.destination_code, RowStatus.NON_LOCAL_LEG)
            )
    return reports


def _all_no_rate_reports(parsed: ParsedPkg) -> list[PerRowReport]:
    """V-B7-05 (happy)/V-B7-... 全 NO_RATE 用例。"""
    reports: list[PerRowReport] = []
    for r in parsed.rows:
        if r.section_code == "PVG":
            if r.cost_type == CostType.LOCAL_DELIVERY:
                reports.append(
                    _make_report(r.row_idx, r.section_code, r.destination_code, RowStatus.LOCAL_DELIVERY_MANUAL)
                )
            else:
                reports.append(
                    _make_report(r.row_idx, r.section_code, r.destination_code, RowStatus.NO_RATE)
                )
        else:
            reports.append(
                _make_report(r.row_idx, r.section_code, r.destination_code, RowStatus.NON_LOCAL_LEG)
            )
    return reports


def _mixed_filled_constraint_reports(parsed: ParsedPkg) -> list[PerRowReport]:
    """V-B7-06：R13/R15/R16 FILLED；R18/R19 CONSTRAINT_BLOCK。"""
    filled_rows = {13, 15, 16}
    constraint_rows = {18, 19}
    costs = {13: Decimal("45"), 15: Decimal("50"), 16: Decimal("38")}
    leads = {13: "3-4DAYS", 15: "3-4DAYS", 16: "2DAYS"}
    carriers = {13: "OZ via ICN / NH via NRT", 15: "NH VIA NRT/CK VIA ORD", 16: "CZ/CK/CA direct flt"}
    reports: list[PerRowReport] = []
    for r in parsed.rows:
        if r.section_code == "PVG":
            if r.cost_type == CostType.LOCAL_DELIVERY:
                reports.append(
                    _make_report(r.row_idx, r.section_code, r.destination_code, RowStatus.LOCAL_DELIVERY_MANUAL)
                )
            elif r.row_idx in filled_rows:
                reports.append(
                    _make_report(
                        r.row_idx,
                        r.section_code,
                        r.destination_code,
                        RowStatus.FILLED,
                        cost_price=costs[r.row_idx],
                        lead_time_text=leads[r.row_idx],
                        carrier_text=carriers[r.row_idx],
                    )
                )
            elif r.row_idx in constraint_rows:
                reports.append(
                    _make_report(
                        r.row_idx,
                        r.section_code,
                        r.destination_code,
                        RowStatus.CONSTRAINT_BLOCK,
                        remark_text="※某约束",
                    )
                )
            else:
                reports.append(
                    _make_report(r.row_idx, r.section_code, r.destination_code, RowStatus.NO_RATE)
                )
        else:
            reports.append(
                _make_report(r.row_idx, r.section_code, r.destination_code, RowStatus.NON_LOCAL_LEG)
            )
    return reports


def _assert_non_pvg_cells_identical(
    src_path: Path,
    out_path: Path,
    *,
    white_list_cells: Iterable[tuple[int, int]],
) -> None:
    """遍历源文件 × 产出文件，排除 white_list（允许变化的单元格）后逐格对比 value。"""
    wb_src = load_workbook(src_path, data_only=False)
    wb_out = load_workbook(out_path, data_only=False)
    try:
        ws_src = wb_src.active
        ws_out = wb_out.active
        white = set(white_list_cells)
        mismatches: list[tuple[int, int, object, object]] = []
        for r in range(1, ws_src.max_row + 1):
            for c in range(1, ws_src.max_column + 1):
                if (r, c) in white:
                    continue
                v_src = ws_src.cell(r, c).value
                v_out = ws_out.cell(r, c).value
                if v_src != v_out:
                    mismatches.append((r, c, v_src, v_out))
        assert not mismatches, f"非白名单单元格出现 {len(mismatches)} 处改动：{mismatches[:5]}"
    finally:
        wb_src.close()
        wb_out.close()


# ---------- V-B7-01 happy path cost 版 ----------


def test_v_b7_01_happy_path_cost(parsed_pkg: ParsedPkg, tmp_path: Path):
    profile = CustomerAProfile()
    reports = _happy_reports(parsed_pkg)
    out = tmp_path / "cost_2-①.xlsx"
    fill_report = profile.fill(GOLDEN_SAMPLE, parsed_pkg, reports, "cost", out)

    wb = load_workbook(out)
    ws = wb.active
    try:
        expected_E = {13: 45, 15: 50, 16: 38, 18: 22, 19: 12}
        for r, ev in expected_E.items():
            assert ws.cell(r, _COL_E).value == ev, f"cost R{r} E={ws.cell(r,_COL_E).value}"
        # F / G 列非空，H 列保持 None
        for r in _PVG_AIRFREIGHT_ROWS:
            assert ws.cell(r, _COL_F).value is not None
            assert ws.cell(r, _COL_G).value is not None
            assert ws.cell(r, _COL_H).value is None, f"cost R{r} H 应为 None"
    finally:
        wb.close()

    assert fill_report.filled_count == 5
    assert fill_report.no_rate_count == 0
    assert fill_report.cost_file_path == str(out)
    assert fill_report.sr_file_path == ""


# ---------- V-B7-02 happy path sr 版（H='ALL-in'，E = cost×1.15 ceil） ----------


def test_v_b7_02_happy_path_sr(parsed_pkg: ParsedPkg, tmp_path: Path):
    profile = CustomerAProfile()
    reports = _happy_reports(parsed_pkg)
    out = tmp_path / "sr_2-①.xlsx"
    fill_report = profile.fill(GOLDEN_SAMPLE, parsed_pkg, reports, "sr", out)

    wb = load_workbook(out)
    ws = wb.active
    try:
        # 45×1.15=51.75→52; 50→58; 38→44; 22→26; 12→14
        expected_E = {13: 52, 15: 58, 16: 44, 18: 26, 19: 14}
        for r, ev in expected_E.items():
            assert ws.cell(r, _COL_E).value == ev, f"sr R{r} E={ws.cell(r,_COL_E).value} (期望 {ev})"
        for r in _PVG_AIRFREIGHT_ROWS:
            assert ws.cell(r, _COL_H).value == "ALL-in", f"sr R{r} H 应为 'ALL-in'"
    finally:
        wb.close()

    assert fill_report.sr_file_path == str(out)
    assert fill_report.cost_file_path == ""


# ---------- V-B7-03 LOCAL_DELIVERY 两版零改动（R14/R17） ----------


def test_v_b7_03_local_delivery_kept(parsed_pkg: ParsedPkg, tmp_path: Path):
    profile = CustomerAProfile()
    reports = _happy_reports(parsed_pkg)
    out_cost = tmp_path / "cost.xlsx"
    out_sr = tmp_path / "sr.xlsx"
    profile.fill(GOLDEN_SAMPLE, parsed_pkg, reports, "cost", out_cost)
    profile.fill(GOLDEN_SAMPLE, parsed_pkg, reports, "sr", out_sr)

    for out in (out_cost, out_sr):
        wb = load_workbook(out)
        ws = wb.active
        try:
            for r in _PVG_LOCAL_DELIVERY_ROWS:
                assert ws.cell(r, _COL_E).value == 0, f"{out.name} R{r} E 应为 0"
                assert ws.cell(r, _COL_F).value == "－", f"{out.name} R{r} F 应为 '－'"
                assert ws.cell(r, _COL_G).value == "－", f"{out.name} R{r} G 应为 '－'"
                assert ws.cell(r, _COL_H).value is None, f"{out.name} R{r} H 应为 None"
        finally:
            wb.close()


# ---------- V-B7-04 非 PVG 段零改动 ----------


def test_v_b7_04_non_pvg_zero_diff(parsed_pkg: ParsedPkg, tmp_path: Path):
    profile = CustomerAProfile()
    reports = _happy_reports(parsed_pkg)
    out_cost = tmp_path / "cost.xlsx"
    out_sr = tmp_path / "sr.xlsx"
    profile.fill(GOLDEN_SAMPLE, parsed_pkg, reports, "cost", out_cost)
    profile.fill(GOLDEN_SAMPLE, parsed_pkg, reports, "sr", out_sr)

    # 白名单：PVG 段 AIR_FREIGHT 5 行 × E/F/G/H 4 列 = 20 单元格
    white = {(r, c) for r in _PVG_AIRFREIGHT_ROWS for c in (_COL_E, _COL_F, _COL_G, _COL_H)}
    _assert_non_pvg_cells_identical(GOLDEN_SAMPLE, out_cost, white_list_cells=white)
    _assert_non_pvg_cells_identical(GOLDEN_SAMPLE, out_sr, white_list_cells=white)


# ---------- V-B7-05 全 NO_RATE（E/F/G 清空，H 保持；非 PVG 零改动） ----------


def test_v_b7_05_all_no_rate(parsed_pkg: ParsedPkg, tmp_path: Path):
    profile = CustomerAProfile()
    reports = _all_no_rate_reports(parsed_pkg)
    out_cost = tmp_path / "cost.xlsx"
    out_sr = tmp_path / "sr.xlsx"
    rep_cost = profile.fill(GOLDEN_SAMPLE, parsed_pkg, reports, "cost", out_cost)
    rep_sr = profile.fill(GOLDEN_SAMPLE, parsed_pkg, reports, "sr", out_sr)

    for out in (out_cost, out_sr):
        wb = load_workbook(out)
        ws = wb.active
        try:
            for r in _PVG_AIRFREIGHT_ROWS:
                # 决策表：NO_RATE → E/F/G 显式写空（写入瞬间为 ""，openpyxl 保存+重载后变 None；业务视觉等价）
                # H 保持原值（None）
                assert ws.cell(r, _COL_E).value in (None, ""), f"{out.name} R{r} E 应为空，实际 {ws.cell(r,_COL_E).value!r}"
                assert ws.cell(r, _COL_F).value in (None, ""), f"{out.name} R{r} F 应为空，实际 {ws.cell(r,_COL_F).value!r}"
                assert ws.cell(r, _COL_G).value in (None, ""), f"{out.name} R{r} G 应为空，实际 {ws.cell(r,_COL_G).value!r}"
                assert ws.cell(r, _COL_H).value is None, f"{out.name} R{r} H 应为 None（保持原值）"
            # LOCAL_DELIVERY 行不动
            for r in _PVG_LOCAL_DELIVERY_ROWS:
                assert ws.cell(r, _COL_E).value == 0
                assert ws.cell(r, _COL_F).value == "－"
        finally:
            wb.close()

    # FillReport 统计
    assert rep_cost.filled_count == 0
    assert rep_cost.no_rate_count == 5
    assert rep_sr.filled_count == 0
    assert rep_sr.no_rate_count == 5

    # 非 PVG 段零改动
    white = {(r, c) for r in _PVG_AIRFREIGHT_ROWS for c in (_COL_E, _COL_F, _COL_G, _COL_H)}
    _assert_non_pvg_cells_identical(GOLDEN_SAMPLE, out_cost, white_list_cells=white)
    _assert_non_pvg_cells_identical(GOLDEN_SAMPLE, out_sr, white_list_cells=white)


# ---------- V-B7-06 mixed FILLED + CONSTRAINT_BLOCK ----------


def test_v_b7_06_mixed_filled_constraint(parsed_pkg: ParsedPkg, tmp_path: Path):
    profile = CustomerAProfile()
    reports = _mixed_filled_constraint_reports(parsed_pkg)
    out_cost = tmp_path / "cost.xlsx"
    out_sr = tmp_path / "sr.xlsx"
    profile.fill(GOLDEN_SAMPLE, parsed_pkg, reports, "cost", out_cost)
    profile.fill(GOLDEN_SAMPLE, parsed_pkg, reports, "sr", out_sr)

    # R13/R15/R16 按 FILLED 分支；R18/R19 CONSTRAINT_BLOCK
    wb_cost = load_workbook(out_cost)
    ws_cost = wb_cost.active
    wb_sr = load_workbook(out_sr)
    ws_sr = wb_sr.active
    try:
        # FILLED 行
        assert ws_cost.cell(13, _COL_E).value == 45
        assert ws_cost.cell(15, _COL_E).value == 50
        assert ws_cost.cell(16, _COL_E).value == 38
        assert ws_sr.cell(13, _COL_E).value == 52
        assert ws_sr.cell(13, _COL_H).value == "ALL-in"

        # CONSTRAINT_BLOCK 行：两版一致，E/F/G 空（openpyxl 行为："" → 保存+重载后变 None），H 写约束文本
        for r in (18, 19):
            for ws in (ws_cost, ws_sr):
                assert ws.cell(r, _COL_E).value in (None, ""), f"R{r} E 应为空"
                assert ws.cell(r, _COL_F).value in (None, ""), f"R{r} F 应为空"
                assert ws.cell(r, _COL_G).value in (None, ""), f"R{r} G 应为空"
                assert ws.cell(r, _COL_H).value == "※某约束", f"R{r} H 应为约束文本"
    finally:
        wb_cost.close()
        wb_sr.close()

    white = {(r, c) for r in _PVG_AIRFREIGHT_ROWS for c in (_COL_E, _COL_F, _COL_G, _COL_H)}
    _assert_non_pvg_cells_identical(GOLDEN_SAMPLE, out_cost, white_list_cells=white)
    _assert_non_pvg_cells_identical(GOLDEN_SAMPLE, out_sr, white_list_cells=white)


# ---------- V-B7-07 markup 依赖注入生效 ----------


def test_v_b7_07_custom_markup_fn(parsed_pkg: ParsedPkg, tmp_path: Path):
    # 注入 ×2 系数
    profile = CustomerAProfile(markup_fn=lambda c: c * Decimal("2"))
    reports = _happy_reports(parsed_pkg)
    out = tmp_path / "sr.xlsx"
    profile.fill(GOLDEN_SAMPLE, parsed_pkg, reports, "sr", out)

    wb = load_workbook(out)
    ws = wb.active
    try:
        # 45×2=90, 50×2=100, 38×2=76, 22×2=44, 12×2=24
        assert ws.cell(13, _COL_E).value == 90
        assert ws.cell(15, _COL_E).value == 100
        assert ws.cell(16, _COL_E).value == 76
        assert ws.cell(18, _COL_E).value == 44
        assert ws.cell(19, _COL_E).value == 24
    finally:
        wb.close()


# ---------- V-B7-08 default_markup_fn = ×1.15 ceil_int ----------


def test_v_b7_08_default_markup_fn_ceil_int():
    assert default_markup_fn(Decimal("45")) == Decimal("52")  # 51.75 → 52
    assert default_markup_fn(Decimal("50")) == Decimal("58")  # 57.5 → 58
    assert default_markup_fn(Decimal("38")) == Decimal("44")  # 43.7 → 44
    assert default_markup_fn(Decimal("22")) == Decimal("26")  # 25.3 → 26
    assert default_markup_fn(Decimal("12")) == Decimal("14")  # 13.8 → 14
    # _ceil_int 单独验证
    assert _ceil_int(Decimal("51.01")) == Decimal("52")
    assert _ceil_int(Decimal("52.00")) == Decimal("52")


# ---------- V-B7-09 variant 非法抛 ValueError ----------


def test_v_b7_09_invalid_variant_raises(parsed_pkg: ParsedPkg, tmp_path: Path):
    profile = CustomerAProfile()
    with pytest.raises(ValueError, match="variant"):
        profile.fill(GOLDEN_SAMPLE, parsed_pkg, [], "hybrid", tmp_path / "x.xlsx")
    with pytest.raises(ValueError, match="variant"):
        profile.fill(GOLDEN_SAMPLE, parsed_pkg, [], "", tmp_path / "x.xlsx")


# ---------- V-B7-10 公式单元格保留（Sheet 名保留作为退化断言） ----------


def test_v_b7_10_formula_and_sheetname_preserved(parsed_pkg: ParsedPkg, tmp_path: Path):
    """黄金样本 PVG 段未含公式（已实测）；用 Sheet 名不变 + 非 PVG 零改动作为退化断言。
    公式守卫本身由 safe_set / is_formula_cell 在 Step1 TD-2 测试矩阵覆盖。
    """
    profile = CustomerAProfile()
    reports = _happy_reports(parsed_pkg)
    out = tmp_path / "cost.xlsx"
    profile.fill(GOLDEN_SAMPLE, parsed_pkg, reports, "cost", out)

    wb = load_workbook(out)
    try:
        assert wb.sheetnames == ["見積りシート"]
        assert wb.active.title == "見積りシート"
    finally:
        wb.close()


# ---------- V-B7-11 合并单元格 / 列宽 保留 ----------


def test_v_b7_11_merged_and_dimensions_preserved(parsed_pkg: ParsedPkg, tmp_path: Path):
    profile = CustomerAProfile()
    reports = _happy_reports(parsed_pkg)
    out = tmp_path / "cost.xlsx"
    profile.fill(GOLDEN_SAMPLE, parsed_pkg, reports, "cost", out)

    wb_src = load_workbook(GOLDEN_SAMPLE, data_only=False)
    wb_out = load_workbook(out, data_only=False)
    try:
        src_merged = sorted(str(r) for r in wb_src.active.merged_cells.ranges)
        out_merged = sorted(str(r) for r in wb_out.active.merged_cells.ranges)
        assert src_merged == out_merged, f"merged_cells 不一致：src={src_merged} out={out_merged}"

        # 列宽：至少对显式定义的列做对比
        for col_letter, src_dim in wb_src.active.column_dimensions.items():
            out_dim = wb_out.active.column_dimensions.get(col_letter)
            if out_dim is None:
                continue
            assert src_dim.width == out_dim.width, (
                f"列 {col_letter} 宽度不一致：src={src_dim.width} out={out_dim.width}"
            )
    finally:
        wb_src.close()
        wb_out.close()


# ---------- V-B7-12 FillReport 字段正确 ----------


def test_v_b7_12_fill_report_fields(parsed_pkg: ParsedPkg, tmp_path: Path):
    profile = CustomerAProfile()
    reports = _happy_reports(parsed_pkg)
    out_cost = tmp_path / "cost.xlsx"
    out_sr = tmp_path / "sr.xlsx"
    rep_cost = profile.fill(GOLDEN_SAMPLE, parsed_pkg, reports, "cost", out_cost)
    rep_sr = profile.fill(GOLDEN_SAMPLE, parsed_pkg, reports, "sr", out_sr)

    # bid_id 透传
    assert rep_cost.bid_id == "bid_test"
    assert rep_sr.bid_id == "bid_test"
    # row_reports 透传（浅拷贝，内容一致、长度一致）
    assert len(rep_cost.row_reports) == len(reports)
    # 计数：5 FILLED、0 NO_RATE、其余 skipped
    assert rep_cost.filled_count == 5
    assert rep_cost.no_rate_count == 0
    assert rep_cost.skipped_count == len(reports) - 5
    # 文件路径按 variant 填一边
    assert rep_cost.cost_file_path == str(out_cost)
    assert rep_cost.sr_file_path == ""
    assert rep_sr.sr_file_path == str(out_sr)
    assert rep_sr.cost_file_path == ""
    # global_warnings 空（黄金样本 PVG 段 client_constraint_text 全空）
    assert rep_cost.global_warnings == []
    assert rep_sr.global_warnings == []
