"""Air adapter 验收用例（V-A01..V-A20，对齐架构任务单 §10）。"""
from __future__ import annotations

import json
from datetime import date
from decimal import Decimal
from pathlib import Path

import pytest

from app.services.step1_rates.adapters.air import AirAdapter
from app.services.step1_rates.adapters.ocean import OceanAdapter
from app.services.step1_rates.adapters.ocean_ngb import OceanNgbAdapter
from app.services.step1_rates.entities import ParsedRateBatch, Step1FileType
from app.services.step1_rates.registry import RateAdapterRegistry


REAL_SAMPLE_DIR = (
    Path(__file__).resolve().parents[4]
    / "資料"  # placeholder so lint doesn't complain
)

REAL_AIR_FILE = (
    Path(__file__).resolve().parents[4]
    / "资料"
    / "2026.04.21"
    / "RE_ 今後の進め方に関するご提案"
    / "【Air】 Market Price updated on  Apr 20.xlsx"
)
REAL_OCEAN_FILE = (
    Path(__file__).resolve().parents[4]
    / "资料"
    / "2026.04.21"
    / "RE_ 今後の進め方に関するご提案"
    / "【Ocean】 Sea Net Rate_2026_Apr.21 - Apr.30.xlsx"
)
REAL_OCEAN_NGB_FILE = (
    Path(__file__).resolve().parents[4]
    / "资料"
    / "2026.04.21"
    / "RE_ 今後の進め方に関するご提案"
    / "【Ocean-NGB】 Ocean FCL rate sheet  HHENGB 2026 APR.xlsx"
)


@pytest.fixture(scope="module")
def real_batch() -> ParsedRateBatch:
    if not REAL_AIR_FILE.exists():
        pytest.skip(f"Air 真实样本不可用：{REAL_AIR_FILE}")
    return AirAdapter().parse(REAL_AIR_FILE)


@pytest.fixture(scope="module")
def weekly_records(real_batch: ParsedRateBatch):
    return [r for r in real_batch.records if r.record_kind == "air_weekly"]


@pytest.fixture(scope="module")
def surcharge_records(real_batch: ParsedRateBatch):
    return [r for r in real_batch.records if r.record_kind == "air_surcharge"]


@pytest.fixture(scope="module")
def current_week_weekly(weekly_records):
    return [r for r in weekly_records if r.extras.get("sheet_name") == "Apr 20 to Apr 26"]


# ------------- V-A01 -------------

def test_v_a01_adapter_metadata(real_batch: ParsedRateBatch) -> None:
    """V-A01：parse_air_file 返回 ParsedRateBatch，file_type=air，adapter_key=air。"""
    assert isinstance(real_batch, ParsedRateBatch)
    assert real_batch.file_type == Step1FileType.air
    assert real_batch.adapter_key == "air"
    assert real_batch.source_file == REAL_AIR_FILE.name
    assert real_batch.metadata.get("parser_version") == "air_v1"


# ------------- V-A02 -------------

def test_v_a02_detect_air_file_via_registry() -> None:
    """V-A02：adapter_key=='air'，且 AirAdapter.detect 对真实 Air 文件命中。"""
    adapter = AirAdapter()
    assert adapter.key == "air"
    assert adapter.file_type == Step1FileType.air
    assert adapter.priority == 10
    assert adapter.detect(REAL_AIR_FILE) is True


# ------------- V-A03 -------------

def test_v_a03_total_records(real_batch: ParsedRateBatch) -> None:
    """V-A03：记录总数 = 42 × 2 + 63 = 147。"""
    assert len(real_batch.records) == 147


# ------------- V-A04 -------------

def test_v_a04_record_kind_distribution(
    weekly_records, surcharge_records
) -> None:
    """V-A04：air_weekly=84，air_surcharge=63。"""
    assert len(weekly_records) == 84
    assert len(surcharge_records) == 63


# ------------- V-A05 -------------

def test_v_a05_batch_effective_range(real_batch: ParsedRateBatch) -> None:
    """V-A05：effective_from=2026-04-20，effective_to=2026-04-26（当前周）。"""
    assert real_batch.effective_from == date(2026, 4, 20)
    assert real_batch.effective_to == date(2026, 4, 26)


# ------------- V-A06 -------------

def test_v_a06_bkk_row_14(current_week_weekly) -> None:
    """V-A06：R14 BKK CK direct — 日价、remark、airports 全对得上。"""
    row = next(r for r in current_week_weekly if r.extras.get("row_index") == 14)
    assert row.destination_port_name == "BKK"
    assert row.price_day1 == Decimal("15.5")
    assert row.price_day2 == Decimal("15.5")
    assert row.price_day3 == Decimal("18")
    assert row.price_day4 == Decimal("18")
    assert row.price_day5 == Decimal("18")
    assert row.price_day6 == Decimal("18")
    assert row.price_day7 == Decimal("18")
    assert row.remarks == "CK D357"
    assert row.extras["destination_airports"] == ["BKK"]
    assert row.extras["density_hint"] is None


# ------------- V-A07 -------------

def test_v_a07_dxb_dwc_row_33(current_week_weekly) -> None:
    """V-A07：R33 DXB/DWC — 主字段保留斜杠原文，airports 拆成两个。"""
    row = next(r for r in current_week_weekly if r.extras.get("row_index") == 33)
    assert row.destination_port_name == "DXB/DWC"
    assert row.extras["destination_airports"] == ["DXB", "DWC"]
    assert row.extras["density_hint"] is None
    # price_day1..price_day3=42，price_day4..price_day7=43
    assert row.price_day1 == Decimal("42")
    assert row.price_day4 == Decimal("43")


# ------------- V-A08 -------------

def test_v_a08_fra_dense_row_40(current_week_weekly) -> None:
    """V-A08：R40 FRA DENSE — density_hint=DENSE，airports=[FRA]。"""
    row = next(r for r in current_week_weekly if r.extras.get("row_index") == 40)
    assert row.destination_port_name == "FRA DENSE"
    assert row.extras["density_hint"] == "DENSE"
    assert row.extras["destination_airports"] == ["FRA"]
    # FRA VOLUME 同时验证 strip（R41 原文末尾有空格）
    row41 = next(r for r in current_week_weekly if r.extras.get("row_index") == 41)
    assert row41.destination_port_name == "FRA VOLUME"
    assert row41.extras["density_hint"] == "VOLUME"
    assert row41.extras["destination_airports"] == ["FRA"]
    # raw_destination 应保留末尾空格
    assert row41.extras["raw_destination"] == "FRA VOLUME "


# ------------- V-A09 -------------

def test_v_a09_sin_must_go_22(current_week_weekly) -> None:
    """V-A09：R18 SIN Must go 22 — has_must_go=True，must_go_value=Decimal('22')。"""
    row = next(r for r in current_week_weekly if r.extras.get("row_index") == 18)
    assert row.destination_port_name == "SIN"
    assert row.remarks == "Must go 22"
    assert row.extras["has_must_go"] is True
    assert row.extras["must_go_value"] == Decimal("22")
    assert row.extras["is_case_by_case"] is False


# ------------- V-A10 -------------

def test_v_a10_case_by_case_row_27(current_week_weekly) -> None:
    """V-A10：R27 KTI Case by case — is_case_by_case=True，且服务列原文换行保留。"""
    row = next(r for r in current_week_weekly if r.extras.get("row_index") == 27)
    assert row.destination_port_name == "KTI"
    assert row.extras["is_case_by_case"] is True
    assert row.extras["has_must_go"] is False
    # raw_service 保留原始换行；service_desc 内部空白折叠但保留换行
    assert row.extras["raw_service"] == "BR/SQ 3-4 days\nservice"
    assert "\n" in row.service_desc


# ------------- V-A11 -------------

def test_v_a11_surcharges_aa_row_5(surcharge_records) -> None:
    """V-A11：Surcharges R5 AA American Airlines — airline_code / carrier_name / area / from / myc 字段精确对。"""
    row = next(r for r in surcharge_records if r.extras.get("row_index") == 5)
    assert row.carrier_name == "AA - American Airlines"
    assert row.airline_code == "AA"
    assert row.extras["area"] == "TC-3"
    assert row.extras["from_region"] == "CHINA / SHA"
    assert row.extras["myc_min_is_dash"] is True
    assert row.extras["myc_min_value"] is None
    assert row.extras["myc_fee_per_kg"] == Decimal("1.52")
    assert row.extras["myc_fee_is_dash"] is False
    assert row.valid_from == date(2026, 3, 13)
    assert row.extras["destination_scope"] == "TC-1,2"


# ------------- V-A12 -------------

def test_v_a12_surcharges_ba_row_6_all_dash(surcharge_records) -> None:
    """V-A12：R6 BA British Airways — 四费率字段全为 dash，all_fees_dash=True。"""
    row = next(r for r in surcharge_records if r.extras.get("row_index") == 6)
    assert row.carrier_name == "BA - British Airways"
    assert row.airline_code == "BA"
    assert row.extras["myc_min_is_dash"] is True
    assert row.extras["myc_fee_is_dash"] is True
    assert row.extras["msc_min_is_dash"] is True
    assert row.extras["msc_fee_is_dash"] is True
    assert row.extras["all_fees_dash"] is True
    assert row.extras["all_fees_empty"] is False
    # AREA 靠 forward-fill 填成 TC-3
    assert row.extras["area"] == "TC-3"


# ------------- V-A13 -------------

def test_v_a13_surcharges_y8_row_66_empty_vs_dash(surcharge_records) -> None:
    """V-A13：R66 Y8 TC-1 — MYC MIN/MSC Min 为真空（非 dash）。"""
    row = next(r for r in surcharge_records if r.extras.get("row_index") == 66)
    assert row.carrier_name == "Y8 - Yangtze River Express Airlines"
    assert row.airline_code == "Y8"
    assert row.extras["myc_min_is_dash"] is False
    assert row.extras["myc_min_value"] is None  # 真空
    assert row.extras["msc_min_is_dash"] is False
    assert row.extras["msc_min_value"] is None  # 真空
    # 对照同行 G、I 是数值
    assert row.extras["myc_fee_per_kg"] == Decimal("15")
    assert row.extras["msc_fee_per_kg"] == Decimal("1.2")
    assert row.extras["destination_scope"] == "TC-1"


# ------------- V-A14 -------------

def test_v_a14_surcharges_currency_all_cny(surcharge_records) -> None:
    """V-A14：Surcharges 全部记录 currency=='CNY'（F2 明示）。"""
    assert len(surcharge_records) > 0
    for row in surcharge_records:
        assert row.currency == "CNY", (
            f"row {row.extras.get('row_index')} currency={row.currency}"
        )
        assert row.extras.get("currency_source") == "F2"


# ------------- V-A15 -------------

def test_v_a15_weekly_currency_all_cny_with_assumption(weekly_records) -> None:
    """V-A15：周表全部 currency=='CNY'（来源：Surcharges F2 声明），currency_source='from_surcharges_F2'。"""
    assert len(weekly_records) > 0
    for row in weekly_records:
        assert row.currency == "CNY"
        assert row.extras.get("currency_source") == "from_surcharges_F2"
        assert row.origin_port_name == "PVG"
        assert row.extras.get("origin_source") == "default_air_PVG"


# ------------- V-A16 -------------

def test_v_a16_warnings_include_weekly_count_and_currency_dedup(
    real_batch: ParsedRateBatch,
) -> None:
    """V-A16：warnings 去重；命中 Surcharges F2 时不应再出现"未声明"警告；W-A02（≥2 张周表）唯一。"""
    warnings = real_batch.warnings
    # 去重后应为唯一
    assert len(warnings) == len(set(warnings))

    # 真实文件 Surcharges F2 = 'CURRENCY : CNY'，命中后不该打"未声明"或"fallback"
    currency_warnings = [
        w for w in warnings
        if "weekly sheet currency not declared" in w
        or "weekly currency falls back to CNY" in w
    ]
    assert currency_warnings == [], (
        f"命中 Surcharges F2 不应有 weekly currency 警告，实际 {currency_warnings}"
    )

    multi_week_warnings = [
        w for w in warnings if "workbook contains" in w and "weekly sheets" in w
    ]
    assert len(multi_week_warnings) == 1, f"W-A02 期望唯一，实际 {multi_week_warnings}"
    assert "Apr 20 - Apr 26" in multi_week_warnings[0] or "2026-04-20" in multi_week_warnings[0]


# ------------- V-A17 -------------

def test_v_a17_detect_matrix_three_adapters_exclusive() -> None:
    """V-A17：三份真实文件走 registry 分别命中各自 adapter，不串门。"""
    registry = RateAdapterRegistry([AirAdapter(), OceanAdapter(), OceanNgbAdapter()])
    assert REAL_AIR_FILE.exists()
    assert REAL_OCEAN_FILE.exists()
    assert REAL_OCEAN_NGB_FILE.exists()

    assert registry.resolve(REAL_AIR_FILE).key == "air"
    assert registry.resolve(REAL_OCEAN_FILE).key == "ocean"
    assert registry.resolve(REAL_OCEAN_NGB_FILE).key == "ocean_ngb"

    # 再直接查 AirAdapter.detect 对另外两份必须返回 False
    air = AirAdapter()
    assert air.detect(REAL_OCEAN_FILE) is False
    assert air.detect(REAL_OCEAN_NGB_FILE) is False


# ------------- V-A18 -------------

def test_v_a18_origin_all_pvg(real_batch: ParsedRateBatch) -> None:
    """V-A18：所有记录（周表+Surcharges）origin_port_name=='PVG'，origin_port_id is None。"""
    assert len(real_batch.records) == 147
    for row in real_batch.records:
        assert row.origin_port_name == "PVG", (
            f"row {row.record_kind}/{row.extras.get('row_index')} "
            f"origin_port_name={row.origin_port_name}"
        )
        assert row.origin_port_id is None


# ------------- V-A19 -------------

def test_v_a19_to_legacy_dict_json_serializable(real_batch: ParsedRateBatch) -> None:
    """V-A19：ParsedRateBatch.to_legacy_dict() 使用 default=str 可序列化为 JSON。"""
    payload = real_batch.to_legacy_dict()
    # total 字段
    assert payload["total_rows"] == 147
    assert payload["file_type"] == "air"
    assert payload["adapter_key"] == "air"
    # 与 Ocean 同款：Decimal 和 date 走 default=str（现有 ocean 测试/service 一致行为）
    dumped = json.dumps(payload, default=str, ensure_ascii=False)
    assert len(dumped) > 1000  # 至少有实质内容
    # 反解后行数一致
    reloaded = json.loads(dumped)
    assert len(reloaded["records"]) == 147


# ------------- V-A20 -------------

def test_v_a20_no_uncaught_exception_on_full_workbook() -> None:
    """V-A20：整本工作簿解析完成无未捕获 Exception；重跑一次确认。"""
    adapter = AirAdapter()
    batch = adapter.parse(REAL_AIR_FILE)  # 不在 fixture scope，独立再跑一次
    assert len(batch.records) == 147
    # 所有记录字段可访问（迭代一遍不炸）
    for row in batch.records:
        _ = (row.record_kind, row.currency, row.origin_port_name,
             dict(row.extras))


# ------------- V-A-CURR-01 -------------

def _build_air_xlsx(tmp_path: Path, *, surcharges_f2: str | None) -> Path:
    """构造一个最小可解析的 Air xlsx：1 张周表 + 可选 Surcharges sheet。"""
    from openpyxl import Workbook

    wb = Workbook()
    weekly = wb.active
    weekly.title = "Apr 20 to Apr 26"
    # 周表 A1='Destinations'，C1 含年份让 _extract_year_from_header 能命中
    weekly.cell(1, 1, "Destinations")
    weekly.cell(1, 2, "Service")
    weekly.cell(1, 3, "2026/04/20")
    # 1 行价：BKK / CK direct / 7 个价格 / 备注列
    weekly.cell(2, 1, "BKK")
    weekly.cell(2, 2, "CK direct")
    for col in range(3, 10):
        weekly.cell(2, col, 15.5)
    weekly.cell(2, 10, "")  # remark

    if surcharges_f2 is not None:
        sheet = wb.create_sheet("Surcharges")
        sheet["F2"] = surcharges_f2
        # 表头行 4：放最简表头（任务单不要求表头通过校验，只要能跑出 currency）
        for col, header in enumerate(
            ["AREA", "FROM", "AIRLINES", "Effective Date",
             "MYC MIN", "MYC FEE/KG", "MSC MIN", "MSC FEE/KG",
             "Destination", "Remarks"],
            start=2,
        ):
            sheet.cell(4, col, header)
        # 不放数据行，避免噪音

    out = tmp_path / "air_test.xlsx"
    wb.save(out)
    return out


def test_v_a_curr_01_weekly_inherits_surcharges_usd(tmp_path: Path) -> None:
    """V-A-CURR-01：Surcharges F2='CURRENCY : USD' → weekly records currency=='USD'，无未声明警告。"""
    path = _build_air_xlsx(tmp_path, surcharges_f2="CURRENCY : USD")
    batch = AirAdapter().parse(path)
    weekly = [r for r in batch.records if r.record_kind == "air_weekly"]
    assert len(weekly) >= 1
    for row in weekly:
        assert row.currency == "USD", (
            f"row {row.extras.get('row_index')} currency={row.currency}"
        )
        assert row.extras.get("currency_source") == "from_surcharges_F2"

    bad = [
        w for w in batch.warnings
        if "weekly sheet currency not declared" in w
        or "weekly currency falls back to CNY" in w
    ]
    assert bad == [], f"命中 Surcharges F2 不应有 weekly currency 警告，实际 {bad}"


# ------------- V-A-CURR-02 -------------

def test_v_a_curr_02_weekly_falls_back_when_no_surcharges(tmp_path: Path) -> None:
    """V-A-CURR-02：无 Surcharges sheet → weekly currency=='CNY'，warnings 含 fallback 提示。"""
    path = _build_air_xlsx(tmp_path, surcharges_f2=None)
    batch = AirAdapter().parse(path)
    weekly = [r for r in batch.records if r.record_kind == "air_weekly"]
    assert len(weekly) >= 1
    for row in weekly:
        assert row.currency == "CNY"
        assert row.extras.get("currency_source") == "fallback_no_surcharges"

    fallback_warnings = [
        w for w in batch.warnings
        if "weekly currency falls back to CNY" in w
    ]
    assert len(fallback_warnings) == 1, (
        f"应含 1 条 fallback 警告，实际 {fallback_warnings}"
    )
    assert "Surcharges sheet missing" in fallback_warnings[0]
