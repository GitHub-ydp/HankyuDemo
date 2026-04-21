from __future__ import annotations

from decimal import Decimal

from openpyxl import Workbook

from app.services.step1_rates.adapters.ocean import OceanAdapter


def _build_worksheet(data_rows: list[list[object]]):
    """Build an in-memory FCL worksheet.

    Layout of data rows (columns A..D matter for pairing):
      A: destination (To)
      B: carrier (Shipping Line)
      C: container type (20FT / 40FT / 40HQ) — hard-coded read at row[2]
      D: freight (numeric)
    """
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "FCL TEST"

    # Row 1: origin marker — `_read_origin_name_from_row` detects "From:" prefix
    worksheet.cell(row=1, column=1, value="From: SHANGHAI")

    # Row 2: header — `_build_fcl_column_map` matches normalized text
    worksheet.cell(row=2, column=1, value="To")
    worksheet.cell(row=2, column=2, value="Shipping Line")
    worksheet.cell(row=2, column=3, value="Container")
    worksheet.cell(row=2, column=4, value="Freight")

    # Row 3..: data rows
    for offset, values in enumerate(data_rows):
        row_num = 3 + offset
        for col_offset, value in enumerate(values):
            worksheet.cell(row=row_num, column=1 + col_offset, value=value)

    return worksheet


def test_pair_20ft_and_40hq_two_rows() -> None:
    adapter = OceanAdapter()
    worksheet = _build_worksheet(
        [
            ["LAX", "MSC", "20FT", 850],
            ["LAX", "MSC", "40HQ", 1650],
        ]
    )

    records, meta, warnings = adapter._parse_fcl_sheet(
        worksheet,
        db=None,
        surcharge_mode="other",
        source_file="test.xlsx",
    )

    assert len(records) == 1, f"expected 1 paired record, got {len(records)}; warnings={warnings}"
    record = records[0]
    assert record.record_kind == "fcl"
    assert record.carrier_name == "MSC"
    assert record.container_20gp == Decimal("850")
    # 40HQ writes container_40hq, then _append_pending_record mirrors 40hq -> 40gp
    assert record.container_40gp == Decimal("1650")
    assert record.container_40hq == Decimal("1650")


def test_pair_20ft_and_40hq_and_40ft_three_rows() -> None:
    adapter = OceanAdapter()
    worksheet = _build_worksheet(
        [
            ["LAX", "MSC", "20FT", 850],
            ["LAX", "MSC", "40HQ", 1650],
            ["LAX", "MSC", "40FT", 1600],
        ]
    )

    records, meta, warnings = adapter._parse_fcl_sheet(
        worksheet,
        db=None,
        surcharge_mode="other",
        source_file="test.xlsx",
    )

    assert len(records) == 1, f"expected 1 paired record, got {len(records)}; warnings={warnings}"
    record = records[0]
    assert record.record_kind == "fcl"
    assert record.carrier_name == "MSC"
    assert record.container_20gp == Decimal("850")
    # 40FT writes container_40gp=1600 (overwriting the 40hq->40gp mirror value of 1650)
    assert record.container_40gp == Decimal("1600")
    # 40FT must not clear container_40hq previously written by 40HQ row
    assert record.container_40hq == Decimal("1650")


def test_orphan_40hq_without_preceding_20ft() -> None:
    adapter = OceanAdapter()
    worksheet = _build_worksheet(
        [
            ["LAX", "MSC", "40HQ", 2000],
        ]
    )

    records, meta, warnings = adapter._parse_fcl_sheet(
        worksheet,
        db=None,
        surcharge_mode="other",
        source_file="test.xlsx",
    )

    assert len(records) == 1, f"expected 1 orphan record, got {len(records)}; warnings={warnings}"
    record = records[0]
    assert record.record_kind == "fcl"
    assert record.carrier_name == "MSC"
    assert record.container_20gp is None
    # Orphan path mirrors 40hq -> 40gp
    assert record.container_40gp == Decimal("2000")
    assert record.container_40hq == Decimal("2000")
    assert any("stored orphan 40FT/40HQ row" in w for w in warnings), (
        f"expected orphan warning, got: {warnings}"
    )
