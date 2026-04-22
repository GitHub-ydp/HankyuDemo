"""T-W2 验收：AirWriter round-trip。"""
from __future__ import annotations

from datetime import datetime
from io import BytesIO

from openpyxl import load_workbook

from app.services.step1_rates.writers.air import AirWriter
from app.services.step1_rates.writers.templates import resolve_template_path


def _load_writer_output(batch_id: str):
    content, filename = AirWriter().write(batch_id)
    return load_workbook(BytesIO(content), data_only=False), filename


def test_air_filename_follows_template(air_batch_id):
    _, filename = _load_writer_output(air_batch_id)
    assert filename.startswith("【Air】 Market Price updated on ")
    assert filename.endswith(".xlsx")


def test_air_weekly_destination_and_service_preserved(air_batch_id):
    wb, _ = _load_writer_output(air_batch_id)
    ws = wb["Apr 20 to Apr 26"]
    # V-W07: BKK row 14 destination + service
    assert ws.cell(14, 1).value == "BKK"
    assert ws.cell(14, 2).value == "CK direct 2 days"
    # V-W07: KUL row 22 preserves "servcie " typo and trailing space
    assert ws.cell(22, 2).value == "MH/FM 3 days servcie "
    # V-W08: KTI row 27 preserves embedded newline
    assert ws.cell(27, 2).value == "BR/SQ 3-4 days\nservice"


def test_air_weekly_prices_unchanged(air_batch_id):
    wb, _ = _load_writer_output(air_batch_id)
    ws = wb["Apr 20 to Apr 26"]
    assert ws.cell(14, 3).value == 15.5
    assert ws.cell(14, 5).value == 18
    assert ws.cell(40, 3).value == 38
    assert ws.cell(40, 5).value == 41


def test_air_surcharges_area_and_from_unchanged(air_batch_id):
    """V-W09: AREA / FROM 合并 anchor cell 保留原值，合并区仍存在。"""
    wb, _ = _load_writer_output(air_batch_id)
    ws = wb["Surcharges"]
    assert ws.cell(5, 2).value == "TC-3"
    assert "CHINA / SHA" in (ws.cell(5, 3).value or "")
    merged_set = {str(r) for r in ws.merged_cells.ranges}
    assert "C5:C67" in merged_set


def test_air_surcharges_airline_and_dates(air_batch_id):
    wb, _ = _load_writer_output(air_batch_id)
    ws = wb["Surcharges"]
    assert ws.cell(5, 4).value == "AA - American Airlines"
    assert isinstance(ws.cell(5, 5).value, datetime)
    assert ws.cell(5, 5).value == datetime(2026, 3, 13)
    # dash cells
    assert ws.cell(5, 6).value == "-"
    assert ws.cell(5, 9).value == "-"
    assert ws.cell(5, 7).value == 1.52


def test_air_surcharges_all_dash_row(air_batch_id):
    """V-W12: BA row 6 所有 4 项费率都 '-'。"""
    wb, _ = _load_writer_output(air_batch_id)
    ws = wb["Surcharges"]
    for col in (6, 7, 8, 9):
        assert ws.cell(6, col).value == "-", f"row 6 col {col} expected '-'"


def test_air_roundtrip_preserves_merged_and_columns(air_batch_id):
    template_path = resolve_template_path(air_batch_id)
    original = load_workbook(template_path, data_only=False)
    wb, _ = _load_writer_output(air_batch_id)
    # V-W03 / V-W05: merged_cells / column_dimensions set invariant
    for sheet_name in original.sheetnames:
        orig_ws = original[sheet_name]
        new_ws = wb[sheet_name]
        orig_merged = {str(r) for r in orig_ws.merged_cells.ranges}
        new_merged = {str(r) for r in new_ws.merged_cells.ranges}
        assert orig_merged == new_merged, f"merged_cells differ in {sheet_name}"

        orig_cols = {
            key: (dim.width, dim.hidden)
            for key, dim in orig_ws.column_dimensions.items()
        }
        new_cols = {
            key: (dim.width, dim.hidden)
            for key, dim in new_ws.column_dimensions.items()
        }
        assert orig_cols == new_cols, f"column_dimensions differ in {sheet_name}"
