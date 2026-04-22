"""T-W1 验收：base / naming 单元测试。"""
from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal

from openpyxl import Workbook

from app.services.step1_rates.entities import Step1FileType
from app.services.step1_rates.writers.base import (
    is_formula_cell,
    pick_raw,
    safe_set,
    stamp_document_properties,
)
from app.services.step1_rates.writers.naming import build_filename


def _cell():
    wb = Workbook()
    ws = wb.active
    return ws.cell(row=1, column=1)


def test_is_formula_cell_detects_equals_prefix():
    cell = _cell()
    cell.value = "=SUM(A2:A5)"
    assert is_formula_cell(cell) is True


def test_is_formula_cell_false_for_plain_text():
    cell = _cell()
    cell.value = "AT COST"
    assert is_formula_cell(cell) is False


def test_safe_set_skips_formula_cell():
    cell = _cell()
    cell.value = "=A2"
    assert safe_set(cell, 999.0) is False
    assert cell.value == "=A2"


def test_safe_set_allows_explicit_overwrite_formula():
    cell = _cell()
    cell.value = "=A2"
    assert safe_set(cell, 12.0, allow_overwrite_formula=True) is True
    assert cell.value == 12.0


def test_safe_set_skips_none_value():
    cell = _cell()
    cell.value = "original"
    assert safe_set(cell, None) is False
    assert cell.value == "original"


def test_safe_set_coerces_decimal_to_float():
    cell = _cell()
    assert safe_set(cell, Decimal("3.50")) is True
    assert cell.value == 3.5


def test_safe_set_writes_strings_and_dates():
    wb = Workbook()
    ws = wb.active
    c1 = ws.cell(row=1, column=1)
    c2 = ws.cell(row=1, column=2)
    assert safe_set(c1, "hello\nworld") is True
    assert c1.value == "hello\nworld"
    d = datetime(2026, 4, 20)
    assert safe_set(c2, d) is True
    assert c2.value == d


def test_pick_raw_first_non_none_non_empty():
    record = {"raw_a": None, "raw_b": "", "raw_c": "keep", "raw_d": "fallback"}
    assert pick_raw(record, "raw_a", "raw_b", "raw_c", "raw_d") == "keep"


def test_pick_raw_default_when_all_missing():
    record = {"raw_a": None}
    assert pick_raw(record, "raw_a", "raw_x", default="def") == "def"


def test_build_filename_air():
    name = build_filename(Step1FileType.air, date(2026, 4, 20), date(2026, 4, 26))
    assert name == "【Air】 Market Price updated on Apr 20.xlsx"


def test_build_filename_ocean_includes_year_and_range():
    name = build_filename(Step1FileType.ocean, date(2026, 4, 21), date(2026, 4, 30))
    assert name == "【Ocean】 Sea Net Rate_2026_Apr.21 - Apr.30.xlsx"


def test_build_filename_ocean_ngb_uppercase_month():
    name = build_filename(Step1FileType.ocean_ngb, date(2026, 4, 1), date(2026, 4, 30))
    assert name == "【Ocean-NGB】 Ocean FCL rate sheet  HHENGB 2026 APR.xlsx"


def test_build_filename_appends_hhmmss_when_now_set():
    now = datetime(2026, 4, 22, 14, 30, 5)
    name = build_filename(
        Step1FileType.air, date(2026, 4, 20), date(2026, 4, 26), now=now
    )
    assert name.endswith("_143005.xlsx")


def test_stamp_document_properties_records_batch_and_writer_version():
    wb = Workbook()
    stamp_document_properties(wb, batch_id="abc-123", exported_at=datetime(2026, 4, 22))
    assert "abc-123" in (wb.properties.title or "")
    assert wb.properties.subject.startswith("exported 2026-04-22")
    assert "step1-writer" in (wb.properties.description or "")
