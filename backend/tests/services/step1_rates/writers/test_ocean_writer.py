"""T-W3 验收：OceanWriter round-trip（RW1 兜底：数值 + 已保留 raw 文本）。"""
from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook

from app.services.step1_rates.writers.ocean import OceanWriter
from app.services.step1_rates.writers.templates import resolve_template_path


def _load_writer_output(batch_id: str):
    content, filename = OceanWriter().write(batch_id)
    return load_workbook(BytesIO(content), data_only=False), filename


def test_ocean_filename_follows_template(ocean_batch_id):
    _, filename = _load_writer_output(ocean_batch_id)
    assert filename.startswith("【Ocean】 Sea Net Rate_")
    assert filename.endswith(".xlsx")


def test_ocean_lcl_freight_original_text_preserved(ocean_batch_id):
    wb, _ = _load_writer_output(ocean_batch_id)
    ws = wb["LCL N RATE"]
    # parser 保留了 freight_raw，writer 用它覆盖；值应与模板一致
    assert ws.cell(10, 2).value == "5/CBM, 10/TON"
    assert ws.cell(11, 2).value == "26/CBM, 29/TON"


def test_ocean_lcl_chinese_remark_preserved(ocean_batch_id):
    """V-W13: LCL row 27 col 12 中文注释原样。"""
    wb, _ = _load_writer_output(ocean_batch_id)
    ws = wb["LCL N RATE"]
    assert ws.cell(27, 12).value == "关封货物只接受新加坡转拼的服务，运费另询"


def test_ocean_lcl_comments_preserved(ocean_batch_id):
    """V-W14: LCL L10 & B26 的 Zhang Jieyi 批注保留。"""
    wb, _ = _load_writer_output(ocean_batch_id)
    ws = wb["LCL N RATE"]
    l10 = ws["L10"].comment
    b26 = ws["B26"].comment
    assert l10 is not None, "L10 comment missing"
    assert b26 is not None, "B26 comment missing"
    assert l10.author == "Zhang Jieyi"
    assert b26.author == "Zhang Jieyi"
    assert "THE TERMINAL DOC CHARGE USD50/BL" in b26.text


def test_ocean_merged_cells_invariant(ocean_batch_id):
    """V-W12: JP / 其他港 / LCL 合并区整体保留。"""
    template_path = resolve_template_path(ocean_batch_id)
    original = load_workbook(template_path, data_only=False)
    wb, _ = _load_writer_output(ocean_batch_id)
    for sheet_name in original.sheetnames:
        orig_merged = {str(r) for r in original[sheet_name].merged_cells.ranges}
        new_merged = {str(r) for r in wb[sheet_name].merged_cells.ranges}
        assert orig_merged == new_merged, f"merged differ in {sheet_name}"


def test_ocean_column_dimensions_invariant(ocean_batch_id):
    template_path = resolve_template_path(ocean_batch_id)
    original = load_workbook(template_path, data_only=False)
    wb, _ = _load_writer_output(ocean_batch_id)
    for sheet_name in original.sheetnames:
        orig = {
            k: (d.width, d.hidden)
            for k, d in original[sheet_name].column_dimensions.items()
        }
        new = {
            k: (d.width, d.hidden)
            for k, d in wb[sheet_name].column_dimensions.items()
        }
        assert orig == new, f"column_dimensions differ in {sheet_name}"
