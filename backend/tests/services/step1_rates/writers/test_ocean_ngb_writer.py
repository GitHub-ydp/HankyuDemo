"""T-W4 验收：OceanNgbWriter 公式纪律。"""
from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook

from app.services.step1_rates.writers.ocean_ngb import OceanNgbWriter
from app.services.step1_rates.writers.templates import resolve_template_path


EXPECTED_FORMULA_COUNT = 1687


def _load_writer_output(batch_id: str):
    content, filename = OceanNgbWriter().write(batch_id)
    return load_workbook(BytesIO(content), data_only=False), filename


def _count_formulas(ws) -> int:
    count = 0
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.startswith("="):
                count += 1
    return count


def test_ngb_filename_follows_template(ocean_ngb_batch_id):
    _, filename = _load_writer_output(ocean_ngb_batch_id)
    assert filename.startswith("【Ocean-NGB】 Ocean FCL rate sheet  HHENGB ")
    assert filename.endswith(".xlsx")


def test_ngb_rate_sheet_preserves_1687_formulas(ocean_ngb_batch_id):
    """V-W15: Rate sheet 公式数量完好（parser 当前 stub 0 records → 全保留）。"""
    wb, _ = _load_writer_output(ocean_ngb_batch_id)
    assert _count_formulas(wb["Rate"]) == EXPECTED_FORMULA_COUNT


def test_ngb_specific_formulas_still_exist(ocean_ngb_batch_id):
    """V-W16: 抽样 Lv.2/Lv.3 典型公式保留。"""
    template_path = resolve_template_path(ocean_ngb_batch_id)
    original = load_workbook(template_path, data_only=False)
    wb, _ = _load_writer_output(ocean_ngb_batch_id)
    orig_rate = original["Rate"]
    new_rate = wb["Rate"]
    sampled = 0
    for row in orig_rate.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                new_val = new_rate[cell.coordinate].value
                assert new_val == cell.value, (
                    f"formula at {cell.coordinate} changed: {cell.value!r} → {new_val!r}"
                )
                sampled += 1
                if sampled >= 50:
                    return
    assert sampled > 0, "did not sample any formulas — unexpected"


def test_ngb_sample_and_shipping_line_sheets_untouched(ocean_ngb_batch_id):
    """V-W17: sample / Shipping line name 两个 sheet 的所有 cell 与原件一致。"""
    template_path = resolve_template_path(ocean_ngb_batch_id)
    original = load_workbook(template_path, data_only=False)
    wb, _ = _load_writer_output(ocean_ngb_batch_id)
    for sheet_name in ("sample", "Shipping line name"):
        orig_ws = original[sheet_name]
        new_ws = wb[sheet_name]
        assert orig_ws.max_row == new_ws.max_row
        assert orig_ws.max_column == new_ws.max_column
        for row_idx in range(1, orig_ws.max_row + 1):
            for col_idx in range(1, orig_ws.max_column + 1):
                orig_v = orig_ws.cell(row_idx, col_idx).value
                new_v = new_ws.cell(row_idx, col_idx).value
                assert orig_v == new_v, (
                    f"{sheet_name}!{orig_ws.cell(row_idx, col_idx).coordinate} "
                    f"changed: {orig_v!r} → {new_v!r}"
                )


def test_ngb_merged_cells_and_columns_invariant(ocean_ngb_batch_id):
    template_path = resolve_template_path(ocean_ngb_batch_id)
    original = load_workbook(template_path, data_only=False)
    wb, _ = _load_writer_output(ocean_ngb_batch_id)
    for sheet_name in original.sheetnames:
        orig_merged = {str(r) for r in original[sheet_name].merged_cells.ranges}
        new_merged = {str(r) for r in wb[sheet_name].merged_cells.ranges}
        assert orig_merged == new_merged, f"merged differ in {sheet_name}"

        orig_cols = {
            k: (d.width, d.hidden)
            for k, d in original[sheet_name].column_dimensions.items()
        }
        new_cols = {
            k: (d.width, d.hidden)
            for k, d in wb[sheet_name].column_dimensions.items()
        }
        assert orig_cols == new_cols, f"column_dimensions differ in {sheet_name}"

    # freeze_panes 保留（V-W05）
    assert original["Rate"].freeze_panes == wb["Rate"].freeze_panes


def test_ngb_document_properties_stamped(ocean_ngb_batch_id):
    wb, _ = _load_writer_output(ocean_ngb_batch_id)
    assert ocean_ngb_batch_id in (wb.properties.title or "")
    assert (wb.properties.description or "").startswith("step1-writer")
